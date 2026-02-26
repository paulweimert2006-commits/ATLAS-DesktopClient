"""
VU-Listen-Parser fuer Provisionsdaten.

Parst Excel-Sheets von Allianz, SwissLife und VB.
Extrahiert aus services/provision_import.py in den Domain Layer.
"""

import re
import hashlib
import logging
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass, field
from datetime import datetime

logger = logging.getLogger(__name__)

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl nicht installiert -- Excel-Import nicht verfuegbar")

from .normalization import normalize_vb_name
from .relevance import is_commission_relevant


def compute_row_hash(vu_name: str, vsnr: str, betrag: float,
                     datum: str, art: str = '') -> str:
    """Eindeutiger Hash pro Provisions-Zeile fuer Duplikat-Schutz."""
    raw = f"{vu_name}|{vsnr}|{betrag:.2f}|{datum or ''}|{art}"
    return hashlib.sha256(raw.encode('utf-8')).hexdigest()


def parse_amount(val) -> Optional[float]:
    """Betrag aus Excel-Zelle parsen (deutsch/englisch)."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s == '-':
        return None
    s = s.replace(' ', '').replace('€', '').replace('EUR', '')
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def parse_date(val) -> Optional[str]:
    """Datum aus Excel-Zelle parsen -> YYYY-MM-DD."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s:
        return None
    for fmt in ('%d.%m.%Y', '%d.%m.%y', '%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


def col_index(letter: str) -> int:
    """Spaltenbuchstabe -> 1-basierter Index (A=1, B=2, ..., AG=33)."""
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result


@dataclass
class ParseResult:
    """Ergebnis eines Sheet-Parsings."""
    rows: List[Dict] = field(default_factory=list)
    sheet_name: str = ''
    vu_name: str = ''
    total_rows: int = 0
    skipped_rows: int = 0
    errors: List[str] = field(default_factory=list)


VU_COLUMN_MAPPINGS = {
    'Allianz': {
        'vsnr_col': 'A',
        'betrag_col': 'D',
        'art_col': 'F',
        'datum_col': 'G',
        'courtage_rate_col': 'K',
        'vn_col': 'AE',
        'konditionssatz_col': None,
    },
    'SwissLife': {
        'vsnr_col': 'Y',
        'betrag_col': 'N',
        'art_col': 'O',
        'datum_col': 'C',
        'courtage_rate_col': None,
        'vn_col': 'U',
        'konditionssatz_col': None,
    },
    'VB': {
        'vsnr_col': 'B',
        'betrag_col': 'O',
        'lastschrift_col': 'P',
        'art_col': 'K',
        'datum_col': 'AR',
        'courtage_rate_col': None,
        'vn_col': 'C',
        'konditionssatz_col': 'M',
    },
}

VU_HEADER_SIGNATURES = {
    'Allianz': {
        'sheet_names': ['allianz'],
        'header_keywords': ['vtnr', 'provisions-betrag', 'courtagesatz',
                            'auszahlungs-datum'],
    },
    'SwissLife': {
        'sheet_names': ['swisslife', 'swiss life'],
        'header_keywords': ['versicherungsnummer', 'buchwert',
                            'abrechnungsnummer', 'konditionssatz'],
    },
    'VB': {
        'sheet_names': ['vb', 'volkswohlbund'],
        'header_keywords': ['vart', 'gutschrift', 'lastschrift',
                            'abrechnung von'],
    },
}


def map_art(raw: str, vu_name: str) -> str:
    """Provisions-Art aus VU-spezifischem Wert in Enum-Wert mappen."""
    s = str(raw or '').strip().upper()
    if not s:
        return 'ap'
    if s in ('RB', 'ST', 'STORNO', 'RÜCK', 'RUECK', 'RÜCKBELASTUNG'):
        return 'rueckbelastung'
    if s in ('BP', 'FP', 'FOLGEPROV', 'BESTANDSPROV', 'BEST'):
        return 'bp'
    if s in ('AP', 'EV', 'EV-PF', 'ABSCHL', 'ABSCHLUSSPROV'):
        return 'ap'
    return 'sonstige'


def parse_vu_sheet(wb, sheet_name: str) -> ParseResult:
    """Ein einzelnes VU-Sheet parsen (iter_rows fuer Performance)."""
    result = ParseResult(sheet_name=sheet_name, vu_name=sheet_name)

    if sheet_name not in wb.sheetnames:
        result.errors.append(f"Sheet '{sheet_name}' nicht gefunden")
        return result

    mapping = VU_COLUMN_MAPPINGS.get(sheet_name)
    if not mapping:
        result.errors.append(f"Kein Column-Mapping fuer '{sheet_name}'")
        return result

    ws = wb[sheet_name]

    vsnr_idx = col_index(mapping['vsnr_col'])
    betrag_idx = col_index(mapping['betrag_col']) if mapping['betrag_col'] else None
    lastschrift_idx = (col_index(mapping['lastschrift_col'])
                       if mapping.get('lastschrift_col') else None)
    art_idx = col_index(mapping['art_col']) if mapping['art_col'] else None
    datum_idx = col_index(mapping['datum_col']) if mapping['datum_col'] else None
    courtage_rate_idx = (col_index(mapping['courtage_rate_col'])
                         if mapping.get('courtage_rate_col') else None)
    vn_idx = col_index(mapping['vn_col']) if mapping.get('vn_col') else None
    konditionssatz_idx = (col_index(mapping['konditionssatz_col'])
                          if mapping.get('konditionssatz_col') else None)

    if betrag_idx is None:
        result.errors.append(f"{sheet_name}: Betrag-Spalte nicht konfiguriert")
        return result

    max_col_needed = max(filter(None, [
        vsnr_idx, betrag_idx, lastschrift_idx, art_idx, datum_idx,
        courtage_rate_idx, vn_idx, konditionssatz_idx,
    ]))

    logger.info(f"Parsing {sheet_name}: vsnr={vsnr_idx}, betrag={betrag_idx}, "
                f"art={art_idx}, datum={datum_idx}")

    row_num = 0
    for row in ws.iter_rows(min_row=2, max_col=max_col_needed, values_only=False):
        row_num += 1
        result.total_rows += 1
        try:
            vsnr_cell = row[vsnr_idx - 1] if vsnr_idx - 1 < len(row) else None
            vsnr_raw = vsnr_cell.value if vsnr_cell else None
            if vsnr_raw is None or str(vsnr_raw).strip() == '':
                result.skipped_rows += 1
                continue

            vsnr = str(vsnr_raw).strip()

            betrag_val = (row[betrag_idx - 1].value
                          if betrag_idx and betrag_idx - 1 < len(row) else None)
            betrag = parse_amount(betrag_val)
            if (betrag is None or betrag == 0) and lastschrift_idx:
                ls_val = (row[lastschrift_idx - 1].value
                          if lastschrift_idx - 1 < len(row) else None)
                ls_betrag = parse_amount(ls_val)
                if ls_betrag is not None and ls_betrag != 0:
                    betrag = -abs(ls_betrag)
            if betrag is None or betrag == 0:
                result.skipped_rows += 1
                continue

            art_val = (row[art_idx - 1].value
                       if art_idx and art_idx - 1 < len(row) else None)
            art_raw = str(art_val or '')
            art = map_art(art_raw, sheet_name)

            datum_val = (row[datum_idx - 1].value
                         if datum_idx and datum_idx - 1 < len(row) else None)
            datum = parse_date(datum_val)

            vn_val = (row[vn_idx - 1].value
                      if vn_idx and vn_idx - 1 < len(row) else None)
            vn_name = str(vn_val or '').strip() if vn_val else None
            if vn_name and sheet_name == 'VB':
                vn_name = normalize_vb_name(vn_name)

            courtage_rate = None
            if courtage_rate_idx and courtage_rate_idx - 1 < len(row):
                rate_val = row[courtage_rate_idx - 1].value
                courtage_rate = parse_amount(rate_val)

            konditionssatz = None
            if konditionssatz_idx and konditionssatz_idx - 1 < len(row):
                kond_val = row[konditionssatz_idx - 1].value
                konditionssatz = (str(kond_val).strip()
                                  if kond_val is not None else None)

            if betrag < 0:
                art = 'rueckbelastung'

            row_hash = compute_row_hash(sheet_name, vsnr, betrag, datum, art)

            relevant = is_commission_relevant(
                vu_name=sheet_name,
                courtage_rate=courtage_rate,
                buchungsart_raw=art_raw,
                konditionssatz=konditionssatz,
            )

            result.rows.append({
                'vsnr': vsnr,
                'betrag': round(betrag, 2),
                'art': art,
                'buchungsart_raw': art_raw,
                'auszahlungsdatum': datum,
                'versicherungsnehmer': vn_name,
                'provisions_basissumme': None,
                'rate_nummer': None,
                'rate_anzahl': None,
                'row_hash': row_hash,
                'courtage_rate': (round(courtage_rate, 2)
                                  if courtage_rate is not None else None),
                'konditionssatz': konditionssatz,
                'source_row': row_num + 1,
                'is_relevant': relevant,
            })
        except Exception as e:
            result.errors.append(f"Zeile {row_num + 1}: {str(e)}")
            result.skipped_rows += 1

    logger.info(f"VU-Sheet '{sheet_name}': {len(result.rows)} Zeilen geparst, "
                f"{result.skipped_rows} uebersprungen, {len(result.errors)} Fehler")
    return result


def parse_vu_liste(filepath: str, selected_sheets: List[str] = None,
                   on_progress=None) -> List[ParseResult]:
    """Alle relevanten VU-Sheets aus einer Excel-Datei parsen."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl nicht installiert")

    import time
    t0 = time.time()

    def _log(msg):
        logger.info(msg)
        if on_progress:
            on_progress(msg)

    _log(f"Oeffne Excel: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    _log(f"Geoeffnet in {time.time() - t0:.1f}s, Sheets: {wb.sheetnames}")

    sheets_to_parse = selected_sheets or list(VU_COLUMN_MAPPINGS.keys())
    results = []

    for sheet_name in sheets_to_parse:
        if sheet_name in wb.sheetnames:
            t1 = time.time()
            _log(f"Parse Sheet '{sheet_name}'...")
            pr = parse_vu_sheet(wb, sheet_name)
            _log(f"Sheet '{sheet_name}' fertig in {time.time() - t1:.1f}s: "
                 f"{len(pr.rows)} Zeilen, {pr.skipped_rows} skip, "
                 f"{len(pr.errors)} err")
            results.append(pr)

    wb.close()
    _log(f"Gesamt: {time.time() - t0:.1f}s")
    return results


def get_available_vu_sheets(filepath: str) -> List[str]:
    """Verfuegbare VU-Sheets in einer Excel-Datei ermitteln."""
    if not HAS_OPENPYXL:
        return []
    wb = openpyxl.load_workbook(filepath, read_only=True)
    available = [s for s in wb.sheetnames if s in VU_COLUMN_MAPPINGS]
    wb.close()
    return available


def detect_vu_format(filepath: str) -> List[Tuple[str, float]]:
    """VU-Format erkennen anhand Sheet-Namen und Header.

    Returns: Liste von (vu_name, confidence) Tupeln, absteigend.
    """
    if not HAS_OPENPYXL:
        return []
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    results = []

    sheets_lower = {s.lower(): s for s in wb.sheetnames}

    for vu_name, sig in VU_HEADER_SIGNATURES.items():
        if vu_name in wb.sheetnames:
            results.append((vu_name, 1.0))
            continue
        for sn_pattern in sig.get('sheet_names', []):
            if sn_pattern in sheets_lower:
                results.append((vu_name, 0.9))
                break

    if not results:
        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                headers = []
                for row in ws.iter_rows(min_row=1, max_row=1,
                                        values_only=True):
                    headers = [str(v or '').strip().lower() for v in row]
                    break
                header_str = ' '.join(headers)
                for vu_name, sig in VU_HEADER_SIGNATURES.items():
                    kws = sig.get('header_keywords', [])
                    matches = sum(1 for kw in kws if kw in header_str)
                    if matches >= 2:
                        results.append((vu_name, 0.6 + 0.1 * matches))
                        break
            except Exception:
                continue

    wb.close()
    results.sort(key=lambda x: x[1], reverse=True)
    seen = set()
    unique = []
    for vu, conf in results:
        if vu not in seen:
            seen.add(vu)
            unique.append((vu, conf))
    return unique


def compute_file_hash(filepath: str) -> str:
    """SHA256-Hash einer Datei berechnen."""
    h = hashlib.sha256()
    with open(filepath, 'rb') as f:
        for chunk in iter(lambda: f.read(65536), b''):
            h.update(chunk)
    return h.hexdigest()

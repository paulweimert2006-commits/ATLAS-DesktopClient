"""
Dokumenten-Verarbeitungs-Service

Automatische Klassifikation und Verschiebung von Dokumenten in Boxen.
Unterstuetzt parallele Verarbeitung fuer bessere Performance.
"""

import logging
from typing import List, Optional, Tuple, Callable
from dataclasses import dataclass
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
import threading
import tempfile
import os

from api.documents import Document, DocumentsAPI, BOX_TYPES
from api.openrouter import OpenRouterClient, ExtractedDocumentData, DocumentClassification
from api.client import APIClient
from api.processing_history import ProcessingHistoryAPI
from api.processing_settings import ProcessingSettingsAPI
from api.document_rules import DocumentRulesAPI, DocumentRulesSettings

logger = logging.getLogger(__name__)

# Parallele Verarbeitung
DEFAULT_MAX_WORKERS = 8  # Anzahl gleichzeitiger Verarbeitungen


@dataclass
class ProcessingResult:
    """Ergebnis der Verarbeitung eines Dokuments."""
    document_id: int
    original_filename: str
    success: bool
    target_box: str
    category: Optional[str] = None
    new_filename: Optional[str] = None
    error: Optional[str] = None
    cost_usd: float = 0.0


@dataclass
class BatchProcessingResult:
    """Ergebnis eines kompletten Verarbeitungslaufs mit Kosten-Tracking."""
    results: List[ProcessingResult]
    total_documents: int
    successful_documents: int
    failed_documents: int
    duration_seconds: float
    # Kosten-Tracking
    credits_before: Optional[float] = None
    credits_after: Optional[float] = None
    total_cost_usd: Optional[float] = None
    cost_per_document_usd: Optional[float] = None
    currency: str = 'USD'
    provider: str = 'openrouter'
    
    @property
    def success_rate(self) -> float:
        """Erfolgsrate in Prozent."""
        if self.total_documents == 0:
            return 0.0
        return (self.successful_documents / self.total_documents) * 100
    
    def get_cost_summary(self) -> str:
        """Formatierte Kosten-Zusammenfassung."""
        if self.total_cost_usd is None:
            return "Kosten nicht verfuegbar"
        
        cost_str = f"Gesamtkosten: ${self.total_cost_usd:.4f} USD"
        if self.cost_per_document_usd is not None and self.successful_documents > 0:
            cost_str += f" | Pro Dokument: ${self.cost_per_document_usd:.6f} USD"
        return cost_str


class DocumentProcessor:
    """
    Verarbeitet Dokumente aus der Eingangsbox und verschiebt sie in Ziel-Boxen.
    
    Workflow:
    1. Dokumente aus Eingangsbox holen
    2. Fuer jedes Dokument:
       a) In Verarbeitungsbox verschieben
       b) Klassifizieren (XML-Roh, GDV, PDF)
       c) Bei PDF: KI-Klassifikation (Courtage/Sach/Leben/Sonstige)
       d) Optional: Umbenennen
       e) In Ziel-Box verschieben
    """
    
    def __init__(self, api_client: APIClient):
        self.api_client = api_client
        self.docs_api = DocumentsAPI(api_client)
        self.history_api = ProcessingHistoryAPI(api_client)
        self.settings_api = ProcessingSettingsAPI(api_client)
        self.doc_rules_api = DocumentRulesAPI(api_client)
        self.openrouter: Optional[OpenRouterClient] = None
        # Content-Hash-Cache fuer Deduplizierung (thread-safe)
        # Spart KI-Kosten wenn identische Dokumente mehrfach verarbeitet werden
        self._classification_cache: dict = {}  # hash -> (target_box, category, new_filename, vu_name, classification_source, classification_confidence, classification_reason)
        self._cache_lock = threading.Lock()
        # KI-Einstellungen (einmal pro Verarbeitungslauf geladen)
        self._ai_settings: Optional[dict] = None
        # Dokumenten-Regeln (einmal pro Verarbeitungslauf geladen)
        self._doc_rules: Optional[DocumentRulesSettings] = None
        
    def _get_openrouter(self) -> OpenRouterClient:
        """Lazy-Init des OpenRouter-Clients."""
        if self.openrouter is None:
            self.openrouter = OpenRouterClient(self.api_client)
        return self.openrouter
    
    def _get_cached_classification(self, content_hash: Optional[str]) -> Optional[dict]:
        """
        Prueft ob eine Klassifikation fuer diesen Content-Hash bereits im Cache liegt.
        
        Kostenoptimierung: Identische Dokumente werden nur 1x per KI klassifiziert.
        
        Args:
            content_hash: SHA256-Hash des Dokument-Inhalts
            
        Returns:
            Cached Klassifikations-dict oder None
        """
        if not content_hash:
            return None
        with self._cache_lock:
            return self._classification_cache.get(content_hash)
    
    def _cache_classification(self, content_hash: Optional[str], result: dict) -> None:
        """
        Speichert eine Klassifikation im Cache fuer Deduplizierung.
        
        Args:
            content_hash: SHA256-Hash des Dokument-Inhalts
            result: Klassifikations-Ergebnis als dict
        """
        if not content_hash:
            return
        with self._cache_lock:
            self._classification_cache[content_hash] = result
            logger.debug(f"Klassifikation gecached fuer Hash {content_hash[:12]}...")
    
    def _load_ai_settings(self) -> dict:
        """
        Laedt KI-Einstellungen vom Server (einmal pro Verarbeitungslauf).
        
        Cached das Ergebnis in self._ai_settings, damit nicht pro Dokument
        ein API-Call gemacht wird.
        
        Returns:
            Dict mit stage1_*, stage2_* Feldern oder leeres Dict bei Fehler
        """
        if self._ai_settings is not None:
            return self._ai_settings
        
        try:
            self._ai_settings = self.settings_api.get_ai_settings()
            if self._ai_settings:
                logger.info(
                    f"KI-Settings geladen: S1={self._ai_settings.get('stage1_model')}, "
                    f"S2={'aktiv' if self._ai_settings.get('stage2_enabled') else 'deaktiviert'} "
                    f"({self._ai_settings.get('stage2_model')})"
                )
            else:
                logger.warning("KI-Settings leer, verwende Defaults")
                self._ai_settings = {}
        except Exception as e:
            logger.warning(f"KI-Settings konnten nicht geladen werden, verwende Defaults: {e}")
            self._ai_settings = {}
        
        return self._ai_settings
    
    def _get_classify_kwargs(self) -> dict:
        """
        Erstellt die kwargs fuer classify_sparte_with_date() aus den geladenen Settings.
        
        Returns:
            Dict mit stage1_*, stage2_* Parametern oder leeres Dict (Defaults)
        """
        settings = self._load_ai_settings()
        if not settings:
            return {}
        
        kwargs = {}
        
        # Stufe 1
        if settings.get('stage1_prompt'):
            kwargs['stage1_prompt'] = settings['stage1_prompt']
        if settings.get('stage1_model'):
            kwargs['stage1_model'] = settings['stage1_model']
        if settings.get('stage1_max_tokens'):
            kwargs['stage1_max_tokens'] = int(settings['stage1_max_tokens'])
        
        # Stufe 2
        kwargs['stage2_enabled'] = bool(settings.get('stage2_enabled', True))
        if settings.get('stage2_prompt'):
            kwargs['stage2_prompt'] = settings['stage2_prompt']
        if settings.get('stage2_model'):
            kwargs['stage2_model'] = settings['stage2_model']
        if settings.get('stage2_max_tokens'):
            kwargs['stage2_max_tokens'] = int(settings['stage2_max_tokens'])
        if settings.get('stage2_trigger'):
            kwargs['stage2_trigger'] = settings['stage2_trigger']
        
        return kwargs
    
    def process_inbox(self, 
                      progress_callback: Optional[Callable[[int, int, str], None]] = None,
                      max_workers: int = DEFAULT_MAX_WORKERS
                      ) -> BatchProcessingResult:
        """
        Verarbeitet alle Dokumente in der Eingangsbox PARALLEL.
        
        Inkludiert Kosten-Tracking:
        - Guthaben vor/nach der Verarbeitung
        - Gesamtkosten und Kosten pro Dokument
        - Logging in der Datenbank
        
        Args:
            progress_callback: Optional - Callback fuer Fortschritt (current, total, message)
            max_workers: Anzahl paralleler Verarbeitungen (default: 4)
            
        Returns:
            BatchProcessingResult mit allen Ergebnissen und Kosten
        """
        results = []
        start_time = datetime.now()
        
        # KI-Settings fuer diesen Lauf laden (einmalig, nicht pro Dokument)
        self._ai_settings = None  # Cache zuruecksetzen
        self._load_ai_settings()
        
        # Dokumenten-Regeln fuer diesen Lauf laden (einmalig)
        self._doc_rules = None
        self._load_document_rules()
        
        # Dokumente aus Eingangsbox holen
        inbox_docs = self.docs_api.list_by_box('eingang')
        
        # Manuell ausgeschlossene Dokumente ueberspringen
        excluded_docs = [d for d in inbox_docs if d.processing_status == 'manual_excluded']
        if excluded_docs:
            logger.info(
                f"{len(excluded_docs)} Dokument(e) uebersprungen (manuell bearbeitet)"
            )
            inbox_docs = [d for d in inbox_docs if d.processing_status != 'manual_excluded']
        
        total = len(inbox_docs)
        
        if total == 0:
            logger.info("Keine Dokumente in der Eingangsbox")
            return BatchProcessingResult(
                results=[],
                total_documents=0,
                successful_documents=0,
                failed_documents=0,
                duration_seconds=0.0
            )
        
        # ============================================
        # KOSTEN-TRACKING: Guthaben VOR Verarbeitung
        # ============================================
        credits_before = None
        credits_provider = 'openrouter'
        try:
            openrouter = self._get_openrouter()
            credits_info = openrouter.get_credits()
            if credits_info:
                credits_provider = credits_info.get('provider', 'openrouter')
                if credits_provider == 'openai':
                    usage_usd = credits_info.get('total_usage')
                    period = credits_info.get('period', '')
                    if usage_usd is not None:
                        logger.info(f"OpenAI-Kosten im Zeitraum {period}: ${usage_usd:.4f} USD")
                    else:
                        logger.info("OpenAI aktiv (Billing-API nicht verfuegbar fuer Service-Accounts)")
                    credits_before = usage_usd
                else:
                    credits_before = credits_info.get('balance', 0.0)
                    logger.info(f"OpenRouter-Guthaben vor Verarbeitung: ${credits_before:.6f} USD")
        except Exception as e:
            logger.warning(f"Konnte Guthaben nicht abrufen: {e}")
        
        logger.info(f"Verarbeite {total} Dokument(e) aus der Eingangsbox (parallel, {max_workers} Worker)")
        
        # Thread-sicherer Counter fuer Progress
        completed_count = [0]  # Liste als mutable Container
        progress_lock = threading.Lock()
        
        def process_with_progress(doc: Document) -> ProcessingResult:
            """Wrapper der Progress-Callback aufruft."""
            result = self._process_document(doc)
            
            # Thread-sicher den Counter erhoehen
            with progress_lock:
                completed_count[0] += 1
                current = completed_count[0]
            
            if progress_callback:
                if result.success:
                    status = "OK"
                elif result.error:
                    status = "FEHLER"
                else:
                    status = "SONSTIGE"
                progress_callback(current, total, f"{status}: {doc.original_filename}")
            
            return result
        
        # Parallele Verarbeitung mit ThreadPoolExecutor
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # Alle Dokumente gleichzeitig starten
            future_to_doc = {
                executor.submit(process_with_progress, doc): doc 
                for doc in inbox_docs
            }
            
            # Ergebnisse einsammeln sobald fertig
            for future in as_completed(future_to_doc):
                doc = future_to_doc[future]
                try:
                    result = future.result()
                    results.append(result)
                    
                    if result.success:
                        logger.info(f"Dokument {doc.id} -> {result.target_box}: {result.new_filename or doc.original_filename}")
                    elif result.error:
                        logger.error(f"Dokument {doc.id} Fehler: {result.error}")
                    else:
                        logger.info(f"Dokument {doc.id} -> {result.target_box}: nicht zugeordnet ({result.category or 'unbekannt'})")
                        
                except Exception as e:
                    logger.exception(f"Unerwarteter Fehler bei Dokument {doc.id}")
                    results.append(ProcessingResult(
                        document_id=doc.id,
                        original_filename=doc.original_filename,
                        success=False,
                        target_box='eingang',
                        error=str(e)
                    ))
        
        # ============================================
        # KOSTEN-TRACKING: Guthaben NACH Verarbeitung
        # HINWEIS: Credits-After wird NICHT sofort abgefragt!
        # OpenRouter aktualisiert das Guthaben verzoegert (1-3 Minuten).
        # Die eigentliche Kosten-Berechnung erfolgt per Timer in der UI
        # (siehe archive_boxes_view.py -> _start_delayed_cost_check).
        # ============================================
        
        # Dauer berechnen
        end_time = datetime.now()
        duration = (end_time - start_time).total_seconds()
        
        successful_count = sum(1 for r in results if r.success)
        failed_count = total - successful_count
        
        # Akkumulierte Kosten aus den Einzel-Ergebnissen
        accumulated_cost = sum(r.cost_usd for r in results)
        cost_per_doc = accumulated_cost / total if total > 0 else 0.0
        
        logger.info(f"Verarbeitung abgeschlossen: {successful_count}/{total} erfolgreich in {duration:.1f}s")
        logger.info(f"Akkumulierte KI-Kosten: ${accumulated_cost:.6f} USD (${cost_per_doc:.6f}/Dok)")
        
        if credits_provider == 'openai':
            if credits_before is not None:
                logger.info(f"OpenAI-Usage vor Verarbeitung: ${credits_before:.6f} USD")
        elif credits_before is not None:
            logger.info(f"Guthaben vor Verarbeitung: ${credits_before:.6f} USD")
            logger.info("Kosten-Berechnung erfolgt verzoegert (OpenRouter-Guthaben braucht 1-3 Min)")
        
        return BatchProcessingResult(
            results=results,
            total_documents=total,
            successful_documents=successful_count,
            failed_documents=failed_count,
            duration_seconds=duration,
            credits_before=credits_before,
            credits_after=None,
            total_cost_usd=accumulated_cost,
            cost_per_document_usd=cost_per_doc,
            provider=credits_provider
        )
    
    def log_batch_complete(self,
                          batch_result: 'BatchProcessingResult') -> Optional[int]:
        """
        Loggt den Abschluss eines Batch-Verarbeitungslaufs in der Datenbank.
        
        Wird direkt nach Verarbeitungsende aufgerufen, OHNE Kosten
        (diese werden spaeter per log_delayed_costs nachgetragen).
        
        Returns:
            ID des History-Eintrags (fuer spaeteres Update) oder None
        """
        try:
            action_details = {
                'batch_type': 'inbox_processing',
                'total_documents': batch_result.total_documents,
                'successful_documents': batch_result.successful_documents,
                'failed_documents': batch_result.failed_documents,
                'duration_seconds': round(batch_result.duration_seconds, 2),
                'timestamp': datetime.now().isoformat(),
                'provider': batch_result.provider,
                'cost_pending': True
            }
            
            if batch_result.credits_before is not None:
                action_details['credits_before_usd'] = round(batch_result.credits_before, 6)
            if batch_result.total_cost_usd is not None and batch_result.total_cost_usd > 0:
                action_details['accumulated_cost_usd'] = round(batch_result.total_cost_usd, 6)
                action_details['cost_per_document_usd'] = round(batch_result.cost_per_document_usd or 0, 8)
            
            entry_id = self.history_api.create(
                document_id=None,
                action='batch_complete',
                new_status='completed',
                previous_status='processing',
                action_details=action_details,
                success=batch_result.failed_documents == 0,
                duration_ms=int(batch_result.duration_seconds * 1000),
                classification_source='batch_processor',
                classification_result=f'{batch_result.successful_documents}/{batch_result.total_documents} OK'
            )
            
            logger.debug(f"Batch-Abschluss geloggt (Kosten ausstehend): ID={entry_id}")
            return entry_id
            
        except Exception as e:
            logger.warning(f"Batch-Abschluss-Logging fehlgeschlagen: {e}")
            return None
    
    def log_delayed_costs(self,
                          history_entry_id: int,
                          batch_result: 'BatchProcessingResult',
                          credits_after: float,
                          provider: str = 'openrouter') -> Optional[dict]:
        """
        Traegt die Kosten nachtraeglich in einen bestehenden History-Eintrag ein.
        
        Kosten-Quellen (nach Prioritaet):
        1. Akkumulierte Server-Kosten aus ai_requests (praezise, provider-unabhaengig)
        2. OpenRouter Balance-Diff (Fallback fuer OpenRouter)
        
        Args:
            history_entry_id: ID des batch_complete History-Eintrags
            batch_result: Das BatchProcessingResult mit akkumulierten Kosten
            credits_after: Das Guthaben NACH der Verarbeitung (verzoegert abgefragt)
            provider: Aktiver Provider ('openrouter' oder 'openai')
            
        Returns:
            Dict mit berechneten Kosten oder None bei Fehler
        """
        try:
            successful_count = batch_result.successful_documents
            
            # Primaere Quelle: akkumulierte Server-Kosten (aus model_pricing)
            accumulated_cost = batch_result.total_cost_usd
            
            if accumulated_cost and accumulated_cost > 0:
                total_cost = accumulated_cost
                cost_source = 'accumulated'
            elif provider == 'openrouter' and batch_result.credits_before is not None:
                total_cost = batch_result.credits_before - (credits_after or 0)
                cost_source = 'balance_diff'
            else:
                total_cost = accumulated_cost or 0
                cost_source = 'accumulated_fallback'
            
            cost_per_doc = total_cost / successful_count if successful_count > 0 else (
                total_cost / batch_result.total_documents if batch_result.total_documents > 0 else 0
            )
            
            logger.info(f"=== KOSTEN-ZUSAMMENFASSUNG ({provider.upper()}, {cost_source}) ===")
            if provider == 'openrouter' and batch_result.credits_before is not None:
                logger.info(f"Guthaben vorher:  ${batch_result.credits_before:.6f} USD")
                logger.info(f"Guthaben nachher: ${credits_after:.6f} USD")
                balance_diff = batch_result.credits_before - (credits_after or 0)
                logger.info(f"Balance-Diff:     ${balance_diff:.6f} USD")
            logger.info(f"Server-Kosten:    ${accumulated_cost or 0:.6f} USD (aus model_pricing)")
            logger.info(f"Gesamtkosten:     ${total_cost:.6f} USD")
            if cost_per_doc:
                logger.info(f"Kosten/Dokument:  ${cost_per_doc:.8f} USD ({batch_result.total_documents} Dokumente)")
            logger.info(f"==========================================")
            
            cost_details = {
                'batch_type': 'cost_update',
                'reference_entry_id': history_entry_id,
                'provider': provider,
                'cost_source': cost_source,
                'accumulated_cost_usd': round(accumulated_cost or 0, 6),
                'credits_before_usd': round(batch_result.credits_before or 0, 6),
                'credits_after_usd': round(credits_after or 0, 6),
                'total_cost_usd': round(total_cost, 6),
                'cost_per_document_usd': round(cost_per_doc, 8),
                'total_documents': batch_result.total_documents,
                'successful_documents': successful_count,
                'failed_documents': batch_result.failed_documents,
                'duration_seconds': round(batch_result.duration_seconds, 2),
                'timestamp': datetime.now().isoformat(),
                'cost_pending': False
            }
            
            self.history_api.create(
                document_id=None,
                action='batch_cost_update',
                new_status='completed',
                previous_status='completed',
                action_details=cost_details,
                success=True,
                duration_ms=0,
                classification_source='cost_tracker',
                classification_result=f'${total_cost:.4f} USD ({successful_count} Dok.)'
            )
            
            return {
                'credits_before': batch_result.credits_before,
                'credits_after': credits_after,
                'total_cost_usd': total_cost,
                'cost_per_document_usd': cost_per_doc,
                'successful_documents': successful_count,
                'provider': provider,
                'cost_source': cost_source
            }
            
        except Exception as e:
            logger.warning(f"Verzoegertes Kosten-Logging fehlgeschlagen: {e}")
            return None
    
    def _process_document(self, doc: Document) -> ProcessingResult:
        """
        Verarbeitet ein einzelnes Dokument.
        
        LOGIK (BiPRO-Code-basiert, optimiert, v2.0.5):
        1. In Verarbeitungsbox verschieben
        1b. Content-Hash-Deduplizierung (spart 100% KI-Kosten bei Duplikaten)
        2. XML-Rohdateien -> Roh Archiv (keine KI)
        3. GDV per BiPRO-Code (999xxx) + Content-Verifikation -> GDV Box (KEINE KI!)
           Bei fehlgeschlagener Verifikation: PDF validieren + KI-Klassifikation
        4. GDV per Dateiendung/Content -> GDV Box + Metadaten aus Datensatz (KEINE KI!)
        5. PDF mit BiPRO-Code + PDF-Validierung vor KI:
           a) Courtage (300xxx) -> Courtage Box + KI nur fuer VU+Datum (~200 Token)
           b) VU-Dokumente -> Sparten-KI + minimale Benennung
        6. PDF ohne BiPRO-Code + PDF-Validierung vor KI -> Sparten-KI
        7. Tabellarische Dateien (CSV/TSV/Excel) -> KI-Klassifikation per Text
        8. Rest -> Sonstige
        
        KOSTENOPTIMIERUNGEN:
        - Content-Hash-Cache: Duplikate werden ohne KI klassifiziert
        - PDF-Validierung: Korrupte PDFs ueberspringen KI (-> Beschaedigte_Datei)
        - Verschluesselungserkennung: Versuch mit bekannten Passwoertern zu entsperren
        - GDV-Verifikation: 999er-Codes mit nicht-GDV-Inhalt -> KI-Klassifikation fuer gueltige PDFs
        
        Args:
            doc: Das zu verarbeitende Dokument
            
        Returns:
            ProcessingResult mit Ergebnis
        """
        start_time = datetime.now()
        previous_status = doc.processing_status or 'pending'
        _doc_cost_usd = 0.0
        
        try:
            # 0. Re-Verifikation: Dokument nochmal frisch vom Server holen
            #    und pruefen ob es noch verarbeitet werden soll
            #    (Schutz gegen Server-Caching bei list_by_box)
            fresh_doc = self.docs_api.get_document(doc.id)
            if fresh_doc and fresh_doc.processing_status == 'manual_excluded':
                logger.info(
                    f"Dokument {doc.id} ({doc.original_filename}): "
                    f"Uebersprungen (manuell ausgeschlossen)"
                )
                return ProcessingResult(
                    document_id=doc.id,
                    original_filename=doc.original_filename,
                    success=True,
                    target_box=doc.box_type,
                    category='manual_excluded'
                )
            
            # 1. Status: downloaded -> processing (In Verarbeitungsbox verschieben)
            self.docs_api.update(doc.id, 
                                 box_type='verarbeitung', 
                                 processing_status='processing')
            
            logger.debug(f"Dokument {doc.id}: Status -> processing")
            
            # History: Start der Verarbeitung
            self._log_history(doc.id, 'start_processing', 'processing',
                              previous_status=previous_status,
                              action_details={'source_box': doc.box_type})
            
            target_box = 'sonstige'
            category = None
            new_filename = None
            
            # Audit-Metadaten fuer Klassifikation
            classification_source = None      # ki_gpt4o, rule_bipro, fallback, etc.
            classification_confidence = None  # high, medium, low
            classification_reason = None      # Begruendung
            
            # AI-Data Persistierung: Volltext und ki_result werden in den
            # Verarbeitungszweigen gesetzt und nach Archive gespeichert.
            # WICHTIG: Volltext muss INNERHALB des tempfile-Blocks extrahiert werden,
            # da der pdf_path danach nicht mehr existiert!
            _ai_extracted_text = None   # Volltext aller Seiten (String)
            _ai_page_count = None       # Anzahl Seiten mit Text
            _ki_result_for_ai = None    # KI-Ergebnis mit _usage/_raw_response/_prompt_text
            
            # 1b. Content-Hash-Deduplizierung: Wenn identisches Dokument bereits
            #     klassifiziert wurde, KI ueberspringen (spart 100% Token-Kosten)
            cached = self._get_cached_classification(doc.content_hash)
            if cached:
                target_box = cached['target_box']
                category = cached['category']
                new_filename = cached.get('new_filename')
                classification_source = 'cache_dedup'
                classification_confidence = cached.get('classification_confidence', 'high')
                classification_reason = f'Deduplizierung: identischer Inhalt bereits klassifiziert (Hash {doc.content_hash[:12] if doc.content_hash else "?"}...)'
                logger.info(
                    f"Dokument {doc.id} ({doc.original_filename}): "
                    f"Duplikat erkannt -> {target_box} (aus Cache)"
                )
            
            # 2. XML-Rohdateien -> Roh Archiv (keine KI)
            elif self._is_xml_raw(doc):
                target_box = 'roh'
                category = 'xml_raw'
                classification_source = 'rule_pattern'
                classification_confidence = 'high'
                classification_reason = 'XML-Rohdatei erkannt (Dateiname-Pattern)'
                logger.info(f"XML-Rohdatei erkannt: {doc.original_filename} -> roh")
            
            # 3. GDV ueber BiPRO-Code (999xxx)
            # Diese Dateien werden vom BiPRO manchmal als .pdf geliefert, sind aber GDV-Datensaetze.
            # WICHTIG: Verifizierung per Content-Check! Nicht blind dem BiPRO-Code vertrauen,
            # da manche 999er-Dateien korrupt oder kein gueltiges GDV sind.
            elif self._is_bipro_gdv(doc):
                gdv_verified = False
                
                # Herunterladen und GDV-Content verifizieren
                try:
                    with tempfile.TemporaryDirectory() as tmpdir:
                        local_path = self.docs_api.download(doc.id, tmpdir)
                        if local_path:
                            # Pruefen ob Datei tatsaechlich GDV-Inhalt hat (erste Zeile = '0001')
                            vu_nummer, absender, datum_iso = self._extract_gdv_metadata(local_path)
                            
                            # GDV ist verifiziert wenn VU-Nummer oder Absender extrahiert werden konnten
                            # (Fallback-Werte 'Xvu'/'kDatum' zaehlen NICHT als erfolgreich)
                            from config.processing_rules import GDV_FALLBACK_VU
                            if (vu_nummer and vu_nummer != GDV_FALLBACK_VU) or absender:
                                gdv_verified = True
                                target_box = 'gdv'
                                category = 'gdv_bipro'
                                classification_source = 'rule_bipro'
                                classification_confidence = 'high'
                                classification_reason = f'BiPRO-Code {doc.bipro_category} + GDV-Content verifiziert'
                                logger.info(f"GDV per BiPRO-Code verifiziert: {doc.original_filename} (Code: {doc.bipro_category}) -> gdv")
                                
                                parts = []
                                if absender:
                                    parts.append(self._slugify(absender))
                                elif vu_nummer:
                                    parts.append(vu_nummer)
                                if datum_iso and datum_iso != 'kDatum':
                                    parts.append(datum_iso)
                                if vu_nummer and absender:
                                    parts.append(f"VU{vu_nummer}")
                                if parts:
                                    new_filename = '_'.join(parts) + '.gdv'
                                    logger.info(f"GDV-Metadaten: Absender={absender}, VU={vu_nummer}, Datum={datum_iso}")
                except Exception as e:
                    logger.warning(f"GDV-Verifikation fehlgeschlagen (BiPRO): {e}")
                
                # Wenn GDV-Content NICHT verifiziert: PDF validieren und ggf. KI-Klassifikation
                # VEMA und andere VUs liefern unter 999xxx-Codes teilweise echte PDFs
                # (z.B. Concordia Lebensversicherung Beispielrechnungen), die NICHT GDV sind.
                # Diese PDFs sind NICHT beschaedigt - der BiPRO-Code war nur falsch.
                if not gdv_verified:
                    ext = doc.file_extension.lower()
                    if ext == '.pdf' or doc.is_pdf:
                        logger.info(
                            f"BiPRO-Code {doc.bipro_category} ist kein GDV: "
                            f"{doc.original_filename} -> Pruefe PDF und starte KI-Klassifikation"
                        )
                        try:
                            with tempfile.TemporaryDirectory() as tmpdir_fallback:
                                local_path_fb = self.docs_api.download(doc.id, tmpdir_fallback)
                                if local_path_fb:
                                    is_valid, repaired_path = self._validate_pdf(local_path_fb)
                                    if not is_valid:
                                        # PDF korrupt oder verschluesselt (kein Passwort bekannt)
                                        target_box = 'sonstige'
                                        # Verschluesselt vs. korrupt unterscheiden
                                        try:
                                            import fitz as _fitz_check
                                            _doc_check = _fitz_check.open(local_path_fb)
                                            _is_enc = _doc_check.is_encrypted and _doc_check.needs_pass
                                            _doc_check.close()
                                        except Exception:
                                            _is_enc = False
                                        if _is_enc:
                                            category = 'pdf_encrypted'
                                            # Original-Dateiname beibehalten bei verschluesselten PDFs
                                            new_filename = None
                                            classification_source = 'rule_validation'
                                            classification_confidence = 'high'
                                            classification_reason = f'BiPRO-Code {doc.bipro_category} + PDF verschluesselt (kein Passwort)'
                                        else:
                                            category = 'pdf_corrupt_bipro'
                                            new_filename = "Beschaedigte_Datei.pdf"
                                            classification_source = 'rule_validation'
                                            classification_confidence = 'high'
                                            classification_reason = f'BiPRO-Code {doc.bipro_category} + PDF korrupt/nicht lesbar'
                                    else:
                                        # PDF ist gueltig -> KI-Klassifikation (wie Schritt 5b/6)
                                        pdf_path = repaired_path or local_path_fb
                                        self._check_and_log_empty_pages(doc, pdf_path)
                                        openrouter = self._get_openrouter()
                                        ki_result = openrouter.classify_sparte_with_date(pdf_path, **self._get_classify_kwargs())
                                        if ki_result:
                                            _doc_cost_usd += ki_result.get('_server_cost_usd', 0) or 0
                                        
                                        _ai_extracted_text, _ai_page_count = self._extract_full_text(pdf_path)
                                        _ki_result_for_ai = ki_result
                                        
                                        if ki_result is None:
                                            ki_result = {}
                                        
                                        sparte = ki_result.get('sparte', 'sonstige')
                                        date_iso = ki_result.get('document_date_iso')
                                        ki_vu_name = ki_result.get('vu_name')
                                        ki_confidence = ki_result.get('confidence', 'medium')
                                        ki_doc_name = ki_result.get('document_name')
                                        
                                        target_box = sparte
                                        category = f'sparte_{sparte}'
                                        
                                        classification_source = 'ki_gpt4o_mini' if ki_confidence == 'high' else 'ki_gpt4o_zweistufig'
                                        classification_confidence = ki_confidence
                                        classification_reason = (
                                            f'KI-Klassifikation (BiPRO 999xxx nicht-GDV PDF): '
                                            f'{sparte} ({ki_confidence})'
                                        )
                                        
                                        logger.info(
                                            f"999xxx-Fallback KI: {doc.original_filename} -> "
                                            f"{sparte} (confidence: {ki_confidence})"
                                        )
                                        
                                        # Benennung (analog Schritt 6)
                                        vu_slug = self._slugify(ki_vu_name) if ki_vu_name else 'Unbekannt'
                                        if sparte == 'courtage':
                                            if date_iso:
                                                new_filename = f"{vu_slug}_Courtage_{date_iso}.pdf"
                                            else:
                                                new_filename = f"{vu_slug}_Courtage.pdf"
                                        elif sparte == 'sonstige' and ki_doc_name:
                                            doc_slug = self._slugify(ki_doc_name)
                                            new_filename = f"{vu_slug}_{doc_slug}.pdf"
                                        elif sparte in ['sach', 'leben', 'kranken']:
                                            new_filename = f"{vu_slug}_{sparte.capitalize()}.pdf"
                                        elif date_iso:
                                            new_filename = f"{vu_slug}_{sparte.capitalize()}_{date_iso}.pdf"
                        except Exception as e:
                            logger.warning(f"999xxx-Fallback KI-Klassifikation fehlgeschlagen: {e}")
                            target_box = 'sonstige'
                            category = 'pdf_error'
                            new_filename = "Beschaedigte_Datei.pdf"
                            classification_source = 'fallback'
                            classification_confidence = 'low'
                            classification_reason = f'BiPRO 999xxx Fallback KI fehlgeschlagen: {str(e)[:100]}'
                    else:
                        logger.warning(
                            f"BiPRO-Code {doc.bipro_category} behauptet GDV, aber Content-Verifikation fehlgeschlagen: "
                            f"{doc.original_filename} -> kein PDF, nicht klassifizierbar"
                        )
                        target_box = 'sonstige'
                        category = 'unknown_bipro'
                        classification_source = 'fallback'
                        classification_confidence = 'low'
                        classification_reason = f'BiPRO-Code {doc.bipro_category} aber Content nicht verifizierbar'
            
            # 4. GDV-Dateien per Dateiendung/Content -> GDV Box + Metadaten aus Datensatz
            # Diese Prüfung kommt NACH BiPRO-Code, da BiPRO-Code zuverlässiger ist
            elif self._is_gdv_file(doc):
                target_box = 'gdv'
                category = 'gdv'
                classification_source = 'rule_extension'
                classification_confidence = 'high'
                classification_reason = 'GDV-Datei erkannt (Dateiendung/Content)'
                
                # VU, Absender und Datum aus GDV-Datensatz extrahieren (KEINE KI!)
                try:
                    with tempfile.TemporaryDirectory() as tmpdir:
                        local_path = self.docs_api.download(doc.id, tmpdir)
                        if local_path:
                            vu_nummer, absender, datum_iso = self._extract_gdv_metadata(local_path)
                            if vu_nummer or absender or datum_iso:
                                # Dateiname anpassen fuer chronologische Sortierung
                                # Format: Absender_Datum_VU.gdv (Absender ist der Versicherer-Name)
                                parts = []
                                
                                # Absender (Versicherer-Name) hat Prioritaet, falls vorhanden
                                if absender:
                                    parts.append(self._slugify(absender))
                                elif vu_nummer:
                                    parts.append(vu_nummer)
                                
                                if datum_iso:
                                    parts.append(datum_iso)
                                
                                # VU-Nummer als Fallback wenn kein Absender
                                if vu_nummer and not absender:
                                    pass  # Bereits hinzugefuegt
                                elif vu_nummer and absender:
                                    parts.append(f"VU{vu_nummer}")  # Als Zusatz
                                
                                if parts:
                                    new_filename = '_'.join(parts) + '.gdv'
                                    logger.info(f"GDV-Metadaten: Absender={absender}, VU={vu_nummer}, Datum={datum_iso}")
                except Exception as e:
                    logger.warning(f"GDV-Metadaten-Extraktion fehlgeschlagen: {e}")
            
            # 5. PDF mit BiPRO-Kategorie -> Box nach BiPRO-Code
            # WICHTIG: Dies ist ein elif, um die if/elif-Kette nicht zu brechen!
            elif doc.is_pdf and doc.bipro_category:
                logger.debug(f"PDF mit BiPRO-Kategorie: {doc.original_filename} -> Code: {doc.bipro_category}")
                
                # 5a. Courtage-PDFs (BiPRO-Codes 300xxx)
                if self._is_bipro_courtage(doc):
                    target_box = 'courtage'
                    category = 'courtage_bipro'
                    classification_source = 'rule_bipro'
                    classification_confidence = 'high'
                    classification_reason = f'BiPRO-Code {doc.bipro_category} identifiziert Courtage-Dokument'
                    logger.info(f"Courtage per BiPRO-Code erkannt: {doc.original_filename} -> courtage")
                    
                    # KI nur fuer VU + Datum (minimal, ~200 Token)
                    try:
                        with tempfile.TemporaryDirectory() as tmpdir:
                            local_path = self.docs_api.download(doc.id, tmpdir)
                            if local_path:
                                # PDF-Validierung vor KI-Call (spart Tokens bei korrupten PDFs)
                                is_valid, repaired_path = self._validate_pdf(local_path)
                                if not is_valid:
                                    logger.warning(f"Courtage-PDF korrupt, ueberspringe KI: {doc.original_filename}")
                                    new_filename = "Beschaedigte_Datei_Courtage.pdf"
                                else:
                                    pdf_path = repaired_path or local_path
                                    # Leere-Seiten-Erkennung (informativ, blockiert nicht)
                                    self._check_and_log_empty_pages(doc, pdf_path)
                                    openrouter = self._get_openrouter()
                                    result = openrouter.classify_courtage_minimal(pdf_path)
                                    
                                    # AI-Data: Volltext extrahieren + ki_result merken
                                    # (muss im tempfile-Block passieren, da pdf_path danach geloescht wird)
                                    _ai_extracted_text, _ai_page_count = self._extract_full_text(pdf_path)
                                    _ki_result_for_ai = result
                                    
                                    if result:
                                        insurer = result.get('insurer') or 'Unbekannt'
                                        date_iso = result.get('document_date_iso') or ''
                                        
                                        # Dateiname: VU_Courtage_Datum.pdf
                                        insurer_slug = self._slugify(insurer)
                                        if date_iso:
                                            new_filename = f"{insurer_slug}_Courtage_{date_iso}.pdf"
                                        else:
                                            new_filename = f"{insurer_slug}_Courtage.pdf"
                                        
                                        # KI-Quelle zusaetzlich dokumentieren
                                        classification_source = 'ki_courtage_minimal'
                                        classification_reason = f'Courtage via BiPRO + KI-Extraktion: {insurer}, {date_iso}'
                                        
                                        logger.info(f"Courtage klassifiziert: {insurer}, {date_iso}")
                    except Exception as e:
                        logger.warning(f"Courtage-KI fehlgeschlagen: {e}")
                
                # 5b. VU-Dokumente (alle anderen BiPRO-Codes) -> KI fuer Sparte
                else:
                    from config.processing_rules import get_bipro_document_type
                    
                    doc_type = get_bipro_document_type(doc.bipro_category)
                    logger.info(f"VU-Dokument per BiPRO-Code: {doc.original_filename} -> Typ: {doc_type}")
                    
                    # KI: Sparte + Datum bestimmen
                    try:
                        with tempfile.TemporaryDirectory() as tmpdir:
                            local_path = self.docs_api.download(doc.id, tmpdir)
                            if local_path:
                                # PDF-Validierung vor KI-Call (spart Tokens bei korrupten PDFs)
                                is_valid, repaired_path = self._validate_pdf(local_path)
                                if not is_valid:
                                    target_box = 'sonstige'
                                    # Verschluesselt vs. korrupt unterscheiden
                                    try:
                                        import fitz as _fitz_vu
                                        _doc_vu = _fitz_vu.open(local_path)
                                        _enc_vu = _doc_vu.is_encrypted and _doc_vu.needs_pass
                                        _doc_vu.close()
                                    except Exception:
                                        _enc_vu = False
                                    if _enc_vu:
                                        logger.warning(f"VU-PDF verschluesselt, kein Passwort: {doc.original_filename}")
                                        category = 'pdf_encrypted'
                                        new_filename = None
                                        classification_source = 'rule_validation'
                                        classification_confidence = 'high'
                                        classification_reason = f'PDF verschluesselt (kein Passwort), KI uebersprungen'
                                    else:
                                        logger.warning(f"VU-PDF korrupt, ueberspringe KI: {doc.original_filename}")
                                        category = 'pdf_corrupt'
                                        new_filename = "Beschaedigte_Datei.pdf"
                                        classification_source = 'rule_validation'
                                        classification_confidence = 'high'
                                        classification_reason = f'PDF korrupt/nicht lesbar, KI uebersprungen'
                                else:
                                    pdf_path = repaired_path or local_path
                                    # Leere-Seiten-Erkennung (informativ, blockiert nicht)
                                    self._check_and_log_empty_pages(doc, pdf_path)
                                    openrouter = self._get_openrouter()
                                    ki_result = openrouter.classify_sparte_with_date(pdf_path, **self._get_classify_kwargs())
                                    if ki_result:
                                        _doc_cost_usd += ki_result.get('_server_cost_usd', 0) or 0
                                    
                                    # AI-Data: Volltext extrahieren + ki_result merken
                                    _ai_extracted_text, _ai_page_count = self._extract_full_text(pdf_path)
                                    _ki_result_for_ai = ki_result
                                    
                                    # Schutz gegen None-Rueckgabe bei KI-Fehler
                                    if ki_result is None:
                                        ki_result = {}
                                    
                                    sparte = ki_result.get('sparte', 'sonstige')
                                    date_iso = ki_result.get('document_date_iso')
                                    ki_vu_name = ki_result.get('vu_name')
                                    ki_confidence = ki_result.get('confidence', 'medium')
                                    ki_doc_name = ki_result.get('document_name')  # Dokumentname bei sonstige (Stufe 2)
                                    
                                    target_box = sparte
                                    category = f'sparte_{sparte}'
                                    
                                    # Audit-Metadaten - Confidence aus KI uebernehmen (BUG-0007 Fix: == 'high' statt != 'medium')
                                    classification_source = 'ki_gpt4o_mini' if ki_confidence == 'high' else 'ki_gpt4o_zweistufig'
                                    classification_confidence = ki_confidence
                                    classification_reason = f'KI-Sparten-Klassifikation: {sparte} ({ki_confidence}), BiPRO-Typ: {doc_type}'
                                    
                                    logger.info(f"Sparte klassifiziert: {sparte} (confidence: {ki_confidence})")
                                    
                                    # VU-Name: KI-Ergebnis hat Vorrang, dann Dokument-Metadaten
                                    vu_name = ki_vu_name or getattr(doc, 'vu_name', None) or 'Unbekannt'
                                    
                                    parts = []
                                    parts.append(self._slugify(vu_name))
                                    
                                    if sparte == 'sonstige' and ki_doc_name:
                                        # Stufe 2 hat einen Dokumentnamen geliefert
                                        parts.append(self._slugify(ki_doc_name))
                                    else:
                                        parts.append(sparte.capitalize())
                                    
                                    # Dokumenttyp bei Sonstige (aus BiPRO) als Fallback
                                    if sparte == 'sonstige' and not ki_doc_name and doc_type and doc_type != 'unbekannt':
                                        parts.append(doc_type.capitalize())
                                    
                                    # Datum nur bei Courtage
                                    if sparte == 'courtage' and date_iso:
                                        parts.append(date_iso)
                                    
                                    new_filename = '_'.join(parts) + '.pdf'
                                    logger.info(f"VU-Dokument benannt: {new_filename}")
                    except Exception as e:
                        logger.warning(f"Sparten-KI fehlgeschlagen: {e}")
                        target_box = 'sonstige'
                        category = 'pdf_error'
                        classification_source = 'fallback'
                        classification_confidence = 'low'
                        classification_reason = f'KI-Klassifikation fehlgeschlagen: {str(e)[:100]}'
            
            # 5c. Dateinamen-basierte Courtage-Erkennung (Vermittlerabrechnung im Dateinamen)
            elif doc.is_pdf and 'vermittlerabrechnung' in (doc.original_filename or '').lower():
                target_box = 'courtage'
                category = 'courtage_filename'
                classification_source = 'rule_filename'
                classification_confidence = 'high'
                classification_reason = f'Dateiname enthaelt "Vermittlerabrechnung" -> Courtage'
                logger.info(f"Courtage per Dateiname erkannt: {doc.original_filename} -> courtage")
                
                try:
                    with tempfile.TemporaryDirectory() as tmpdir:
                        local_path = self.docs_api.download(doc.id, tmpdir)
                        if local_path:
                            is_valid, repaired_path = self._validate_pdf(local_path)
                            if not is_valid:
                                logger.warning(f"Vermittlerabrechnung-PDF korrupt: {doc.original_filename}")
                                new_filename = "Beschaedigte_Datei_Courtage.pdf"
                            else:
                                pdf_path = repaired_path or local_path
                                self._check_and_log_empty_pages(doc, pdf_path)
                                openrouter = self._get_openrouter()
                                result = openrouter.classify_courtage_minimal(pdf_path)
                                
                                _ai_extracted_text, _ai_page_count = self._extract_full_text(pdf_path)
                                _ki_result_for_ai = result
                                
                                if result:
                                    _doc_cost_usd += result.get('_server_cost_usd', 0) or 0
                                    insurer = result.get('insurer') or 'Unbekannt'
                                    date_iso = result.get('document_date_iso') or ''
                                    
                                    insurer_slug = self._slugify(insurer)
                                    if date_iso:
                                        new_filename = f"{insurer_slug}_Courtage_{date_iso}.pdf"
                                    else:
                                        new_filename = f"{insurer_slug}_Courtage.pdf"
                                    
                                    classification_source = 'rule_filename_ki'
                                    classification_reason = f'Dateiname "Vermittlerabrechnung" + KI: {insurer}, {date_iso}'
                                    logger.info(f"Vermittlerabrechnung klassifiziert: {insurer}, {date_iso}")
                except Exception as e:
                    logger.warning(f"Vermittlerabrechnung-KI fehlgeschlagen: {e}")
            
            # 6. PDFs ohne BiPRO-Kategorie -> KI fuer Sparte
            elif doc.is_pdf:
                logger.debug(f"PDF ohne BiPRO-Kategorie: {doc.original_filename}")
                
                try:
                    with tempfile.TemporaryDirectory() as tmpdir:
                        local_path = self.docs_api.download(doc.id, tmpdir)
                        if local_path:
                            # PDF-Validierung vor KI-Call (spart Tokens bei korrupten PDFs)
                            is_valid, repaired_path = self._validate_pdf(local_path)
                            if not is_valid:
                                target_box = 'sonstige'
                                # Verschluesselt vs. korrupt unterscheiden
                                try:
                                    import fitz as _fitz_s6
                                    _doc_s6 = _fitz_s6.open(local_path)
                                    _enc_s6 = _doc_s6.is_encrypted and _doc_s6.needs_pass
                                    _doc_s6.close()
                                except Exception:
                                    _enc_s6 = False
                                if _enc_s6:
                                    logger.warning(f"PDF verschluesselt, kein Passwort: {doc.original_filename}")
                                    category = 'pdf_encrypted'
                                    new_filename = None
                                    classification_source = 'rule_validation'
                                    classification_confidence = 'high'
                                    classification_reason = f'PDF verschluesselt (kein Passwort), KI uebersprungen'
                                else:
                                    logger.warning(f"PDF korrupt, ueberspringe KI: {doc.original_filename}")
                                    category = 'pdf_corrupt'
                                    new_filename = "Beschaedigte_Datei.pdf"
                                    classification_source = 'rule_validation'
                                    classification_confidence = 'high'
                                    classification_reason = f'PDF korrupt/nicht lesbar, KI uebersprungen'
                            else:
                                pdf_path = repaired_path or local_path
                                # Leere-Seiten-Erkennung (informativ, blockiert nicht)
                                self._check_and_log_empty_pages(doc, pdf_path)
                                openrouter = self._get_openrouter()
                                ki_result = openrouter.classify_sparte_with_date(pdf_path, **self._get_classify_kwargs())
                                if ki_result:
                                    _doc_cost_usd += ki_result.get('_server_cost_usd', 0) or 0
                                
                                # AI-Data: Volltext extrahieren + ki_result merken
                                _ai_extracted_text, _ai_page_count = self._extract_full_text(pdf_path)
                                _ki_result_for_ai = ki_result
                                
                                # Schutz gegen None-Rueckgabe bei KI-Fehler
                                if ki_result is None:
                                    ki_result = {}
                                
                                sparte = ki_result.get('sparte', 'sonstige')
                                date_iso = ki_result.get('document_date_iso')
                                ki_vu_name = ki_result.get('vu_name')
                                ki_confidence = ki_result.get('confidence', 'medium')
                                ki_doc_name = ki_result.get('document_name')
                                
                                target_box = sparte
                                category = f'sparte_{sparte}'
                                
                                # Audit-Metadaten - Confidence aus KI (BUG-0007 Fix: == 'high' statt != 'medium')
                                classification_source = 'ki_gpt4o_mini' if ki_confidence == 'high' else 'ki_gpt4o_zweistufig'
                                classification_confidence = ki_confidence
                                classification_reason = f'KI-Sparten-Klassifikation ohne BiPRO: {sparte} ({ki_confidence})'
                                
                                logger.info(f"PDF klassifiziert: {sparte} (confidence: {ki_confidence})")
                                
                                # Benennung je nach Sparte
                                vu_slug = self._slugify(ki_vu_name) if ki_vu_name else 'Unbekannt'
                                if sparte == 'courtage':
                                    if date_iso:
                                        new_filename = f"{vu_slug}_Courtage_{date_iso}.pdf"
                                    else:
                                        new_filename = f"{vu_slug}_Courtage.pdf"
                                elif sparte == 'sonstige' and ki_doc_name:
                                    # Stufe 2 hat Dokumentnamen geliefert
                                    doc_slug = self._slugify(ki_doc_name)
                                    new_filename = f"{vu_slug}_{doc_slug}.pdf"
                                elif sparte in ['sach', 'leben', 'kranken']:
                                    new_filename = f"{vu_slug}_{sparte.capitalize()}.pdf"
                                elif date_iso:
                                    new_filename = f"{vu_slug}_{sparte.capitalize()}_{date_iso}.pdf"
                except Exception as e:
                    logger.warning(f"PDF-KI fehlgeschlagen: {e}")
                    target_box = 'sonstige'
                    category = 'pdf_error'
                    classification_source = 'fallback'
                    classification_confidence = 'low'
                    classification_reason = f'KI-Klassifikation fehlgeschlagen: {str(e)[:100]}'
            
            # 7. Tabellarische Dateien (CSV/TSV/Excel) -> KI-Klassifikation per Text
            elif self._is_spreadsheet(doc):
                logger.debug(f"Tabellendatei erkannt: {doc.original_filename}")
                try:
                    with tempfile.TemporaryDirectory() as tmpdir:
                        local_path = self.docs_api.download(doc.id, tmpdir)
                        if local_path:
                            # Erste Zeilen als Text extrahieren fuer KI
                            csv_text = self._extract_spreadsheet_text(local_path)
                            if csv_text and csv_text.strip():
                                openrouter = self._get_openrouter()
                                ki_result = openrouter._classify_sparte_request(csv_text[:2500])
                                
                                _ai_extracted_text = csv_text
                                _ai_page_count = 1
                                _ki_result_for_ai = ki_result
                                
                                if ki_result is None:
                                    ki_result = {}
                                
                                sparte = ki_result.get('sparte', 'sonstige')
                                ki_vu_name = ki_result.get('vu_name')
                                ki_confidence = ki_result.get('confidence', 'medium')
                                
                                target_box = sparte
                                category = f'spreadsheet_{sparte}'
                                classification_source = 'ki_spreadsheet'
                                classification_confidence = ki_confidence
                                classification_reason = f'Tabellendatei KI-klassifiziert: {sparte} ({ki_confidence})'
                                
                                logger.info(f"Tabelle klassifiziert: {doc.original_filename} -> {sparte} ({ki_confidence})")
                            else:
                                target_box = 'sonstige'
                                category = 'spreadsheet_empty'
                                classification_source = 'rule_pattern'
                                classification_confidence = 'low'
                                classification_reason = 'Tabellendatei ohne lesbaren Text'
                except Exception as e:
                    logger.warning(f"Tabellen-Klassifikation fehlgeschlagen: {e}")
                    target_box = 'sonstige'
                    category = 'spreadsheet_error'
                    classification_source = 'fallback'
                    classification_confidence = 'low'
                    classification_reason = f'Tabellen-Klassifikation fehlgeschlagen: {str(e)[:100]}'
            
            # 8. Rest (unbekannte Dateitypen) -> Sonstige
            else:
                target_box = 'sonstige'
                category = 'unknown'
                classification_source = 'fallback'
                classification_confidence = 'low'
                classification_reason = 'Unbekannter Dateityp, keine Klassifikation moeglich'
                logger.debug(f"Unbekannter Dateityp: {doc.original_filename} -> sonstige")
            
            # Content-Hash-Cache: KI-Ergebnis fuer spaetere Duplikate speichern
            # Nur wenn nicht selbst aus Cache und nicht pdf_corrupt/pdf_error
            if (classification_source != 'cache_dedup' 
                    and category not in ('pdf_corrupt', 'pdf_error', None)
                    and doc.content_hash):
                self._cache_classification(doc.content_hash, {
                    'target_box': target_box,
                    'category': category,
                    'new_filename': new_filename,
                    'classification_confidence': classification_confidence,
                    'classification_reason': classification_reason,
                })
            
            # Schritt 1: processing -> classified (mit Klassifikations-Metadaten)
            update_kwargs = {
                'box_type': target_box,
                'processing_status': 'classified',
                'document_category': category,
            }
            
            # Audit-Metadaten hinzufuegen wenn vorhanden
            if classification_source:
                update_kwargs['classification_source'] = classification_source
            if classification_confidence:
                update_kwargs['classification_confidence'] = classification_confidence
            if classification_reason:
                # Auf 500 Zeichen begrenzen (DB-Limit)
                update_kwargs['classification_reason'] = classification_reason[:500] if classification_reason else None
            # Timestamp immer setzen wenn klassifiziert wurde
            update_kwargs['classification_timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            self.docs_api.update(doc.id, **update_kwargs)
            
            # Logging: Sonstige als "nicht zugeordnet" markieren
            if target_box == 'sonstige':
                logger.info(f"Dokument {doc.id}: Nicht zugeordnet -> {category}")
            else:
                logger.debug(f"Dokument {doc.id}: Status -> classified")
            
            # History: Klassifikation abgeschlossen
            duration_ms = int((datetime.now() - start_time).total_seconds() * 1000)
            self._log_history(doc.id, 'classify', 'classified',
                              previous_status='processing',
                              classification_source=classification_source,
                              classification_result=f'{category} -> {target_box}',
                              action_details={
                                  'category': category,
                                  'target_box': target_box,
                                  'confidence': classification_confidence,
                                  'reason': classification_reason[:200] if classification_reason else None
                              },
                              duration_ms=duration_ms)
            
            # Schritt 2: Falls umbenannt: classified -> renamed
            current_status = 'classified'
            if new_filename:
                self.docs_api.update(doc.id, 
                                     original_filename=new_filename,
                                     ai_renamed=True,
                                     processing_status='renamed')
                current_status = 'renamed'
                logger.debug(f"Dokument {doc.id}: Status -> renamed ({new_filename})")
                self._log_history(doc.id, 'rename', 'renamed',
                                  previous_status='classified',
                                  action_details={'new_filename': new_filename})
            
            # Schritt 3: classified/renamed -> archived (in Ziel-Box)
            self.docs_api.update(doc.id, processing_status='archived')
            logger.debug(f"Dokument {doc.id}: Status -> archived (in {target_box})")
            
            # History: Archivierung abgeschlossen
            self._log_history(doc.id, 'archive', 'archived',
                              previous_status=current_status,
                              action_details={'final_box': target_box, 'new_filename': new_filename})
            
            # Nachgelagerter Schritt: AI-Daten persistieren (Volltext + KI-Response)
            # Fehler hier brechen die Verarbeitung NICHT ab
            if _ai_extracted_text is not None or _ki_result_for_ai is not None:
                try:
                    self._persist_ai_data(doc, _ai_extracted_text, _ai_page_count, _ki_result_for_ai)
                except Exception as ai_err:
                    logger.warning(f"AI-Daten-Persistierung fehlgeschlagen fuer Dokument {doc.id}: {ai_err}")
            
            # Dokumenten-Regeln anwenden (Duplikate, leere Seiten)
            # Fehler hier brechen die Verarbeitung NICHT ab
            try:
                self._apply_document_rules(doc)
            except Exception as rule_err:
                logger.warning(f"Dokumenten-Regeln fehlgeschlagen fuer Dokument {doc.id}: {rule_err}")
            
            # Erfolg-Logik:
            # - Erfolgreich = GDV, Courtage, Sach, Leben, Kranken, Roh
            # - Nicht zugeordnet = Sonstige (wird als "failed" gezaehlt)
            is_success = target_box not in ['sonstige']
            
            return ProcessingResult(
                document_id=doc.id,
                original_filename=doc.original_filename,
                success=is_success,
                target_box=target_box,
                category=category,
                new_filename=new_filename,
                cost_usd=_doc_cost_usd
            )
            
        except Exception as e:
            logger.exception(f"Fehler bei Verarbeitung von Dokument {doc.id}")
            
            # Fehler markieren mit Status: error
            try:
                self.docs_api.update(doc.id, 
                                     box_type='sonstige',
                                     processing_status='error',
                                     ai_processing_error=str(e)[:500])
                logger.debug(f"Dokument {doc.id}: Status -> error")
                
                # History: Fehler protokollieren
                duration_ms = int((datetime.now() - start_time).total_seconds() * 1000)
                self._log_history(doc.id, 'error', 'error',
                                  previous_status='processing',
                                  success=False,
                                  error_message=str(e)[:500],
                                  duration_ms=duration_ms)
            except Exception:
                pass
            
            return ProcessingResult(
                document_id=doc.id,
                original_filename=doc.original_filename,
                success=False,
                target_box='sonstige',
                error=str(e)
            )
    
    def _validate_pdf(self, pdf_path: str) -> Tuple[bool, Optional[str]]:
        """
        Validiert ein PDF, erkennt Verschluesselung und versucht bei Fehler Reparatur.
        
        Kostenoptimiert: Verhindert teure KI-Aufrufe fuer korrupte PDFs.
        Verschluesselte PDFs werden automatisch mit bekannten Passwoertern entsperrt.
        
        Args:
            pdf_path: Pfad zur PDF-Datei
            
        Returns:
            (is_valid, repaired_path) - is_valid=True wenn OK oder repariert/entsperrt,
            repaired_path = Pfad zur reparierten/entsperrten Datei (oder None wenn original OK)
        """
        try:
            import fitz  # PyMuPDF
        except ImportError:
            logger.warning("PyMuPDF nicht installiert, ueberspringe PDF-Validierung")
            return (True, None)  # Im Zweifel weiter
        
        try:
            doc = fitz.open(pdf_path)
            page_count = len(doc)
            is_encrypted = doc.is_encrypted
            needs_pass = doc.needs_pass if is_encrypted else False
            doc.close()
            
            # Verschluesselte PDF: Versuche mit bekannten Passwoertern zu entsperren
            if is_encrypted and needs_pass:
                logger.info(f"PDF ist verschluesselt, versuche Entsperrung: {pdf_path}")
                try:
                    from services.pdf_unlock import unlock_pdf_if_needed
                    # api_client aus self holen (DocumentProcessor hat self.docs_api.client)
                    api_client = getattr(self.docs_api, 'client', None)
                    unlocked = unlock_pdf_if_needed(pdf_path, api_client=api_client)
                    if unlocked:
                        logger.info(f"PDF erfolgreich entsperrt: {pdf_path}")
                        return (True, None)  # Original wurde in-place ueberschrieben
                except ValueError as ve:
                    # Kein Passwort passt - PDF bleibt verschluesselt
                    logger.warning(f"PDF-Entsperrung fehlgeschlagen: {ve}")
                    return (False, None)
                except Exception as unlock_err:
                    logger.warning(f"PDF-Entsperrung Fehler: {unlock_err}")
                    return (False, None)
            
            if page_count > 0:
                return (True, None)  # PDF OK
            else:
                logger.warning(f"PDF hat 0 Seiten: {pdf_path}")
                return (False, None)
                
        except Exception as open_error:
            logger.warning(f"PDF defekt ({open_error}), versuche Reparatur: {pdf_path}")
            
            # Reparatur-Versuch: fitz kann defekte PDFs oft retten
            try:
                repaired_path = pdf_path + ".repaired.pdf"
                doc = fitz.open(pdf_path)  # type: ignore[arg-type]
                # Garbage-Collection und Linearisierung
                doc.save(repaired_path, garbage=4, deflate=True, clean=True)
                doc.close()
                
                # Reparierte Datei pruefen
                doc2 = fitz.open(repaired_path)
                page_count = len(doc2)
                doc2.close()
                
                if page_count > 0:
                    logger.info(f"PDF erfolgreich repariert: {repaired_path} ({page_count} Seiten)")
                    return (True, repaired_path)
                else:
                    logger.warning(f"Repariertes PDF hat 0 Seiten")
                    # Aufraeumen
                    try:
                        os.remove(repaired_path)
                    except OSError:
                        pass
                    return (False, None)
                    
            except Exception as repair_error:
                logger.warning(f"PDF-Reparatur fehlgeschlagen: {repair_error}")
                # Aufraeumen falls Datei erstellt wurde
                repaired_path = pdf_path + ".repaired.pdf"
                try:
                    os.remove(repaired_path)
                except OSError:
                    pass
                return (False, None)
    
    def _check_and_log_empty_pages(self, doc: Document, pdf_path: str) -> None:
        """
        Prueft ein PDF auf leere Seiten und speichert das Ergebnis in der DB.
        
        Dieser Schritt ist rein informativ und blockiert NICHT die weitere
        Verarbeitung (KI-Klassifikation, Benennung, Box-Zuweisung laufen
        danach ganz normal weiter).
        
        Wirkung:
            - Speichert empty_page_count + total_page_count via PUT /documents/{id}
              (direkter API-Call, da DocumentsAPI.update() diese Felder nicht kennt)
            - Legt bei Fund einen Activity-Log-Eintrag 'empty_pages_detected' an
            - Bei Fehler: Warnung loggen, Pipeline laeuft weiter (kein Abbruch)
        
        Args:
            doc: Das Document-Objekt
            pdf_path: Pfad zur (ggf. reparierten) PDF-Datei
        """
        try:
            from services.empty_page_detector import get_empty_pages
            
            empty_indices, total_pages = get_empty_pages(pdf_path)
            empty_count = len(empty_indices)
            
            if total_pages == 0:
                # Kein sinnvolles Ergebnis (z.B. PyMuPDF nicht installiert)
                return
            
            # In DB speichern (direkter API-Call, da update() diese Felder nicht kennt)
            try:
                self.docs_api.client.put(
                    f'/documents/{doc.id}',
                    json_data={
                        'empty_page_count': empty_count,
                        'total_page_count': total_pages
                    }
                )
            except Exception as e:
                logger.warning(f"Leere-Seiten-Werte konnten nicht gespeichert werden fuer Dokument {doc.id}: {e}")
            
            # History loggen wenn leere Seiten gefunden
            if empty_count > 0:
                if empty_count == total_pages:
                    detail_msg = f"PDF komplett leer ({total_pages} Seiten)"
                else:
                    detail_msg = f"Leere Seiten erkannt: {empty_count} von {total_pages} (Indizes: {empty_indices})"
                
                self._log_history(
                    document_id=doc.id,
                    action='empty_pages_detected',
                    new_status='processing',
                    success=True,
                    action_details={
                        'empty_page_count': empty_count,
                        'total_page_count': total_pages,
                        'empty_page_indices': empty_indices,
                        'detail': detail_msg
                    }
                )
                logger.info(f"[Leere Seiten] {doc.original_filename}: {detail_msg}")
                
        except Exception as e:
            # Fehler in der Leere-Seiten-Erkennung darf die Pipeline NICHT blockieren
            logger.warning(f"Leere-Seiten-Erkennung fehlgeschlagen fuer {doc.original_filename}: {e}")
    
    def _extract_full_text(self, pdf_path: str) -> tuple:
        """
        Extrahiert Volltext ueber ALLE Seiten einer PDF.
        
        Muss INNERHALB des tempfile-Blocks aufgerufen werden,
        da pdf_path danach nicht mehr existiert.
        
        Args:
            pdf_path: Lokaler Pfad zur PDF-Datei
            
        Returns:
            Tuple (extracted_text: str, pages_with_text: int)
        """
        try:
            import fitz  # PyMuPDF
        except ImportError:
            logger.warning("PyMuPDF nicht verfuegbar fuer Volltext-Extraktion")
            return ("", 0)
        
        extracted_text = ""
        pages_with_text = 0
        
        try:
            pdf_doc = fitz.open(pdf_path)
            for page in pdf_doc:
                page_text = page.get_text("text")
                if page_text and page_text.strip():
                    extracted_text += page_text + "\n"
                    pages_with_text += 1
            pdf_doc.close()
        except Exception as e:
            logger.warning(f"Volltext-Extraktion fehlgeschlagen: {e}")
        
        return (extracted_text, pages_with_text)
    
    def _persist_ai_data(self, doc: Document, extracted_text: str,
                         extracted_page_count: int, ki_result: dict) -> None:
        """
        Persistiert Volltext + KI-Daten in document_ai_data Tabelle.
        
        Laeuft NACH der Klassifikation/Rename/Archive. Ein Fehler hier
        bricht die Verarbeitung NICHT ab (wird im Aufrufer gefangen).
        
        Args:
            doc: Das verarbeitete Dokument
            extracted_text: Bereits extrahierter Volltext (alle Seiten)
            extracted_page_count: Anzahl Seiten mit tatsaechlichem Text
            ki_result: KI-Ergebnis mit optionalen _usage/_raw_response/_prompt_text
        """
        import hashlib
        
        # 1. Extraction-Method bestimmen
        extraction_method = 'text' if (extracted_text and extracted_text.strip()) else 'none'
        
        # 2. SHA256 des Textes berechnen
        text_sha256 = None
        if extracted_text and extracted_text.strip():
            text_sha256 = hashlib.sha256(extracted_text.encode('utf-8')).hexdigest()
        
        # 3. KI-Metadaten aus ki_result extrahieren
        ai_full_response = None
        ai_prompt_text = None
        ai_model = None
        ai_stage = None
        prompt_tokens = None
        completion_tokens = None
        total_tokens = None
        
        if ki_result and isinstance(ki_result, dict):
            # Raw-Response (kann String oder Dict sein bei zweistufig)
            raw_resp = ki_result.get('_raw_response')
            if raw_resp is not None:
                import json as _json
                if isinstance(raw_resp, dict):
                    ai_full_response = _json.dumps(raw_resp, ensure_ascii=False)
                else:
                    ai_full_response = str(raw_resp)
            
            # Prompt-Text (kann String oder Dict sein bei zweistufig)
            prompt = ki_result.get('_prompt_text')
            if prompt is not None:
                import json as _json
                if isinstance(prompt, dict):
                    ai_prompt_text = _json.dumps(prompt, ensure_ascii=False)
                else:
                    ai_prompt_text = str(prompt)
            
            ai_model = ki_result.get('_ai_model')
            ai_stage = ki_result.get('_ai_stage')
            
            # Token-Verbrauch
            usage = ki_result.get('_usage', {})
            if usage:
                prompt_tokens = usage.get('prompt_tokens')
                completion_tokens = usage.get('completion_tokens')
                total_tokens = usage.get('total_tokens')
        
        # 4. Zeichenzaehler berechnen (schnelle Groessenanalyse ohne Text laden)
        text_char_count = len(extracted_text) if extracted_text else 0
        ai_response_char_count = len(ai_full_response) if ai_full_response else 0
        
        # 5. API-Call: POST /documents/{id}/ai-data
        data = {
            'extracted_text': extracted_text if (extracted_text and extracted_text.strip()) else None,
            'extracted_text_sha256': text_sha256,
            'extraction_method': extraction_method,
            'extracted_page_count': extracted_page_count or 0,
            'ai_full_response': ai_full_response,
            'ai_prompt_text': ai_prompt_text,
            'ai_model': ai_model,
            'ai_prompt_version': 'v2.0.2',  # Aktuelle Prompt-Version
            'ai_stage': ai_stage,
            'text_char_count': text_char_count if text_char_count > 0 else None,
            'ai_response_char_count': ai_response_char_count if ai_response_char_count > 0 else None,
            'prompt_tokens': prompt_tokens,
            'completion_tokens': completion_tokens,
            'total_tokens': total_tokens,
        }
        
        result = self.docs_api.save_ai_data(doc.id, data)
        if result:
            # Inhaltsduplikat-Info wird bereits von save_ai_data geloggt
            logger.debug(
                f"AI-Daten gespeichert fuer Dokument {doc.id} ({doc.original_filename}): "
                f"method={extraction_method}, pages={extracted_page_count}, "
                f"tokens={total_tokens}, stage={ai_stage}"
            )
        else:
            from i18n.de import AI_DATA_SAVE_FAILED
            logger.warning(AI_DATA_SAVE_FAILED.format(doc_id=doc.id))
    
    # ================================================================
    # Dokumenten-Regeln
    # ================================================================
    
    def _load_document_rules(self) -> None:
        """Laedt Dokumenten-Regeln vom Server (einmal pro Verarbeitungslauf)."""
        try:
            self._doc_rules = self.doc_rules_api.get_rules()
            if self._doc_rules and self._doc_rules.has_any_rule():
                logger.info(
                    f"Dokumenten-Regeln geladen: "
                    f"Datei-Dup={self._doc_rules.file_dup_action}, "
                    f"Content-Dup={self._doc_rules.content_dup_action}, "
                    f"Partial-Empty={self._doc_rules.partial_empty_action}, "
                    f"Full-Empty={self._doc_rules.full_empty_action}"
                )
            else:
                logger.debug("Dokumenten-Regeln: Keine aktiven Regeln konfiguriert")
        except Exception as e:
            logger.warning(f"Dokumenten-Regeln konnten nicht geladen werden: {e}")
            self._doc_rules = None
    
    def _apply_document_rules(self, doc: Document) -> None:
        """
        Wendet konfigurierte Dokumenten-Regeln an.
        
        Wird nach _persist_ai_data() aufgerufen. Zu diesem Zeitpunkt sind
        alle relevanten Informationen verfuegbar:
        - doc.is_duplicate / doc.previous_version_id (Datei-Duplikat)
        - doc.content_duplicate_of_id (Inhaltsduplikat, nach _persist_ai_data)
        - doc.empty_page_count / doc.total_page_count (nach _check_and_log_empty_pages)
        """
        if not self._doc_rules or not self._doc_rules.has_any_rule():
            return
        
        rules = self._doc_rules
        
        # Dokument-Daten aktualisieren (content_duplicate_of_id wird erst bei
        # _persist_ai_data gesetzt, doc-Objekt hat evtl. noch den alten Wert)
        try:
            fresh_doc_data = self.docs_api.get_document(doc.id)
            if fresh_doc_data:
                doc = fresh_doc_data
        except Exception:
            pass
        
        # 1. Komplett leere Datei
        if doc.is_completely_empty:
            if rules.full_empty_action == 'delete':
                logger.info(f"Dokumenten-Regel: Komplett leere Datei {doc.id} wird geloescht")
                self.docs_api.delete_documents([doc.id])
                return
            elif rules.full_empty_action == 'color_file' and rules.full_empty_color:
                logger.info(f"Dokumenten-Regel: Komplett leere Datei {doc.id} wird markiert ({rules.full_empty_color})")
                self.docs_api.set_document_color(doc.id, rules.full_empty_color)
        
        # 2. Teilweise leere Seiten
        elif doc.has_empty_pages and not doc.is_completely_empty:
            if rules.partial_empty_action == 'remove_pages':
                logger.info(f"Dokumenten-Regel: Leere Seiten entfernen bei Dokument {doc.id}")
                self._remove_empty_pages(doc)
            elif rules.partial_empty_action == 'color_file' and rules.partial_empty_color:
                logger.info(f"Dokumenten-Regel: Datei {doc.id} mit leeren Seiten markiert ({rules.partial_empty_color})")
                self.docs_api.set_document_color(doc.id, rules.partial_empty_color)
        
        # 3. Datei-Duplikat (gleiche SHA256-Pruefsumme)
        if doc.is_duplicate and doc.previous_version_id:
            self._apply_duplicate_rule(
                doc, rules.file_dup_action, rules.file_dup_color,
                doc.previous_version_id, 'Datei-Duplikat')
        
        # 4. Inhaltsduplikat (gleicher Text-Hash)
        if doc.is_content_duplicate and doc.content_duplicate_of_id:
            self._apply_duplicate_rule(
                doc, rules.content_dup_action, rules.content_dup_color,
                doc.content_duplicate_of_id, 'Inhaltsduplikat')
    
    def _apply_duplicate_rule(self, doc: Document, action: str, color: Optional[str],
                              original_id: int, rule_type: str) -> None:
        """Wendet eine Duplikat-Regel auf ein Dokument an."""
        if action == 'none':
            return
        
        if action == 'color_both' and color:
            logger.info(f"Dokumenten-Regel: {rule_type} - Beide markieren ({color}): {doc.id} + {original_id}")
            self.docs_api.set_documents_color([doc.id, original_id], color)
        
        elif action == 'color_new' and color:
            logger.info(f"Dokumenten-Regel: {rule_type} - Neue Datei markieren ({color}): {doc.id}")
            self.docs_api.set_document_color(doc.id, color)
        
        elif action == 'delete_new':
            logger.info(f"Dokumenten-Regel: {rule_type} - Neue Datei loeschen: {doc.id}")
            self.docs_api.delete_documents([doc.id])
        
        elif action == 'delete_old':
            logger.info(f"Dokumenten-Regel: {rule_type} - Alte Datei loeschen: {original_id}")
            self.docs_api.delete_documents([original_id])
    
    def _remove_empty_pages(self, doc: Document) -> None:
        """
        Entfernt leere Seiten aus dem PDF und laedt es neu hoch.
        
        Laed das Dokument herunter, entfernt leere Seiten mit PyMuPDF,
        speichert die bereinigte Version und ersetzt die Datei auf dem Server.
        """
        try:
            import fitz
        except ImportError:
            logger.warning("PyMuPDF nicht installiert, kann leere Seiten nicht entfernen")
            return
        
        from services.empty_page_detector import get_empty_pages
        
        with tempfile.TemporaryDirectory() as tmpdir:
            local_path = self.docs_api.download(doc.id, tmpdir)
            if not local_path:
                logger.warning(f"Dokument {doc.id} konnte nicht heruntergeladen werden")
                return
            
            empty_indices, total = get_empty_pages(local_path)
            if not empty_indices or len(empty_indices) >= total:
                return
            
            fitz_doc = fitz.open(local_path)
            for idx in sorted(empty_indices, reverse=True):
                fitz_doc.delete_page(idx)
            
            cleaned_path = os.path.join(tmpdir, 'cleaned.pdf')
            fitz_doc.save(cleaned_path, garbage=4, deflate=True)
            fitz_doc.close()
            
            self.docs_api.replace_document_file(doc.id, cleaned_path)
            
            new_total = total - len(empty_indices)
            try:
                self.docs_api.client.put(
                    f'/documents/{doc.id}',
                    json_data={'empty_page_count': 0, 'total_page_count': new_total})
            except Exception:
                logger.debug(f"Leere-Seiten-Zaehler Update fehlgeschlagen fuer {doc.id}")
            
            # Vorschau-Cache invalidieren (persistiert ueber App-Neustarts)
            try:
                import glob as _glob
                cache_dir = os.path.join(tempfile.gettempdir(), 'bipro_preview_cache')
                for cached in _glob.glob(os.path.join(cache_dir, f"{doc.id}_*")):
                    os.unlink(cached)
                    logger.debug(f"Vorschau-Cache invalidiert: {cached}")
            except Exception:
                pass
            
            logger.info(
                f"Dokument {doc.id}: {len(empty_indices)} leere Seiten entfernt "
                f"({total} -> {new_total} Seiten)"
            )

    def _slugify(self, text: str) -> str:
        """
        Konvertiert Text in sicheren Dateinamen.
        
        Ersetzt Umlaute und entfernt Sonderzeichen.
        """
        if not text:
            return "unbekannt"
        
        # Umlaute ersetzen
        replacements = {
            'ä': 'ae', 'ö': 'oe', 'ü': 'ue',
            'Ä': 'Ae', 'Ö': 'Oe', 'Ü': 'Ue',
            'ß': 'ss'
        }
        for old, new in replacements.items():
            text = text.replace(old, new)
        
        # Nur alphanumerische Zeichen und Unterstriche
        import re
        text = re.sub(r'[^a-zA-Z0-9_]', '_', text)
        text = re.sub(r'_+', '_', text)  # Mehrfache Unterstriche zusammenfassen
        text = text.strip('_')
        
        return text or "unbekannt"
    
    def _log_history(self,
                     document_id: int,
                     action: str,
                     new_status: str,
                     previous_status: Optional[str] = None,
                     success: bool = True,
                     error_message: Optional[str] = None,
                     classification_source: Optional[str] = None,
                     classification_result: Optional[str] = None,
                     action_details: Optional[dict] = None,
                     duration_ms: Optional[int] = None) -> None:
        """
        Protokolliert einen Verarbeitungsschritt in der History.
        
        Fehler beim Logging werden ignoriert, um die Verarbeitung nicht zu unterbrechen.
        """
        try:
            self.history_api.create(
                document_id=document_id,
                action=action,
                new_status=new_status,
                previous_status=previous_status,
                success=success,
                error_message=error_message,
                classification_source=classification_source,
                classification_result=classification_result,
                action_details=action_details,
                duration_ms=duration_ms
            )
        except Exception as e:
            # Fehler beim History-Logging sollten die Verarbeitung nicht stoppen
            logger.warning(f"History-Logging fehlgeschlagen fuer Dokument {document_id}: {e}")
    
    def _is_bipro_courtage(self, doc: Document) -> bool:
        """
        Prueft ob das Dokument per BiPRO-Code als Courtage klassifiziert ist.
        
        Verwendet die Konfiguration aus config/processing_rules.py
        
        Args:
            doc: Dokument mit bipro_category Attribut
            
        Returns:
            True wenn BiPRO-Code eine Provisionsabrechnung markiert
        """
        from config.processing_rules import is_bipro_courtage_code, BIPRO_COURTAGE_CODES
        
        bipro_category = getattr(doc, 'bipro_category', None)
        
        # Debug-Logging für BiPRO-Kategorien
        if bipro_category:
            is_courtage = is_bipro_courtage_code(bipro_category)
            logger.debug(
                f"BiPRO-Kategorie: {bipro_category} -> "
                f"{'Courtage' if is_courtage else 'NICHT Courtage'} "
                f"(erlaubte Codes: {BIPRO_COURTAGE_CODES})"
            )
            return is_courtage
        else:
            logger.debug(f"Dokument {doc.id} hat keine BiPRO-Kategorie")
            return False
    
    def _is_bipro_gdv(self, doc: Document) -> bool:
        """
        Prueft ob das Dokument per BiPRO-Code als GDV-Datensatz klassifiziert ist.
        
        GDV-Dateien koennen vom BiPRO mit .pdf Endung geliefert werden,
        sind aber tatsaechlich GDV-Bestandsdaten (999er-Codes).
        
        Args:
            doc: Dokument mit bipro_category Attribut
            
        Returns:
            True wenn BiPRO-Code eine GDV-Datei markiert (999xxx)
        """
        from config.processing_rules import is_bipro_gdv_code
        
        bipro_category = getattr(doc, 'bipro_category', None)
        
        if bipro_category:
            is_gdv = is_bipro_gdv_code(bipro_category)
            if is_gdv:
                logger.debug(f"BiPRO-Code {bipro_category} markiert GDV-Datei")
            return is_gdv
        
        return False
    
    def _classify_document(self, doc: Document) -> Tuple[str, str]:
        """
        Klassifiziert ein Dokument basierend auf Typ und Dateiendung.
        
        Bei unbekannten Endungen wird der Dateiinhalt geprueft (Magic-Bytes).
        
        Returns:
            (target_box, category)
        """
        from config.processing_rules import PROCESSING_RULES
        
        filename = doc.original_filename.lower()
        extension = doc.file_extension.lower()
        
        # Bekannte Endungen aus Konfiguration
        known_extensions = PROCESSING_RULES.get('known_extensions', ['.pdf', '.xml', '.gdv', '.txt', ''])
        
        # 1. XML-Rohdateien -> Roh Archiv
        if self._is_xml_raw(doc):
            return ('roh', 'xml_raw')
        
        # 2. GDV-Dateien -> GDV Box
        if self._is_gdv_file(doc):
            return ('gdv', 'gdv')
        
        # 3. PDFs -> muessen von KI klassifiziert werden
        if doc.is_pdf:
            return ('sonstige', 'pdf_pending')
        
        # 4. Unbekannte Endung -> Content-Erkennung (Magic-Bytes)
        if extension not in known_extensions:
            detected_type = self._detect_file_type_by_content(doc)
            
            if detected_type:
                logger.info(f"Unbekannte Endung '{extension}', erkannt als: {detected_type}")
                
                if detected_type == 'gdv':
                    # GDV erkannt -> umbenennen und in GDV Box
                    self._rename_with_extension(doc, '.gdv')
                    return ('gdv', 'gdv_detected')
                    
                elif detected_type == 'pdf':
                    # PDF erkannt -> umbenennen und zur KI
                    self._rename_with_extension(doc, '.pdf')
                    return ('sonstige', 'pdf_pending')
                    
                elif detected_type == 'xml':
                    # XML erkannt -> pruefen ob Roh, sonst Sonstige
                    self._rename_with_extension(doc, '.xml')
                    # Nochmal XML-Roh Check mit neuer Endung
                    if 'roh' in filename:
                        return ('roh', 'xml_raw_detected')
                    return ('sonstige', 'xml_detected')
        
        # 5. Andere Dateien -> Sonstige
        return ('sonstige', 'unknown')
    
    def _detect_file_type_by_content(self, doc: Document) -> Optional[str]:
        """
        Erkennt den Dateityp anhand der Magic-Bytes (ersten Bytes).
        
        Sehr leistungssparend: Liest nur die ersten 256 Bytes.
        
        Args:
            doc: Das zu pruefende Dokument
            
        Returns:
            'pdf', 'xml', 'gdv' oder None wenn nicht erkannt
        """
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                local_path = self.docs_api.download(doc.id, tmpdir)
                
                if not local_path or not os.path.exists(local_path):
                    logger.warning(f"Content-Detection: Download fehlgeschlagen fuer {doc.id}")
                    return None
                
                # Nur erste 256 Bytes lesen (sehr schnell)
                with open(local_path, 'rb') as f:
                    first_bytes = f.read(256)
                
                if not first_bytes:
                    return None
                
                # PDF: Magic-Bytes "%PDF"
                if first_bytes.startswith(b'%PDF'):
                    logger.debug(f"Magic-Bytes: PDF erkannt fuer {doc.original_filename}")
                    return 'pdf'
                
                # XML: Beginnt mit "<?xml" oder "<" (nach Whitespace)
                text_start = first_bytes.lstrip()
                if text_start.startswith(b'<?xml') or (text_start.startswith(b'<') and b'>' in text_start):
                    logger.debug(f"Magic-Bytes: XML erkannt fuer {doc.original_filename}")
                    return 'xml'
                
                # GDV: Erste Zeile beginnt mit "0001" (Vorsatz)
                # Verschiedene Encodings versuchen
                for encoding in ['cp1252', 'latin-1', 'utf-8']:
                    try:
                        first_line = first_bytes.decode(encoding).strip()
                        if first_line.startswith('0001'):
                            logger.debug(f"Magic-Bytes: GDV erkannt fuer {doc.original_filename}")
                            return 'gdv'
                        break  # Encoding funktioniert, aber kein GDV
                    except UnicodeDecodeError:
                        continue
                
                logger.debug(f"Dateityp nicht erkannt fuer {doc.original_filename}")
                return None
                
        except Exception as e:
            logger.warning(f"Content-Detection fehlgeschlagen fuer {doc.id}: {e}")
            return None
    
    def _rename_with_extension(self, doc: Document, new_extension: str) -> bool:
        """
        Benennt ein Dokument um und fuegt die korrekte Endung hinzu.
        
        Args:
            doc: Das umzubenennende Dokument
            new_extension: Die neue Dateiendung (z.B. '.pdf', '.gdv')
            
        Returns:
            True bei Erfolg
        """
        try:
            # Alte Endung entfernen falls vorhanden
            base_name = doc.original_filename
            if '.' in base_name:
                base_name = base_name.rsplit('.', 1)[0]
            
            new_filename = base_name + new_extension
            
            logger.info(f"Umbenennung: {doc.original_filename} -> {new_filename}")
            
            success = self.docs_api.rename_document(doc.id, new_filename)
            return success
            
        except Exception as e:
            logger.warning(f"Umbenennung fehlgeschlagen fuer {doc.id}: {e}")
            return False
    
    def _is_xml_raw(self, doc: Document) -> bool:
        """Prueft ob es sich um eine XML-Rohdatei handelt."""
        from config.processing_rules import PROCESSING_RULES
        
        filename = doc.original_filename
        
        # Pattern-Check
        for pattern in PROCESSING_RULES.get('raw_xml_patterns', []):
            # Einfacher Wildcard-Match
            if pattern.startswith('*'):
                if filename.endswith(pattern[1:]):
                    return True
            elif pattern.endswith('*'):
                if filename.startswith(pattern[:-1]):
                    return True
            elif '*' in pattern:
                prefix, suffix = pattern.split('*', 1)
                if filename.startswith(prefix) and filename.endswith(suffix):
                    return True
            elif filename == pattern:
                return True
        
        # Fallback: XML + "Roh" im Namen
        if doc.is_xml and 'roh' in filename.lower():
            return True
        
        return False
    
    def _is_spreadsheet(self, doc: Document) -> bool:
        """Prueft ob es sich um eine tabellarische Datei handelt (CSV, TSV, Excel)."""
        ext = doc.file_extension.lower()
        return ext in ['.csv', '.tsv', '.xlsx', '.xls']
    
    
    def _extract_spreadsheet_text(self, file_path: str, max_lines: int = 50) -> str:
        """
        Extrahiert Text aus einer Tabellendatei fuer KI-Klassifikation.
        
        Args:
            file_path: Pfad zur Datei
            max_lines: Maximale Anzahl Zeilen
            
        Returns:
            Extrahierter Text (Zeilen getrennt durch Newline)
        """
        ext = os.path.splitext(file_path)[1].lower()
        
        if ext in ['.csv', '.tsv']:
            # CSV/TSV: Als Text lesen mit Encoding-Fallback
            for encoding in ['utf-8', 'cp1252', 'latin-1']:
                try:
                    with open(file_path, 'r', encoding=encoding) as f:
                        lines = []
                        for i, line in enumerate(f):
                            if i >= max_lines:
                                break
                            lines.append(line.rstrip())
                        return '\n'.join(lines)
                except (UnicodeDecodeError, UnicodeError):
                    continue
            return ''
        
        elif ext == '.xlsx':
            # Excel: Erste Zeilen mit openpyxl
            try:
                from openpyxl import load_workbook
                wb = load_workbook(file_path, read_only=True, data_only=True)
                ws = wb.active
                lines = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= max_lines:
                        break
                    cells = [str(c) if c is not None else '' for c in row]
                    lines.append(' | '.join(cells))
                wb.close()
                return '\n'.join(lines)
            except Exception as e:
                logger.warning(f"Excel-Extraktion fehlgeschlagen: {e}")
                return ''
        
        return ''
    
    def _is_gdv_file(self, doc: Document) -> bool:
        """
        Prueft ob es sich um eine GDV-Datei handelt.
        
        Logik:
        1. Bereits als GDV markiert -> True
        2. Eindeutige Endung (.gdv) -> True
        3. Ambige Endung (.txt, keine) -> Content-Pruefung (erste Zeile = '0001')
        """
        from config.processing_rules import PROCESSING_RULES
        
        # Bereits als GDV markiert
        if doc.is_gdv:
            return True
        
        extension = doc.file_extension.lower()
        
        # Eindeutige GDV-Endungen (keine Content-Pruefung noetig)
        gdv_extensions = PROCESSING_RULES.get('gdv_extensions', ['.gdv'])
        if extension in gdv_extensions:
            return True
        
        # Ambige Endungen -> Content-Pruefung
        content_check_extensions = PROCESSING_RULES.get('gdv_content_check_extensions', ['.txt', ''])
        if extension in content_check_extensions:
            return self._check_gdv_content(doc)
        
        return False
    
    def _check_gdv_content(self, doc: Document) -> bool:
        """
        Prueft ob der Dateiinhalt eine GDV-Datei ist.
        
        GDV-Dateien beginnen IMMER mit Satzart '0001' (Vorsatz).
        Laedt nur die ersten 256 Bytes zur Pruefung.
        
        WICHTIG: Prueft ZUERST auf PDF-Magic-Bytes, um False-Positives zu vermeiden!
        
        Args:
            doc: Das zu pruefende Dokument
            
        Returns:
            True wenn erste Zeile mit '0001' beginnt UND KEINE PDF-Datei
        """
        try:
            # Dokument herunterladen (nur fuer Content-Check)
            with tempfile.TemporaryDirectory() as tmpdir:
                local_path = self.docs_api.download(doc.id, tmpdir)
                
                if not local_path or not os.path.exists(local_path):
                    logger.warning(f"Content-Check: Download fehlgeschlagen fuer {doc.id}")
                    return False
                
                # Nur erste 256 Bytes lesen (GDV-Zeilen sind 256 Zeichen)
                with open(local_path, 'rb') as f:
                    first_bytes = f.read(256)
                
                if not first_bytes:
                    logger.debug(f"Content-Check: Datei ist leer fuer {doc.id}")
                    return False
                
                # KRITISCH: PDF-Check VOR GDV-Check!
                # PDFs koennen zufaellig "0001" in den ersten Bytes enthalten
                if first_bytes.startswith(b'%PDF'):
                    logger.debug(f"Content-Check: PDF erkannt (nicht GDV) fuer {doc.original_filename}")
                    return False
                
                # Verschiedene Encodings versuchen
                for encoding in ['cp1252', 'latin-1', 'utf-8']:
                    try:
                        first_line = first_bytes.decode(encoding).strip()
                        break
                    except UnicodeDecodeError:
                        continue
                else:
                    # Kein Encoding funktioniert
                    logger.debug(f"Content-Check: Encoding fehlgeschlagen fuer {doc.id}")
                    return False
                
                # GDV-Dateien beginnen mit Satzart 0001 (Vorsatz)
                is_gdv = first_line.startswith('0001')
                logger.debug(f"Content-Check fuer {doc.original_filename}: is_gdv={is_gdv}")
                return is_gdv
                
        except Exception as e:
            logger.warning(f"Content-Check fehlgeschlagen fuer {doc.id}: {e}")
            return False
    
    def _extract_gdv_metadata(self, filepath: str) -> Tuple[Optional[str], Optional[str], Optional[str]]:
        """
        Extrahiert VU-Nummer, Absender und Datum aus GDV-Datensatz.
        
        Keine KI noetig - liest direkt aus dem Vorsatz (Satzart 0001):
        - VU-Nummer: Position 5-9 (5 Zeichen)
        - Absender: Position 10-39 (30 Zeichen) - Versicherer-Name
        - Datum: Position 70-77 (erstellungsdatum_von, 8 Zeichen, TTMMJJJJ)
        
        Bei Fehlern werden definierte Fallback-Werte verwendet:
        - Xvu: Unbekannter Versicherer
        - kDatum: Kein Datum gefunden
        
        Args:
            filepath: Pfad zur GDV-Datei
            
        Returns:
            (vu_nummer, absender, datum_iso) - mit Fallback-Werten bei Fehlern
        """
        from config.processing_rules import GDV_FALLBACK_VU, GDV_FALLBACK_DATE
        
        try:
            # Verschiedene Encodings versuchen (GDV ist meist CP1252)
            for encoding in ['cp1252', 'latin-1', 'utf-8']:
                try:
                    with open(filepath, 'r', encoding=encoding) as f:
                        for line in f:
                            # Suche nach Satzart 0001 (Vorsatz)
                            if len(line) >= 77 and line[0:4] == '0001':
                                # VU-Nummer: Position 5-9 (0-basiert: 4-9)
                                vu_nummer = line[4:9].strip()
                                
                                # Absender: Position 10-39 (0-basiert: 9-39)
                                absender = line[9:39].strip() if len(line) >= 39 else None
                                
                                # Datum: Position 70-77 (0-basiert: 69-77)
                                datum_raw = line[69:77].strip()
                                
                                # TTMMJJJJ -> YYYY-MM-DD konvertieren
                                datum_iso = None
                                if len(datum_raw) == 8 and datum_raw.isdigit():
                                    tag = datum_raw[0:2]
                                    monat = datum_raw[2:4]
                                    jahr = datum_raw[4:8]
                                    datum_iso = f"{jahr}-{monat}-{tag}"
                                
                                # Fallback-Werte anwenden wenn noetig
                                if not vu_nummer and not absender:
                                    vu_nummer = GDV_FALLBACK_VU
                                    logger.warning(f"GDV: VU fehlt, verwende Fallback '{GDV_FALLBACK_VU}'")
                                
                                if not datum_iso:
                                    datum_iso = GDV_FALLBACK_DATE
                                    logger.warning(f"GDV: Datum fehlt, verwende Fallback '{GDV_FALLBACK_DATE}'")
                                
                                logger.debug(f"GDV-Metadaten: VU={vu_nummer}, Absender={absender}, Datum={datum_iso}")
                                return (vu_nummer, absender, datum_iso)
                    break  # Encoding funktioniert, aber kein Vorsatz gefunden
                except UnicodeDecodeError:
                    continue
            
            # Kein Vorsatz gefunden - komplettes Fallback
            logger.warning(f"Kein GDV-Vorsatz (0001) gefunden in {filepath}, verwende Fallback-Werte")
            return (GDV_FALLBACK_VU, None, GDV_FALLBACK_DATE)
            
        except Exception as e:
            logger.warning(f"GDV-Metadaten-Extraktion fehlgeschlagen: {e}, verwende Fallback-Werte")
            return (GDV_FALLBACK_VU, None, GDV_FALLBACK_DATE)
    
    def _classify_pdf_with_ai(self, doc: Document) -> Optional[DocumentClassification]:
        """
        Klassifiziert ein PDF mit dem zweistufigen KI-System.
        
        Stufe 1 (Triage): Schnelle Kategorisierung mit GPT-4o-mini
        Stufe 2 (Detail): Nur bei courtage/versicherung mit GPT-4o
        
        Bei 'sonstige' in Stufe 1 wird KEINE teure Detailanalyse gemacht.
        
        Args:
            doc: Das zu klassifizierende Dokument
            
        Returns:
            DocumentClassification oder None bei Fehler
        """
        try:
            openrouter = self._get_openrouter()
            
            # Dokument herunterladen
            with tempfile.TemporaryDirectory() as tmpdir:
                local_path = self.docs_api.download(doc.id, tmpdir)
                
                if not local_path:
                    logger.warning(f"Download fehlgeschlagen fuer Dokument {doc.id}")
                    return None
                
                # Zweistufige KI-Klassifikation (Triage -> Detail bei Bedarf)
                classification = openrouter.classify_pdf_smart(local_path)
                
                return classification
                
        except Exception as e:
            logger.error(f"KI-Klassifikation fehlgeschlagen: {e}")
            return None
    
    def _is_courtage_document(self, extracted: ExtractedDocumentData) -> bool:
        """Prueft ob das Dokument eine Courtage/Provisionsabrechnung ist."""
        from config.processing_rules import PROCESSING_RULES
        
        # Schluesselwoerter fuer Courtage
        courtage_keywords = PROCESSING_RULES.get('courtage_keywords', [])
        
        # Versicherungstyp pruefen
        doc_type = (extracted.versicherungstyp or '').lower()
        
        for keyword in courtage_keywords:
            if keyword.lower() in doc_type:
                return True
        
        return False
    
    def _is_leben_category(self, doc_type: str) -> bool:
        """Prueft ob der Dokumenttyp zur Leben-Kategorie gehoert."""
        leben_keywords = [
            'leben', 'life', 'rente', 'pension', 'altersvorsorge',
            'berufsunfähigkeit', 'bu', 'risiko', 'kapital', 'fond'
        ]
        return any(kw in doc_type for kw in leben_keywords)
    
    def _is_sach_category(self, doc_type: str) -> bool:
        """Prueft ob der Dokumenttyp zur Sach-Kategorie gehoert."""
        sach_keywords = [
            'sach', 'haftpflicht', 'hausrat', 'wohngebäude', 'kfz',
            'auto', 'unfall', 'rechtsschutz', 'glas', 'elektronik',
            'transport', 'gewerbe', 'betrieb'
        ]
        return any(kw in doc_type for kw in sach_keywords)
    
    def process_single_document(self, doc_id: int) -> ProcessingResult:
        """
        Verarbeitet ein einzelnes Dokument (manueller Trigger).
        
        Args:
            doc_id: ID des zu verarbeitenden Dokuments
            
        Returns:
            ProcessingResult
        """
        doc = self.docs_api.get_document(doc_id)
        if not doc:
            return ProcessingResult(
                document_id=doc_id,
                original_filename='',
                success=False,
                target_box='',
                error='Dokument nicht gefunden'
            )
        
        return self._process_document(doc)
    
    def classify_document_preview(self, doc: Document) -> Tuple[str, str]:
        """
        Zeigt Vorschau der Klassifikation ohne Aenderungen.
        
        Returns:
            (target_box, category)
        """
        return self._classify_document(doc)

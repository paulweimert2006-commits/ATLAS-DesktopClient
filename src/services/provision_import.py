"""
Import-Service fuer Provisionsmanagement.

Parst VU-Provisionslisten (Allianz, SwissLife, VB) und Xempus-Exporte
und bereitet die Daten fuer den Server-Import auf.

Normalisierungsfunktionen sind in domain/provision/normalization.py
definiert und werden hier re-exportiert fuer Abwaertskompatibilitaet.
"""

import re
import hashlib
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass, field
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl nicht installiert -- Excel-Import nicht verfuegbar")

from domain.provision.normalization import (  # noqa: F401, E402 – Re-Export
    normalize_vsnr,
    normalize_vermittler_name,
    normalize_for_db,
    normalize_vb_name as _normalize_vb_name,
)


def _compute_row_hash(vu_name: str, vsnr: str, betrag: float, datum: str, art: str = '') -> str:
    """Eindeutiger Hash pro Provisions-Zeile fuer Duplikat-Schutz."""
    raw = f"{vu_name}|{vsnr}|{betrag:.2f}|{datum or ''}|{art}"
    return hashlib.sha256(raw.encode('utf-8')).hexdigest()


def _parse_amount(val) -> Optional[float]:
    """Betrag aus Excel-Zelle parsen (deutsch/englisch)."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s == '-':
        return None
    s = s.replace(' ', '').replace('€', '').replace('EUR', '')
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def _parse_date(val) -> Optional[str]:
    """Datum aus Excel-Zelle parsen → YYYY-MM-DD."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s:
        return None
    for fmt in ('%d.%m.%Y', '%d.%m.%y', '%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


def _cell_val(ws, row: int, col: int):
    """Sicherer Zellzugriff (1-basiert fuer Spalte)."""
    cell = ws.cell(row=row, column=col)
    return cell.value


def _col_index(letter: str) -> int:
    """Spaltenbuchstabe → 1-basierter Index (A=1, B=2, ..., AG=33, AV=48, BC=55)."""
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result


@dataclass
class ParseResult:
    """Ergebnis eines Sheet-Parsings."""
    rows: List[Dict] = field(default_factory=list)
    sheet_name: str = ''
    vu_name: str = ''
    total_rows: int = 0
    skipped_rows: int = 0
    errors: List[str] = field(default_factory=list)


# ═══════════════════════════════════════════════════════
# VU-LISTEN COLUMN-MAPPINGS (aus VBA-Makro abgeleitet)
# ═══════════════════════════════════════════════════════

VU_COLUMN_MAPPINGS = {
    'Allianz': {
        'vsnr_col': 'A',
        'betrag_col': 'D',
        'art_col': 'F',
        'datum_col': 'G',
        'courtage_rate_col': 'K',
        'vn_col': 'AE',
        'konditionssatz_col': None,
    },
    'SwissLife': {
        'vsnr_col': 'Y',
        'betrag_col': 'N',
        'art_col': 'O',
        'datum_col': 'C',
        'courtage_rate_col': None,
        'vn_col': 'U',
        'konditionssatz_col': None,
    },
    'VB': {
        'vsnr_col': 'B',
        'betrag_col': 'O',
        'lastschrift_col': 'P',
        'art_col': 'K',
        'datum_col': 'AR',
        'courtage_rate_col': None,
        'vn_col': 'C',
        'konditionssatz_col': 'M',
    },
}

VU_HEADER_SIGNATURES = {
    'Allianz': {
        'sheet_names': ['allianz'],
        'header_keywords': ['vtnr', 'provisions-betrag', 'courtagesatz', 'auszahlungs-datum'],
    },
    'SwissLife': {
        'sheet_names': ['swisslife', 'swiss life'],
        'header_keywords': ['versicherungsnummer', 'buchwert', 'abrechnungsnummer', 'konditionssatz'],
    },
    'VB': {
        'sheet_names': ['vb', 'volkswohlbund'],
        'header_keywords': ['vart', 'gutschrift', 'lastschrift', 'abrechnung von'],
    },
}


def _map_art(raw: str, vu_name: str) -> str:
    """Provisions-Art aus VU-spezifischem Wert in Enum-Wert mappen."""
    s = str(raw or '').strip().upper()
    if not s:
        return 'ap'
    if s in ('RB', 'ST', 'STORNO', 'RÜCK', 'RUECK', 'RÜCKBELASTUNG'):
        return 'rueckbelastung'
    if s in ('BP', 'FP', 'FOLGEPROV', 'BESTANDSPROV', 'BEST'):
        return 'bp'
    if s in ('AP', 'EV', 'EV-PF', 'ABSCHL', 'ABSCHLUSSPROV'):
        return 'ap'
    return 'sonstige'


def parse_vu_sheet(wb, sheet_name: str) -> ParseResult:
    """Ein einzelnes VU-Sheet parsen (iter_rows fuer Performance)."""
    result = ParseResult(sheet_name=sheet_name, vu_name=sheet_name)

    if sheet_name not in wb.sheetnames:
        result.errors.append(f"Sheet '{sheet_name}' nicht gefunden")
        return result

    mapping = VU_COLUMN_MAPPINGS.get(sheet_name)
    if not mapping:
        result.errors.append(f"Kein Column-Mapping fuer '{sheet_name}'")
        return result

    ws = wb[sheet_name]

    vsnr_idx = _col_index(mapping['vsnr_col'])
    betrag_idx = _col_index(mapping['betrag_col']) if mapping['betrag_col'] else None
    lastschrift_idx = _col_index(mapping['lastschrift_col']) if mapping.get('lastschrift_col') else None
    art_idx = _col_index(mapping['art_col']) if mapping['art_col'] else None
    datum_idx = _col_index(mapping['datum_col']) if mapping['datum_col'] else None
    courtage_rate_idx = _col_index(mapping['courtage_rate_col']) if mapping.get('courtage_rate_col') else None
    vn_idx = _col_index(mapping['vn_col']) if mapping.get('vn_col') else None
    konditionssatz_idx = _col_index(mapping['konditionssatz_col']) if mapping.get('konditionssatz_col') else None
    # --- Haupt-Parsing-Logik für VU-Tabellen (Allianz/SwissLife/VB) ---
    # Diese Schleife parst die bereitgestellte Excel-Tabelle Zeile für Zeile (ab Zeile 2, also nach dem Header).
    # Sie extrahiert alle relevanten Werte, normalisiert sie und fügt sie der Ergebnisliste hinzu.
    # Fehler, unmögliche oder unvollständige Zeilen werden übersprungen und protokolliert.

    if betrag_idx is None:
        # Ohne konfigurierten Betrag (Pflichtwert) macht das Parsen keinen Sinn.
        result.errors.append(f"{sheet_name}: Betrag-Spalte nicht konfiguriert")
        return result

    # Ermittelt die höchste benötigte Spaltennummer für iter_rows, abhängig davon, welche Felder in dem Mapping vorkommen.
    max_col_needed = max(filter(None, [vsnr_idx, betrag_idx, lastschrift_idx, art_idx, datum_idx, courtage_rate_idx, vn_idx, konditionssatz_idx]))

    logger.info(f"Parsing {sheet_name}: vsnr={vsnr_idx}, betrag={betrag_idx}, "
                f"art={art_idx}, datum={datum_idx}")

    row_num = 0
    for row in ws.iter_rows(min_row=2, max_col=max_col_needed, values_only=False):
        # Zählt echte Datenzeilen hoch (header = Zeile 1).
        row_num += 1
        result.total_rows += 1
        try:
            # Versicherungsnummer extrahieren, Pflichtfeld für Import.
            vsnr_cell = row[vsnr_idx - 1] if vsnr_idx - 1 < len(row) else None
            vsnr_raw = vsnr_cell.value if vsnr_cell else None
            if vsnr_raw is None or str(vsnr_raw).strip() == '':
                # Leere Zeile überspringen.
                result.skipped_rows += 1
                continue

            vsnr = str(vsnr_raw).strip()

            # Primärer Betrag. Kann positiv oder negativ sein.
            betrag_val = row[betrag_idx - 1].value if betrag_idx and betrag_idx - 1 < len(row) else None
            betrag = _parse_amount(betrag_val)
            if (betrag is None or betrag == 0) and lastschrift_idx:
                # Manche VUs benutzen separate Lastschriften für Rückbelastungen.
                ls_val = row[lastschrift_idx - 1].value if lastschrift_idx - 1 < len(row) else None
                ls_betrag = _parse_amount(ls_val)
                if ls_betrag is not None and ls_betrag != 0:
                    betrag = -abs(ls_betrag)
            if betrag is None or betrag == 0:
                # Ohne Betrag macht Import keinen Sinn.
                result.skipped_rows += 1
                continue

            # Provisionsart (AP, BP, Rückbelastung usw.) – VU-spezifisch kodiert, daher mit Mapping.
            art_val = row[art_idx - 1].value if art_idx and art_idx - 1 < len(row) else None
            art_raw = str(art_val or '')
            art = _map_art(art_raw, sheet_name)

            # Auszahlungs-/Buchungsdatum extrahieren und normalisieren.
            datum_val = row[datum_idx - 1].value if datum_idx and datum_idx - 1 < len(row) else None
            datum = _parse_date(datum_val)

            # Versicherungsnehmer-Name (optional).
            vn_val = row[vn_idx - 1].value if vn_idx and vn_idx - 1 < len(row) else None
            vn_name = str(vn_val or '').strip() if vn_val else None
            if vn_name and sheet_name == 'VB':
                # Volkswohlbund-Namen sind oft uneinheitlich, daher normalization notwendig.
                vn_name = _normalize_vb_name(vn_name)

            # Courtage-Satz (optional, z.B. bei SwissLife/Allianz für spätere Plausiprüfungen).
            courtage_rate = None
            if courtage_rate_idx and courtage_rate_idx - 1 < len(row):
                rate_val = row[courtage_rate_idx - 1].value
                courtage_rate = _parse_amount(rate_val)

            # Konditionssatz (nur SwissLife, optional).
            konditionssatz = None
            if konditionssatz_idx and konditionssatz_idx - 1 < len(row):
                kond_val = row[konditionssatz_idx - 1].value
                konditionssatz = str(kond_val).strip() if kond_val is not None else None

            # Spezielle Regel: Negative Beträge sind IMMER Rückbelastungen.
            if betrag < 0:
                art = 'rueckbelastung'

            # Hash der Zeile für Duplikatsschutz/Identifikation.
            row_hash = _compute_row_hash(sheet_name, vsnr, betrag, datum, art)

            # Ergebnisse an der Ergebnisliste anfügen. 
            # Felder, die für weitere Verarbeitung/Import nötig sind, werden gesetzt.
            result.rows.append({
                'vsnr': vsnr,
                'betrag': round(betrag, 2),
                'art': art,
                'buchungsart_raw': art_raw,
                'auszahlungsdatum': datum,
                'versicherungsnehmer': vn_name,
                'provisions_basissumme': None,  # Kann für späteres ETL noch befüllt werden.
                'rate_nummer': None,
                'rate_anzahl': None,
                'row_hash': row_hash,
                'courtage_rate': round(courtage_rate, 2) if courtage_rate is not None else None,
                'konditionssatz': konditionssatz,
                'source_row': row_num + 1,  # Quellzeile für Fehlerdiagnose.
            })
        except Exception as e:
            # Jede Exception in einer Zeile wird protokolliert, Zeile wird verworfen.
            result.errors.append(f"Zeile {row_num + 1}: {str(e)}")
            result.skipped_rows += 1

    logger.info(f"VU-Sheet '{sheet_name}': {len(result.rows)} Zeilen geparst, "
                f"{result.skipped_rows} uebersprungen, {len(result.errors)} Fehler")
    return result


# --- Parsen mehrerer VU-Sheets aus einer Excel-Provisionsliste ---
def parse_vu_liste(filepath: str, selected_sheets: List[str] = None,
                   on_progress=None) -> List[ParseResult]:
    """Alle relevanten VU-Sheets aus einer Excel-Datei parsen.

    Args:
        on_progress: Optional callback(msg: str) fuer Fortschrittsmeldungen.
    Returns:
        Liste von ParseResult-Objekten mit den geparsten Zeilen pro VU.
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl nicht installiert")

    import time
    t0 = time.time()

    def _log(msg):
        # Lokale Logging-Funktion – informiert auch evt. Callback für UI.
        logger.info(msg)
        if on_progress:
            on_progress(msg)

    _log(f"Oeffne Excel: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    _log(f"Geoeffnet in {time.time() - t0:.1f}s, Sheets: {wb.sheetnames}")

    # Entweder explizite Sheets oder alle konfigurierten Mappings abarbeiten.
    sheets_to_parse = selected_sheets or list(VU_COLUMN_MAPPINGS.keys())
    results = []

    for sheet_name in sheets_to_parse:
        if sheet_name in wb.sheetnames:
            t1 = time.time()
            _log(f"Parse Sheet '{sheet_name}'...")
            pr = parse_vu_sheet(wb, sheet_name)
            _log(f"Sheet '{sheet_name}' fertig in {time.time() - t1:.1f}s: "
                 f"{len(pr.rows)} Zeilen, {pr.skipped_rows} skip, {len(pr.errors)} err")
            results.append(pr)

    wb.close()
    _log(f"Gesamt: {time.time() - t0:.1f}s")
    return results


# --- Hilfsfunktion: Verfügbare, unterstützte VU-Sheets aus Excel bestimmen ---
def get_available_vu_sheets(filepath: str) -> List[str]:
    """Verfuegbare VU-Sheets in einer Excel-Datei ermitteln.
    Gibt nur die Sheets zurück, für die ein Mapping existiert.
    """
    if not HAS_OPENPYXL:
        return []
    wb = openpyxl.load_workbook(filepath, read_only=True)
    available = [s for s in wb.sheetnames if s in VU_COLUMN_MAPPINGS]
    wb.close()
    return available


# --- Detection-Logik: VU-Format einer Excel-Datei automatisch anhand Sheet-Name/Header raten ---
def detect_vu_format(filepath: str) -> List[Tuple[str, float]]:
    """VU-Format einer Excel-Datei anhand der Sheet-Namen und Header erkennen.

    Returns: Liste von (vu_name, confidence) Tupeln, absteigend nach Confidence.

    - Schritt 1: Sheetname exakte Übereinstimmung (confidence 1.0).
    - Schritt 2: Sheetname-Pattern-Match (confidence 0.9).
    - Schritt 3: Header-Keywords versuchen (zählt Treffer und gewichtet sie).
    """
    if not HAS_OPENPYXL:
        return []
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    results = []

    # Mapping: sheetname.lower() -> Originalsheetname
    sheets_lower = {s.lower(): s for s in wb.sheetnames}

    # 1. Step: Sheet-Namen mit Signature
    for vu_name, sig in VU_HEADER_SIGNATURES.items():
        if vu_name in wb.sheetnames:
            results.append((vu_name, 1.0))
            continue
        for sn_pattern in sig.get('sheet_names', []):
            if sn_pattern in sheets_lower:
                results.append((vu_name, 0.9))
                break

    # 2. Step: Falls keine Sheets erkannt, dann Header-Row parsen und nach Keywords suchen
    if not results:
        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                headers = []
                for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                    # Nur die allererste Tabellenzeile lesen.
                    headers = [str(v or '').strip().lower() for v in row]
                    break
                header_str = ' '.join(headers)
                for vu_name, sig in VU_HEADER_SIGNATURES.items():
                    kws = sig.get('header_keywords', [])
                    matches = sum(1 for kw in kws if kw in header_str)
                    if matches >= 2:
                        # Je mehr Treffer, desto größer das Vertrauen.
                        results.append((vu_name, 0.6 + 0.1 * matches))
                        break
            except Exception:
                continue

    wb.close()
    # Duplikate mit höchster Confidence entfernen
    results.sort(key=lambda x: x[1], reverse=True)
    seen = set()
    unique = []
    for vu, conf in results:
        if vu not in seen:
            seen.add(vu)
            unique.append((vu, conf))
    return unique


# --- Re-Exports: Normalisierungs- und Parser-Utils für Abwärtskompatibilität ---
from domain.provision.normalization import normalize_swisslife_vsnr  # noqa: F401, E402

from domain.provision.vu_parser import (  # noqa: F401, E402 – Re-Export
    compute_row_hash,
    parse_amount,
    parse_date,
    col_index,
    compute_file_hash as domain_compute_file_hash,
)
from domain.provision.relevance import (  # noqa: F401, E402 – Re-Export
    is_commission_relevant,
    classify_buchungsart,
)


# ═══════════════════════════════════════════════════════
# XEMPUS-PARSER
# Parsen und Handhabung von Xempus-Excel-Exporten
# ═══════════════════════════════════════════════════════

# Feste Spaltenangaben für IDs usw. (Excel-Buchstabenspalten, 1-basiert)
XEMPUS_ID_COL = 'AM'            # Xempus interne ID (Spalte AM)
XEMPUS_ARBN_ID_COL = 'AN'       # Arbeitnehmer-ID (Spalte AN)
XEMPUS_ARBG_ID_COL = 'AO'       # Arbeitgeber-ID (Spalte AO)
XEMPUS_STATUS_COL = 'E'         # Status der Beratung (Spalte E)
XEMPUS_VSNR_COL = 'O'           # Versicherungsscheinnummer (Spalte O)


def _detect_xempus_columns(ws, header_row: int = 1) -> Dict[str, Optional[int]]:
    """
    Sucht nach für den Parser relevanten Spalten anhand Schlüsselwörtern in der Kopfzeile
    und gibt ein Mapping zurück: Feldname → Spaltenindex (1-basiert).
    """
    mapping = {}
    # Schlüsselwörter für die typischen Xempus-Header (alle Kleinbuchstaben)
    keywords = {
        'versicherungsscheinnummer': ['versicherungsscheinnummer', 'vsnr', 'vertragsnummer'],
        'berater': ['berater'],
        'status': ['status'],
        'versicherer': ['versicherer', 'gesellschaft'],
        'gesamtbeitrag': ['gesamtbeitrag', 'beitrag'],
        'versicherungsnehmer': ['versicherungsnehmer', 'vn'],
        'arbn_name': ['arbn-name', 'arbeitnehmer name', 'arbeitnehmer-name', 'an-name'],
        'arbn_vorname': ['arbn-vorname', 'arbeitnehmer vorname', 'arbeitnehmer-vorname', 'an-vorname'],
        'arbeitgeber': ['arbg', 'arbeitgeber'],
        'sparte': ['sparte', 'produktgruppe'],
        'tarif': ['tarif', 'tarifname', 'produkt'],
        'beginn': ['beginn', 'vertragsbeginn', 'versicherungsbeginn'],
    }

    # Nur die Header-Zeile lesen (~erwartet: 1)
    for row in ws.iter_rows(min_row=header_row, max_row=header_row):
        for col_idx, cell in enumerate(row, start=1):
            header = str(cell.value or '').strip().lower()
            if not header:
                continue
            # Prüfe, ob eines der Schlüsselwörter im Header vorkommt
            for field_name, kw_list in keywords.items():
                if any(kw in header for kw in kw_list):
                    # Nur das erste Vorkommen je Feldname speichern
                    if field_name not in mapping:
                        mapping[field_name] = col_idx
                    break

    return mapping


def _map_xempus_status(raw: str) -> Optional[str]:
    """
    Xempus-Status auf internes Standard-Format mappen.
    Gibt z.B. "abgeschlossen", "storniert", "angebot", "beantragt", "offen" oder None zurück.
    None: Diese Zeile überspringen (z.B. "nicht gewünscht").
    """
    s = raw.strip().lower()
    if s in ('nicht gewünscht', 'nicht gewuenscht'):
        # Nicht gewünschte Beratung: nicht importieren
        return None
    if s in ('abgeschlossen', 'geschlossen', 'policiert'):
        return 'abgeschlossen'
    if s in ('storniert', 'storno'):
        return 'storniert'
    if s in ('angebot', 'angeboten'):
        return 'angebot'
    if s == 'beantragt':
        return 'beantragt'
    if s in ('unberaten', 'entscheidung ausstehend'):
        return 'offen'
    # Default-Fall: Status unbekannt, wird als "offen" betrachtet
    return 'offen'


def parse_xempus(filepath: str, sheet_name: str = 'Beratungen') -> ParseResult:
    """
    Parst den Xempus-Excel-Export für das Sheet 'Beratungen'.
    Gibt einen ParseResult mit allen extrahierten Datensätzen zurück.
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl nicht installiert")

    # Basisergebnis initialisieren
    result = ParseResult(sheet_name=sheet_name, vu_name='Xempus')

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        available = wb.sheetnames
        wb.close()
        result.errors.append(
            f"Sheet '{sheet_name}' nicht gefunden. "
            f"Verfuegbar: {', '.join(available)}"
        )
        return result

    ws = wb[sheet_name]
    columns = _detect_xempus_columns(ws)

    # Mapping von Feldname zu 0-basiertem Spaltenindex festlegen (None, falls Feld nicht vorhanden)
    vsnr_ci = (columns['versicherungsscheinnummer'] - 1) if 'versicherungsscheinnummer' in columns else None
    berater_ci = (columns['berater'] - 1) if 'berater' in columns else None
    status_ci = (columns['status'] - 1) if 'status' in columns else None
    versicherer_ci = (columns['versicherer'] - 1) if 'versicherer' in columns else None
    vn_ci = (columns['versicherungsnehmer'] - 1) if 'versicherungsnehmer' in columns else None
    arbn_name_ci = (columns['arbn_name'] - 1) if 'arbn_name' in columns else None
    arbn_vorname_ci = (columns['arbn_vorname'] - 1) if 'arbn_vorname' in columns else None
    arbg_ci = (columns['arbeitgeber'] - 1) if 'arbeitgeber' in columns else None
    sparte_ci = (columns.get('sparte', 0) - 1) if 'sparte' in columns else None
    tarif_ci = (columns.get('tarif', 0) - 1) if 'tarif' in columns else None
    beitrag_ci = (columns.get('gesamtbeitrag', 0) - 1) if 'gesamtbeitrag' in columns else None
    beginn_ci = (columns.get('beginn', 0) - 1) if 'beginn' in columns else None

    # Xempus interne IDs aus festen Spaltenbuchstaben
    xempus_id_ci = _col_index(XEMPUS_ID_COL) - 1
    arbn_id_ci = _col_index(XEMPUS_ARBN_ID_COL) - 1
    arbg_id_ci = _col_index(XEMPUS_ARBG_ID_COL) - 1

    # Ab Zeile 2 (nach Header) parsen
    for row in ws.iter_rows(min_row=2):
        result.total_rows += 1
        row_num = result.total_rows + 1  # Zeilennummer im Excel (inkl. Header)

        try:
            cells = list(row)
            max_ci = len(cells) - 1

            # Liefert String-Wert (oder ''), falls Spaltenindex gültig ist, sonst leeren String
            def _safe(ci):
                if ci is not None and ci <= max_ci:
                    return str(cells[ci].value or '').strip()
                return ''

            # Status prüfen/mappen (möglicherweise None → überspringen)
            status_raw = _safe(status_ci)
            status = _map_xempus_status(status_raw)
            if status is None:
                result.skipped_rows += 1
                continue

            # Versicherungsscheinnummer und Xempus-ID lesen
            vsnr_raw = cells[vsnr_ci].value if vsnr_ci is not None and vsnr_ci <= max_ci else None
            vsnr = str(vsnr_raw).strip() if vsnr_raw and str(vsnr_raw).strip() else ''
            xempus_id = _safe(xempus_id_ci) or None

            # Ohne VSNR überspringen (dient als Primary Key)
            if not vsnr:
                result.skipped_rows += 1
                continue

            berater = _safe(berater_ci)

            # Arbeitnehmer-Name zusammensetzen, sonst Versicherungsnehmer verwenden
            arbn_nachname = _safe(arbn_name_ci)
            arbn_vorname = _safe(arbn_vorname_ci)
            if arbn_nachname:
                vn_name = f"{arbn_nachname} {arbn_vorname}".strip() if arbn_vorname else arbn_nachname
            else:
                vn_name = _safe(vn_ci) or None

            # Zeilendatensatz zusammenstellen
            row_data = {
                'vsnr': vsnr if vsnr else None,
                'berater': berater,
                'status': status,
                'versicherer': _safe(versicherer_ci) or None,
                'versicherungsnehmer': vn_name,
                'sparte': _safe(sparte_ci) or None,
                'tarif': _safe(tarif_ci) or None,
                'beitrag': _parse_amount(cells[beitrag_ci].value) if beitrag_ci is not None and beitrag_ci <= max_ci else None,
                'beginn': _parse_date(cells[beginn_ci].value) if beginn_ci is not None and beginn_ci <= max_ci else None,
                'xempus_id': xempus_id,
                'arbn_id': _safe(arbn_id_ci) or None,
                'arbg_id': _safe(arbg_id_ci) or None,
            }
            result.rows.append(row_data)
        except Exception as e:
            # Fehlerhafte Zeile überspringen und Meldung speichern
            result.errors.append(f"Zeile {row_num}: {str(e)}")
            result.skipped_rows += 1

    wb.close()
    logger.info(f"Xempus '{sheet_name}': {len(result.rows)} Zeilen geparst, "
                f"{result.skipped_rows} uebersprungen")
    return result


def get_xempus_sheets(filepath: str) -> List[str]:
    """
    Liefert alle verfügbaren Sheetnamen in einer Xempus-Exceldatei.
    Rückgabe: Liste von Strings (Sheetnamen)
    """
    if not HAS_OPENPYXL:
        return []
    wb = openpyxl.load_workbook(filepath, read_only=True)
    sheets = list(wb.sheetnames)
    wb.close()
    return sheets


# Bekannte/standardisierte Sheetnamen in Xempus-Exporten
XEMPUS_KNOWN_SHEETS = [
    'ArbG',           # Arbeitgeber-Stammdaten
    'ArbG-Tarife',    # Arbeitgeber-Tarifdaten
    'ArbG-Zuschüsse', # Arbeitgeber-Zuschüsse/Beiträge
    'ArbN',           # Arbeitnehmer
    'Beratungen',     # Beratungsdaten
]


@dataclass
class XempusFullResult:
    """
    Kompletter Ergebnis-Container für alle Daten eines Xempus-Exports.
    Enthält:
        - beratungen: Alle Zeilen im Beratungen-Sheet (ParseResult)
        - sheets_found: Liste aller gefundenen Sheetnamen
        - sheet_row_counts: Row-Anzahl (Datenzeilen) pro Sheet
        - berater_names: Liste aller unterschiedlichen Beraternamen
    """
    beratungen: ParseResult
    sheets_found: List[str] = field(default_factory=list)
    sheet_row_counts: Dict[str, int] = field(default_factory=dict)
    berater_names: List[str] = field(default_factory=list)


def parse_xempus_full(filepath: str) -> XempusFullResult:
    """
    Parst einen vollständigen Xempus-Export (alle Excel-Sheets).
    Extrahiert insbesondere die Beratungen und zählt pro Sheet die Anzahl Datenzeilen.
    Liefert auch alle gefundenen Beraternamen.
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl nicht installiert")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheets_found = list(wb.sheetnames)

    # Zeilenzahlen je Sheet zählen (Datenzeilen OHNE Header)
    sheet_row_counts: Dict[str, int] = {}
    for sn in sheets_found:
        ws = wb[sn]
        row_count = (ws.max_row or 1) - 1  # Abzüglich Headerzeile
        sheet_row_counts[sn] = max(row_count, 0)

    beratungen = ParseResult(sheet_name='Beratungen', vu_name='Xempus')

    # Falls Sheet 'Beratungen' existiert, alle Zeilen wie oben parsen
    if 'Beratungen' in wb.sheetnames:
        ws = wb['Beratungen']
        columns = _detect_xempus_columns(ws)

        # Alle relevanten Spaltenindices (ggf. None)
        vsnr_ci = (columns['versicherungsscheinnummer'] - 1) if 'versicherungsscheinnummer' in columns else None
        berater_ci = (columns['berater'] - 1) if 'berater' in columns else None
        status_ci = (columns['status'] - 1) if 'status' in columns else None
        versicherer_ci = (columns['versicherer'] - 1) if 'versicherer' in columns else None
        vn_ci = (columns['versicherungsnehmer'] - 1) if 'versicherungsnehmer' in columns else None
        arbn_name_ci = (columns['arbn_name'] - 1) if 'arbn_name' in columns else None
        arbn_vorname_ci = (columns['arbn_vorname'] - 1) if 'arbn_vorname' in columns else None
        arbg_ci = (columns['arbeitgeber'] - 1) if 'arbeitgeber' in columns else None
        sparte_ci = (columns['sparte'] - 1) if 'sparte' in columns else None
        tarif_ci = (columns['tarif'] - 1) if 'tarif' in columns else None
        beitrag_ci = (columns['gesamtbeitrag'] - 1) if 'gesamtbeitrag' in columns else None
        beginn_ci = (columns['beginn'] - 1) if 'beginn' in columns else None

        # IDs über fest zugewiesene Spaltenbuchstaben
        xempus_id_ci = _col_index(XEMPUS_ID_COL) - 1
        arbn_id_ci = _col_index(XEMPUS_ARBN_ID_COL) - 1
        arbg_id_ci = _col_index(XEMPUS_ARBG_ID_COL) - 1

        for row in ws.iter_rows(min_row=2):
            beratungen.total_rows += 1
            row_num = beratungen.total_rows + 1  # Excel-Zeilennummer (mit Header)

            try:
                cells = list(row)
                max_ci = len(cells) - 1

                # Liefert String (oder None), falls Index gültig
                def _safe_str(ci):
                    if ci is not None and ci <= max_ci:
                        return str(cells[ci].value or '').strip()
                    return None

                status_raw = _safe_str(status_ci) or ''
                status = _map_xempus_status(status_raw)
                if status is None:
                    beratungen.skipped_rows += 1
                    continue

                vsnr_raw = cells[vsnr_ci].value if vsnr_ci is not None and vsnr_ci <= max_ci else None
                vsnr = str(vsnr_raw).strip() if vsnr_raw and str(vsnr_raw).strip() else ''

                xempus_id = _safe_str(xempus_id_ci) or None

                # Mindestens VSNR oder interne ID muss vorliegen
                if not vsnr and not xempus_id:
                    beratungen.skipped_rows += 1
                    continue

                # Berater ermitteln
                berater = ''
                if berater_ci is not None and berater_ci <= max_ci:
                    berater = str(cells[berater_ci].value or '').strip()

                # Arbeitnehmer kombinieren, ggf. Versicherungsnehmer nehmen
                arbn_nachname = _safe_str(arbn_name_ci) or ''
                arbn_vorname_val = _safe_str(arbn_vorname_ci) or ''
                if arbn_nachname:
                    vn_name = f"{arbn_nachname} {arbn_vorname_val}".strip() if arbn_vorname_val else arbn_nachname
                else:
                    vn_name = _safe_str(vn_ci)

                # Compose den Dataset
                beratungen.rows.append({
                    'vsnr': vsnr if vsnr else None,
                    'berater': berater,
                    'status': status,
                    'versicherer': _safe_str(versicherer_ci),
                    'versicherungsnehmer': vn_name,
                    'sparte': _safe_str(sparte_ci),
                    'tarif': _safe_str(tarif_ci),
                    'beitrag': _parse_amount(cells[beitrag_ci].value) if beitrag_ci is not None and beitrag_ci <= max_ci else None,
                    'beginn': _parse_date(cells[beginn_ci].value) if beginn_ci is not None and beginn_ci <= max_ci else None,
                    'xempus_id': xempus_id,
                    'arbn_id': _safe_str(arbn_id_ci),
                    'arbg_id': _safe_str(arbg_id_ci),
                })
            except Exception as e:
                # Fehler abfangen und Zeilen überspringen, Fehler mitprotokollieren
                beratungen.errors.append(f"Zeile {row_num}: {str(e)}")
                beratungen.skipped_rows += 1
    else:
        beratungen.errors.append("Sheet 'Beratungen' nicht gefunden")

    # Alle gefundenen Beraternamen aufsammeln (für UI)
    berater_names = sorted(set(
        r['berater'] for r in beratungen.rows if r.get('berater')
    ))

    wb.close()

    logger.info(
        f"Xempus full: {len(sheets_found)} Sheets, "
        f"{len(beratungen.rows)} Beratungen, "
        f"{len(berater_names)} Berater erkannt"
    )

    # Gesamter Schritt als Ergebnis-Objekt
    return XempusFullResult(
        beratungen=beratungen,
        sheets_found=sheets_found,
        sheet_row_counts=sheet_row_counts,
        berater_names=berater_names,
    )


def compute_file_hash(filepath: str) -> str:
    """
    Berechne SHA256-Hash einer Datei (z.B. zur Dublettenprüfung/Änderungserkennung).
    Liefert Hex-String zurück.
    """
    h = hashlib.sha256()
    with open(filepath, 'rb') as f:
        for chunk in iter(lambda: f.read(65536), b''):
            h.update(chunk)
    return h.hexdigest()

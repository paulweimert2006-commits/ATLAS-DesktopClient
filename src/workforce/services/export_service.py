"""
Export-Service: Excel-Generierung fuer Standard- und Delta-SCS-Exporte.

Quelle: app.py Zeilen 2394-2507
"""

import os
from datetime import datetime

import openpyxl

from workforce.constants import SCS_HEADERS
from workforce.helpers import getv, get_from_path, get_safe_employer_name


def map_to_scs_schema(e: dict, employer_name: str, provider_key: str) -> dict:
    """
    Mappt einen Mitarbeiterdatensatz auf das SCS-Tabellenkalkulationsschema.

    Args:
        e: Mitarbeiterdaten
        employer_name: Name des Arbeitgebers
        provider_key: Provider-Schluessel fuer provider-spezifische Logik

    Returns:
        Dict im SCS-Schema
    """
    details = e.get("details", {}) if isinstance(e.get("details"), dict) else None
    graw = getv(e, details, "geschlecht", "gender")
    gl = graw.lower()
    gmap_m = {"mann", "maennlich", "m", "male", "mr", "mr.", "herr", "1"}
    gmap_f = {"frau", "weiblich", "w", "female", "f", "mrs", "mrs.", "2"}
    gender = "Mann" if gl in gmap_m else ("Frau" if gl in gmap_f else graw)

    strasse = getv(e, details, "strasse", "address.street")
    hausnummer = getv(e, details, "hausnummer", "address.streetNumber")
    plz = getv(e, details, "postleitzahl") or getv(e, details, "plz", "address.zipCode")
    ort = getv(e, details, "ort") or getv(e, details, "stadt", "address.city")

    if provider_key == 'hrworks':
        if strasse == 'n/a': strasse = ''
        if hausnummer == 'n/a': hausnummer = ''
        if plz == 'n/a': plz = ''
        if ort == 'n/a': ort = ''

    email = getv(e, details, "email", "email")
    if provider_key == 'personio':
        personal_email = getv(e, details, "persoenliche e-mail")
        work_email = getv(e, details, "e-mail")
        email = personal_email or work_email

    status_raw = getv(e, details, "status", "status")
    sl = status_raw.lower()
    active_vals = {"active", "aktiv"}
    status = "Aktiv" if sl in active_vals else "Ehemalig"

    return {
        "Name": getv(e, details, "nachname", "lastName"),
        "Vorname": getv(e, details, "vorname", "firstName"),
        "Geschlecht": gender,
        "Titel": getv(e, details, "titel", "title"),
        "Geburtsdatum": getv(e, details, "geburtsdatum", "birthday"),
        "Strasse": strasse,
        "Hausnummer": hausnummer,
        "PLZ": plz,
        "Ort": ort,
        "Land": getv(e, details, "land", "address.country") or "D",
        "Kommentar": "",
        "Email": email,
        "Telefon": getv(e, details, "mobilnummer", "phone"),
        "Personalnummer": getv(e, details, "personalnummer", "personnelNumber"),
        "Position": getv(e, details, "position", "position"),
        "Firmeneintritt": getv(e, details, "eintrittsdatum", "joinDate"),
        "Bruttogehalt": getv(e, details, "festgehalt", "salary.amount"),
        "VWL": "",
        "geldwerterVorteil": "",
        "SteuerfreibetragJahr": "",
        "SteuerfreibetragMonat": "",
        "SV_Brutto": "",
        "Steuerklasse": getv(e, details, "lohnsteuerklasse", "taxClass"),
        "Religion": getv(e, details, "kirchensteuer", "religion"),
        "Kinder": getv(e, details, "kinderfreibetrag", "children"),
        "Abteilung": getv(e, details, "abteilung", "organizationUnit.name", "department"),
        "Arbeitsplatz": getv(e, details, "arbeitsplatz", "office"),
        "Arbeitgeber": employer_name,
        "Status": status,
    }


def _add_employer_sheet(wb: openpyxl.Workbook, employer_name: str,
                        employer_cfg: dict = None) -> None:
    """Fuegt ein Arbeitgeber-Blatt zur Workbook hinzu."""
    ws_org = wb.create_sheet("Arbeitgeber")
    org_headers = ["Name", "Strasse", "PLZ", "Ort", "Land", "Kommentar", "Email", "Telefon", "Fax"]
    ws_org.append(org_headers)
    cfg = employer_cfg or {}
    addr = cfg.get("address_json") or cfg.get("address") or {}
    if not isinstance(addr, dict):
        addr = {}
    org_row = {
        "Name": employer_name,
        "Strasse": addr.get("street") or "",
        "PLZ": addr.get("zip_code") or addr.get("zipCode") or "",
        "Ort": addr.get("city") or "",
        "Land": addr.get("country") or "D",
        "Kommentar": cfg.get("comment") or "",
        "Email": cfg.get("email") or "",
        "Telefon": cfg.get("phone") or "",
        "Fax": cfg.get("fax") or "",
    }
    ws_org.append([org_row.get(h, "") for h in org_headers])


def generate_standard_export(employees: list[dict], employer_name: str,
                             provider_key: str, exports_dir: str,
                             employer_cfg: dict = None) -> str:
    """
    Generiert einen vollstaendigen XLSX-Export aller Mitarbeiterdaten.

    Args:
        employees: Liste der Mitarbeiterdaten
        employer_name: Name des Arbeitgebers
        provider_key: Provider-Schluessel
        exports_dir: Verzeichnis fuer Exporte
        employer_cfg: Vollstaendiger Arbeitgeber-Dict (fuer Arbeitgeber-Blatt)

    Returns:
        Dateipfad der generierten Excel-Datei
    """
    rows = [map_to_scs_schema(e, employer_name, provider_key) for e in employees]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mitarbeiter"
    hl = list(SCS_HEADERS)
    ws.append(hl)
    for rd in rows:
        ws.append([rd.get(h) for h in hl])

    _add_employer_sheet(wb, employer_name, employer_cfg)

    t = datetime.now().strftime("%Y%m%d-%H%M%S")
    fn = f"standard_{get_safe_employer_name(employer_name)}_{provider_key}_{t}.xlsx"
    os.makedirs(exports_dir, exist_ok=True)
    fp = os.path.join(exports_dir, fn)
    wb.save(fp)
    return fp


def generate_delta_excel(changed_employees: dict, employer_cfg: dict,
                         exports_dir: str) -> str:
    """
    Generiert eine Delta-SCS-Excel-Datei.

    Args:
        changed_employees: Dict {pid: {hash, flat, core, dates}} fuer exportierte PIDs
        employer_cfg: Arbeitgeber-Konfiguration
        exports_dir: Verzeichnis fuer Exporte

    Returns:
        Dateipfad der generierten Excel-Datei
    """
    employer_name = employer_cfg.get('name', '')
    provider_key = employer_cfg.get('provider_key', '')
    safe_emp = get_safe_employer_name(employer_name)

    wb = openpyxl.Workbook()
    ws_emp = wb.active
    ws_emp.title = "Mitarbeiter"
    ws_emp.append(SCS_HEADERS)
    for pid in sorted(changed_employees.keys()):
        ws_emp.append([changed_employees[pid]["core"].get(h, "") for h in SCS_HEADERS])

    _add_employer_sheet(wb, employer_name, employer_cfg)

    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    os.makedirs(exports_dir, exist_ok=True)
    outfile = os.path.join(exports_dir, f"delta-{safe_emp}-{provider_key}-{ts}.xlsx")
    wb.save(outfile)
    return outfile

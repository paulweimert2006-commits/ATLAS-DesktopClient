"""
Spreadsheet-Extractor — Standalone-Funktion fuer die Textextraktion
aus Tabellendateien (CSV, TSV, XLSX) fuer KI-Klassifikation.

Extrahiert aus services/document_processor.py.
"""

import logging
import os

logger = logging.getLogger(__name__)

__all__ = [
    'extract_text',
]


def extract_text(file_path: str, max_lines: int = 50) -> str:
    """
    Extrahiert Text aus einer Tabellendatei fuer KI-Klassifikation.

    Args:
        file_path: Pfad zur Datei
        max_lines: Maximale Anzahl Zeilen

    Returns:
        Extrahierter Text (Zeilen getrennt durch Newline)
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext in ['.csv', '.tsv']:
        for encoding in ['utf-8', 'cp1252', 'latin-1']:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    lines = []
                    for i, line in enumerate(f):
                        if i >= max_lines:
                            break
                        lines.append(line.rstrip())
                    return '\n'.join(lines)
            except (UnicodeDecodeError, UnicodeError):
                continue
        return ''

    elif ext == '.xlsx':
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            lines = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= max_lines:
                    break
                cells = [str(c) if c is not None else '' for c in row]
                lines.append(' | '.join(cells))
            wb.close()
            return '\n'.join(lines)
        except Exception as e:
            logger.warning(f"Excel-Extraktion fehlgeschlagen: {e}")
            return ''

    return ''

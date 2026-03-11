"""
ACENCIA ATLAS - Spreadsheet Viewer Dialog

Dialog zur Vorschau von CSV- und Excel-Dateien.
"""

import os
import logging

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QToolBar, QTableWidget, QTableWidgetItem, QHeaderView,
    QComboBox, QMessageBox, QApplication, QWidget, QSizePolicy,
)
from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtGui import QFont, QAction

logger = logging.getLogger(__name__)

__all__ = ["SpreadsheetViewerDialog"]

# openpyxl fuer .xlsx (optional)
HAS_OPENPYXL = False
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    pass


class _SpreadsheetLoadWorker(QThread):
    """Laedt CSV/XLSX-Daten im Hintergrund."""
    finished = Signal(dict)
    error = Signal(str)

    def __init__(self, file_path: str, max_rows: int, parent=None):
        super().__init__(parent)
        self._file_path = file_path
        self._max_rows = max_rows

    def run(self):
        import csv
        ext = os.path.splitext(self._file_path)[1].lower()
        try:
            sheets: dict = {}
            if ext == '.csv':
                self._load_csv(sheets)
            elif ext == '.tsv':
                self._load_csv(sheets, delimiter='\t')
            elif ext == '.xlsx':
                self._load_xlsx(sheets)
            elif ext == '.xls':
                self.finished.emit({"__xls__": True})
                return
            else:
                self._load_csv(sheets)
            self.finished.emit(sheets)
        except Exception as e:
            self.error.emit(str(e))

    def _load_csv(self, sheets: dict, delimiter: str = None):
        import csv
        encodings = ['utf-8-sig', 'utf-8', 'cp1252', 'latin-1']
        content = None
        for enc in encodings:
            try:
                with open(self._file_path, 'r', encoding=enc) as f:
                    content = f.read()
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
        if content is None:
            raise ValueError("Encoding nicht erkannt")
        if delimiter is None:
            sniffer = csv.Sniffer()
            try:
                dialect = sniffer.sniff(content[:8192], delimiters=',;\t|')
                delimiter = dialect.delimiter
            except csv.Error:
                delimiter = ';' if ';' in content[:2000] else ','
        lines = content.splitlines()
        reader = csv.reader(lines, delimiter=delimiter)
        rows, headers = [], []
        for i, row in enumerate(reader):
            if i == 0:
                headers = [str(h).strip() for h in row]
            else:
                rows.append([str(cell) for cell in row])
            if i >= self._max_rows:
                break
        if not headers and rows:
            max_cols = max(len(r) for r in rows) if rows else 0
            headers = [f"Spalte {i+1}" for i in range(max_cols)]
        sheet_name = os.path.basename(self._file_path)
        sheets[sheet_name] = (headers, rows, len(lines) - 1)

    def _load_xlsx(self, sheets: dict):
        if not HAS_OPENPYXL:
            sheets["__no_openpyxl__"] = True
            return
        wb = openpyxl.load_workbook(self._file_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers, rows, total_rows = [], [], 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c) if c is not None else "" for c in row]
                else:
                    rows.append([str(c) if c is not None else "" for c in row])
                total_rows = i
                if i >= self._max_rows:
                    break
            if not headers and not rows:
                headers = ["(leer)"]
            if not headers and rows:
                max_cols = max(len(r) for r in rows) if rows else 0
                headers = [f"Spalte {i+1}" for i in range(max_cols)]
            sheets[sheet_name] = (headers, rows, total_rows)
        wb.close()


class SpreadsheetViewerDialog(QDialog):
    """
    Dialog zur Vorschau von CSV- und Excel-Dateien.
    
    Zeigt tabellarische Daten in einem QTableWidget an.
    Unterstuetzt:
    - CSV (.csv) via Python csv-Modul
    - Excel (.xlsx) via openpyxl
    - TSV (.tsv) via Python csv-Modul
    """
    
    # Maximale Zeilen fuer die Vorschau (Performance-Schutz)
    MAX_PREVIEW_ROWS = 5000
    
    def __init__(self, file_path: str, title: str = "", parent=None):
        super().__init__(parent)
        self.file_path = file_path
        self._sheets_data: dict = {}  # {sheet_name: (headers, rows)}
        self._current_sheet: str = ""
        
        from i18n.de import SPREADSHEET_PREVIEW_TITLE
        display_title = title or SPREADSHEET_PREVIEW_TITLE.format(
            filename=os.path.basename(file_path)
        )
        self.setWindowTitle(display_title)
        self.setMinimumSize(900, 600)
        self.resize(1100, 750)
        
        self._load_worker = None
        self._setup_ui()
        self._load_data_async()
    
    def _setup_ui(self):
        """Erstellt die UI-Elemente."""
        from i18n.de import (
            SPREADSHEET_SHEET_LABEL, SPREADSHEET_EXTERN_OPEN,
            SPREADSHEET_CLOSE, SPREADSHEET_NO_DATA
        )
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        
        # Toolbar
        toolbar = QToolBar()
        toolbar.setMovable(False)
        
        # Titel
        title_label = QLabel(f"  {os.path.basename(self.file_path)}")
        title_label.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        toolbar.addWidget(title_label)
        
        toolbar.addSeparator()
        
        # Sheet-Auswahl (nur fuer Excel mit mehreren Blaettern)
        self._sheet_label = QLabel(f"  {SPREADSHEET_SHEET_LABEL} ")
        self._sheet_label.setVisible(False)
        toolbar.addWidget(self._sheet_label)
        
        self._sheet_combo = QComboBox()
        self._sheet_combo.setMinimumWidth(150)
        self._sheet_combo.currentTextChanged.connect(self._on_sheet_changed)
        self._sheet_combo.setVisible(False)
        toolbar.addWidget(self._sheet_combo)
        
        toolbar.addSeparator()
        
        # Zeilen-Info
        self._info_label = QLabel("")
        self._info_label.setFont(QFont("Segoe UI", 9))
        toolbar.addWidget(self._info_label)
        
        # Spacer
        spacer = QWidget()
        spacer.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        toolbar.addWidget(spacer)
        
        # Extern oeffnen
        open_external_btn = QPushButton(SPREADSHEET_EXTERN_OPEN)
        open_external_btn.setToolTip("Mit System-Anwendung oeffnen")
        open_external_btn.clicked.connect(self._open_external)
        toolbar.addWidget(open_external_btn)
        
        toolbar.addSeparator()
        
        # Schliessen
        close_btn = QPushButton(SPREADSHEET_CLOSE)
        close_btn.clicked.connect(self.close)
        toolbar.addWidget(close_btn)
        
        layout.addWidget(toolbar)
        
        # Tabelle
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setDefaultSectionSize(24)
        self.table.setStyleSheet("""
            QTableWidget {
                font-family: 'Segoe UI', 'Consolas', monospace;
                font-size: 12px;
                gridline-color: #e0e0e0;
            }
            QTableWidget::item {
                padding: 2px 6px;
            }
            QHeaderView::section {
                background-color: #f0f0f0;
                padding: 4px 6px;
                border: 1px solid #d0d0d0;
                font-weight: bold;
                font-size: 11px;
            }
            QTableWidget::item:alternate {
                background-color: #fafafa;
            }
        """)
        layout.addWidget(self.table)
        
        # Hinweis-Label (bei leeren Daten)
        self._empty_label = QLabel(SPREADSHEET_NO_DATA)
        self._empty_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._empty_label.setFont(QFont("Segoe UI", 11))
        self._empty_label.setVisible(False)
        layout.addWidget(self._empty_label)
        
        # Inline-Status-Label fuer Fehler/Hinweise (statt modaler Dialoge)
        self._status_label = QLabel("")
        self._status_label.setVisible(False)
        self._status_label.setWordWrap(True)
        layout.addWidget(self._status_label)
    
    def _load_data_async(self):
        """Startet das Laden der Daten im Hintergrund."""
        from i18n.de import SPREADSHEET_LOAD_ERROR

        self._info_label.setText("  Laden...")

        self._load_worker = _SpreadsheetLoadWorker(
            self.file_path, self.MAX_PREVIEW_ROWS, parent=self,
        )

        def _on_loaded(sheets: dict):
            if "__xls__" in sheets:
                self._show_xls_message()
                return
            if "__no_openpyxl__" in sheets:
                from i18n.de import SPREADSHEET_XLSX_NOT_AVAILABLE
                self._status_label.setText(SPREADSHEET_XLSX_NOT_AVAILABLE)
                self._status_label.setStyleSheet(
                    "color: #1e40af; background: #eff6ff; padding: 6px 12px; border-radius: 4px;"
                )
                self._status_label.setVisible(True)
                self._open_external()
                return

            self._sheets_data = sheets
            if self._sheets_data:
                sheet_names = list(self._sheets_data.keys())
                if len(sheet_names) > 1:
                    self._sheet_label.setVisible(True)
                    self._sheet_combo.setVisible(True)
                    self._sheet_combo.addItems(sheet_names)
                self._display_sheet(sheet_names[0])

        def _on_error(msg: str):
            logger.error(f"Fehler beim Laden der Tabelle: {msg}")
            self._status_label.setText(SPREADSHEET_LOAD_ERROR.format(error=msg))
            self._status_label.setStyleSheet(
                "color: #dc2626; background: #fef2f2; padding: 6px 12px; border-radius: 4px;"
            )
            self._status_label.setVisible(True)

        self._load_worker.finished.connect(_on_loaded)
        self._load_worker.error.connect(_on_error)
        self._load_worker.start()
    
    def _show_xls_message(self):
        """Zeigt Hinweis fuer alte .xls Dateien."""
        from i18n.de import SPREADSHEET_XLS_NOT_SUPPORTED
        self._status_label.setText(SPREADSHEET_XLS_NOT_SUPPORTED)
        self._status_label.setStyleSheet(
            "color: #1e40af; background: #eff6ff; padding: 6px 12px; border-radius: 4px;"
        )
        self._status_label.setVisible(True)
        self._open_external()
    
    def _display_sheet(self, sheet_name: str):
        """Zeigt die Daten eines Sheets in der Tabelle an."""
        from i18n.de import (
            SPREADSHEET_ROWS_INFO, SPREADSHEET_MAX_ROWS_INFO,
            SPREADSHEET_NO_DATA
        )
        
        if sheet_name not in self._sheets_data:
            return
        
        self._current_sheet = sheet_name
        headers, rows, total_rows = self._sheets_data[sheet_name]
        
        if not headers and not rows:
            self.table.setVisible(False)
            self._empty_label.setVisible(True)
            self._info_label.setText(SPREADSHEET_NO_DATA)
            return
        
        self.table.setVisible(True)
        self._empty_label.setVisible(False)
        
        # Tabelle befuellen
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(rows))
        
        for row_idx, row_data in enumerate(rows):
            for col_idx, cell_value in enumerate(row_data):
                if col_idx < len(headers):
                    item = QTableWidgetItem(cell_value)
                    self.table.setItem(row_idx, col_idx, item)
        
        # Spaltenbreiten anpassen
        self.table.resizeColumnsToContents()
        
        # Maximale Spaltenbreite begrenzen
        for col in range(self.table.columnCount()):
            if self.table.columnWidth(col) > 300:
                self.table.setColumnWidth(col, 300)
        
        # Info-Label aktualisieren
        shown_rows = len(rows)
        if total_rows > self.MAX_PREVIEW_ROWS:
            self._info_label.setText(
                f"  {SPREADSHEET_MAX_ROWS_INFO.format(shown=shown_rows, total=total_rows)}"
            )
        else:
            self._info_label.setText(
                f"  {SPREADSHEET_ROWS_INFO.format(rows=shown_rows, cols=len(headers))}"
            )
    
    def _on_sheet_changed(self, sheet_name: str):
        """Handler fuer Sheet-Wechsel."""
        if sheet_name and sheet_name in self._sheets_data:
            self._display_sheet(sheet_name)
    
    def _open_external(self):
        """Oeffnet die Datei mit der System-Anwendung (non-blocking)."""
        import subprocess
        import sys
        
        try:
            if sys.platform == 'win32':
                os.startfile(self.file_path)
            elif sys.platform == 'darwin':
                subprocess.Popen(['open', self.file_path])
            else:
                subprocess.Popen(['xdg-open', self.file_path])
        except Exception as e:
            self._status_label.setText(f"Konnte Datei nicht oeffnen: {e}")
            self._status_label.setStyleSheet(
                "color: #dc2626; background: #fef2f2; padding: 6px 12px; border-radius: 4px;"
            )
            self._status_label.setVisible(True)

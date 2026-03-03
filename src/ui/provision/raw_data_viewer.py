# -*- coding: utf-8 -*-
"""
Excel-Rohdaten-Viewer Dialog.

Zeigt den Inhalt einer Original-Excel-Datei in einer QTableView an
und springt optional zu einer bestimmten Zeile (source_row).

Unterstuetzt zwei Betriebsarten:
  1. Datei-Modus: Laedt eine Excel-Datei direkt (filepath)
  2. JSON-Modus:  Zeigt gespeicherte Rohdaten aus dem Backend (headers + rows)
"""

import logging
import os
from typing import Optional, List

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QTableView,
    QPushButton, QLabel, QLineEdit, QHeaderView,
    QFileDialog, QAbstractItemView, QFrame,
)
from PySide6.QtCore import Qt, QAbstractTableModel, QModelIndex, QThread, Signal, QObject
from PySide6.QtGui import QColor
from PySide6.QtPrintSupport import QPrinter, QPrintDialog

from ui.styles.tokens import (
    PRIMARY_0, PRIMARY_500, PRIMARY_900, ACCENT_500,
    FONT_BODY, FONT_SIZE_BODY, FONT_SIZE_CAPTION,
    SUCCESS, WARNING,
)
from infrastructure.threading.worker_utils import run_worker
from i18n import de as texts

logger = logging.getLogger(__name__)


class _ExcelLoadWorker(QThread):
    finished = Signal(list, list)
    error = Signal(str)

    def __init__(self, filepath: str, sheet_name: str = None):
        super().__init__()
        self._filepath = filepath
        self._sheet_name = sheet_name

    def run(self):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self._filepath, read_only=True, data_only=True)
            if self._sheet_name and self._sheet_name in wb.sheetnames:
                ws = wb[self._sheet_name]
            else:
                ws = wb.active

            headers = []
            rows_data = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                str_row = [str(v) if v is not None else '' for v in row]
                if row_idx == 0:
                    headers = str_row
                else:
                    rows_data.append(str_row)

            wb.close()
            self.finished.emit(headers, rows_data)
        except Exception as e:
            self.error.emit(str(e))


class _ExcelTableModel(QAbstractTableModel):

    def __init__(self):
        super().__init__()
        self._headers: List[str] = []
        self._rows: List[List[str]] = []
        self._highlight_row = -1

    def set_data(self, headers: list, rows: list):
        self.beginResetModel()
        self._headers = headers
        self._rows = rows
        self.endResetModel()

    def set_highlight_row(self, row: int):
        self._highlight_row = row
        self.dataChanged.emit(
            self.index(0, 0),
            self.index(self.rowCount() - 1, self.columnCount() - 1),
        )

    def rowCount(self, parent=QModelIndex()):
        return len(self._rows)

    def columnCount(self, parent=QModelIndex()):
        return len(self._headers) if self._headers else 0

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            if section < len(self._headers):
                return self._headers[section]
        if orientation == Qt.Vertical and role == Qt.DisplayRole:
            return str(section + 2)
        return None

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        row_data = self._rows[index.row()]
        col = index.column()

        if role == Qt.DisplayRole:
            if col < len(row_data):
                return row_data[col]
            return ''

        if role == Qt.BackgroundRole:
            if index.row() == self._highlight_row:
                return QColor(ACCENT_500 + '40')

        return None

    def get_headers(self) -> List[str]:
        return self._headers

    def get_rows(self) -> List[List[str]]:
        return self._rows


class RawDataViewerDialog(QDialog):
    """In-App Rohdaten-Viewer mit Zeilen-Navigation, Export und Druck.

    Datei-Modus: filepath angeben.
    JSON-Modus:  headers + rows_data angeben (aus Backend).
    """

    def __init__(self, parent=None, filepath: str = None,
                 sheet_name: str = None, target_row: int = None,
                 headers: List[str] = None, rows_data: List[List[str]] = None,
                 title: str = None):
        super().__init__(parent)
        self._filepath = filepath
        self._sheet_name = sheet_name
        self._target_row = target_row
        self._load_worker = None
        self._search_ctx = QObject(self)
        self._model = _ExcelTableModel()
        self._json_mode = headers is not None and rows_data is not None

        if title:
            self.setWindowTitle(title)
        elif filepath:
            self.setWindowTitle(texts.PROVISION_RAW_VIEWER_TITLE.format(
                filename=filepath.split('\\')[-1].split('/')[-1]))
        else:
            self.setWindowTitle(texts.PROVISION_RAW_VIEWER_TITLE.format(filename=''))

        self.setMinimumSize(1000, 600)
        self._setup_ui()

        if self._json_mode:
            self._on_data_loaded(headers, rows_data)
        elif filepath:
            self._load_file(filepath, sheet_name)

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        toolbar = QHBoxLayout()
        toolbar.setSpacing(8)

        if not self._json_mode:
            self._file_btn = QPushButton(texts.PROVISION_RAW_VIEWER_FILE_SELECT)
            self._file_btn.setCursor(Qt.PointingHandCursor)
            self._file_btn.clicked.connect(self._on_select_file)
            toolbar.addWidget(self._file_btn)

        if self._sheet_name:
            sheet_lbl = QLabel(texts.PROVISION_RAW_VIEWER_SHEET.format(sheet=self._sheet_name))
            sheet_lbl.setStyleSheet(f"color: {PRIMARY_500}; font-size: {FONT_SIZE_CAPTION};")
            toolbar.addWidget(sheet_lbl)

        toolbar.addStretch()

        self._search_input = QLineEdit()
        self._search_input.setPlaceholderText(texts.PROVISION_RAW_VIEWER_SEARCH)
        self._search_input.setMaximumWidth(250)
        self._search_input.textChanged.connect(self._on_search)
        toolbar.addWidget(self._search_input)

        if self._target_row is not None:
            self._jump_btn = QPushButton(
                texts.PROVISION_RAW_VIEWER_ROW.format(row=self._target_row))
            self._jump_btn.setCursor(Qt.PointingHandCursor)
            self._jump_btn.setStyleSheet(f"""
                QPushButton {{
                    background: {ACCENT_500}; color: white;
                    border: none; border-radius: 4px;
                    padding: 6px 12px; font-weight: 600;
                }}
                QPushButton:hover {{ background: {SUCCESS}; }}
            """)
            self._jump_btn.clicked.connect(self._jump_to_target)
            toolbar.addWidget(self._jump_btn)

        layout.addLayout(toolbar)

        self._status_label = QLabel(texts.PROVISION_RAW_VIEWER_LOADING)
        self._status_label.setStyleSheet(f"color: {PRIMARY_500}; font-size: {FONT_SIZE_CAPTION};")
        layout.addWidget(self._status_label)

        self._table = QTableView()
        self._table.setModel(self._model)
        self._table.setAlternatingRowColors(True)
        self._table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._table.setSelectionMode(QAbstractItemView.SingleSelection)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self._table.verticalHeader().setDefaultSectionSize(28)
        layout.addWidget(self._table)

        btn_row = QHBoxLayout()

        export_btn = QPushButton(texts.PM_RAW_BTN_EXPORT)
        export_btn.setCursor(Qt.PointingHandCursor)
        export_btn.clicked.connect(self._export_xlsx)
        btn_row.addWidget(export_btn)

        print_btn = QPushButton(texts.PM_RAW_BTN_PRINT)
        print_btn.setCursor(Qt.PointingHandCursor)
        print_btn.clicked.connect(self._print_table)
        btn_row.addWidget(print_btn)

        btn_row.addStretch()

        close_btn = QPushButton(texts.CLOSE)
        close_btn.setCursor(Qt.PointingHandCursor)
        close_btn.clicked.connect(self.close)
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)

    def _load_file(self, filepath: str, sheet_name: str = None):
        self._status_label.setText(texts.PROVISION_RAW_VIEWER_LOADING)
        self._load_worker = _ExcelLoadWorker(filepath, sheet_name)
        self._load_worker.finished.connect(self._on_data_loaded)
        self._load_worker.error.connect(self._on_load_error)
        self._load_worker.start()

    def _on_data_loaded(self, headers: list, rows: list):
        self._model.set_data(headers, rows)
        self._status_label.setText(f"{len(rows)} Zeilen")
        if self._target_row is not None:
            self._jump_to_target()

    def _on_load_error(self, error: str):
        self._status_label.setText(
            texts.PROVISION_RAW_VIEWER_ERROR.format(error=error))

    def _jump_to_target(self):
        if self._target_row is None:
            return
        model_row = self._target_row - 2
        logger.info(f"Jump: target_row={self._target_row}, model_row={model_row}, "
                     f"total_rows={self._model.rowCount()}")
        if 0 <= model_row < self._model.rowCount():
            self._model.set_highlight_row(model_row)
            idx = self._model.index(model_row, 0)
            self._table.scrollTo(idx, QAbstractItemView.PositionAtCenter)
            self._table.selectRow(model_row)
        else:
            logger.warning(f"Jump: model_row={model_row} ausserhalb [0, {self._model.rowCount()-1}]")

    def _on_select_file(self):
        filepath, _ = QFileDialog.getOpenFileName(
            self,
            texts.PROVISION_RAW_VIEWER_FILE_SELECT,
            '',
            texts.PROVISION_RAW_VIEWER_FILE_FILTER,
        )
        if filepath:
            self._filepath = filepath
            self._load_file(filepath)

    def _on_search(self, text: str):
        self._pending_search_text = text
        if not text:
            return

        def find_match(worker):
            search_text = self._pending_search_text.lower()
            rows = self._model.get_rows()
            col_count = self._model.columnCount()
            for row_idx in range(len(rows)):
                if worker.is_cancelled():
                    return None
                row_data = rows[row_idx]
                for col_idx in range(min(len(row_data), col_count)):
                    val = row_data[col_idx]
                    if val and search_text in str(val).lower():
                        return (row_idx, col_idx)
            return None

        run_worker(
            self._search_ctx, find_match, self._on_search_result,
            debounce_ms=400,
        )

    def _on_search_result(self, result):
        if result is None:
            return
        row_idx, col_idx = result
        self._model.set_highlight_row(row_idx)
        idx = self._model.index(row_idx, col_idx)
        self._table.scrollTo(idx, QAbstractItemView.PositionAtCenter)
        self._table.selectRow(row_idx)

    def _export_xlsx(self):
        try:
            import openpyxl
        except ImportError:
            self._status_label.setText("openpyxl nicht installiert")
            return

        headers = self._model.get_headers()
        rows = self._model.get_rows()
        if not headers and not rows:
            return

        filepath, _ = QFileDialog.getSaveFileName(
            self, texts.PM_RAW_BTN_EXPORT, '',
            texts.PROVISION_RAW_VIEWER_FILE_FILTER,
        )
        if not filepath:
            return
        if not filepath.endswith('.xlsx'):
            filepath += '.xlsx'

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self._sheet_name or 'Rohdaten'

            if headers:
                ws.append(headers)
            for row in rows:
                ws.append(row)

            wb.save(filepath)
            self._status_label.setText(
                texts.PM_RAW_EXPORT_SUCCESS.format(path=os.path.basename(filepath)))
        except Exception as e:
            self._status_label.setText(
                texts.PM_RAW_EXPORT_ERROR.format(error=str(e)))

    def _print_table(self):
        printer = QPrinter(QPrinter.HighResolution)
        dlg = QPrintDialog(printer, self)
        if dlg.exec() != QDialog.Accepted:
            return

        from PySide6.QtGui import QTextDocument
        doc = QTextDocument()

        headers = self._model.get_headers()
        rows = self._model.get_rows()

        html = '<table border="1" cellspacing="0" cellpadding="4" style="border-collapse:collapse;">'
        if headers:
            html += '<tr>'
            html += '<th style="background:#ddd;">#</th>'
            for h in headers:
                html += f'<th style="background:#ddd;">{h}</th>'
            html += '</tr>'
        for idx, row in enumerate(rows):
            bg = ' style="background:#ffffcc;"' if (idx == (self._target_row or 0) - 2) else ''
            html += f'<tr{bg}>'
            html += f'<td>{idx + 2}</td>'
            for cell in row:
                html += f'<td>{cell}</td>'
            html += '</tr>'
        html += '</table>'

        doc.setHtml(html)
        doc.print_(printer)

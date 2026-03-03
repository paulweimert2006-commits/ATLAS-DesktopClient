"""
Auszahlungen & Reports-Panel: Monatsabrechnungen, Pruefberichte, Exporte.

Ersetzt: billing_panel.py
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QTableView,
    QHeaderView, QSplitter, QFrame, QScrollArea, QComboBox,
    QPushButton, QFileDialog, QMenu, QMessageBox,
)
from PySide6.QtCore import (
    Qt, Signal, QTimer, QPoint, QModelIndex,
)
from PySide6.QtGui import QColor
from typing import List, Optional
from datetime import datetime
import calendar
import csv
import os

from api.provision import ProvisionAPI
from domain.provision.entities import BeraterAbrechnung
from ui.styles.tokens import (
    PRIMARY_100, PRIMARY_500, PRIMARY_900, ACCENT_500,
    BG_PRIMARY, BG_SECONDARY, BORDER_DEFAULT,
    SUCCESS, ERROR, WARNING,
    FONT_BODY, FONT_SIZE_BODY, FONT_SIZE_CAPTION,
    PILL_COLORS, ROLE_BADGE_COLORS,
    get_provision_table_style,
)
from ui.provision.widgets import (
    SectionHeader, PillBadgeDelegate, StatementCard,
    PaginationBar, ThreeDotMenuDelegate, ProvisionLoadingOverlay,
    format_eur, get_secondary_button_style, get_combo_style,
)
from ui.provision.workers import (
    AuszahlungenLoadWorker, AuszahlungenPositionenWorker,
    AbrechnungGenerateWorker, AbrechnungStatusWorker,
    StatementExportWorker, StatementBatchExportWorker,
    StatementEmailWorker, StatementBatchEmailWorker,
)
from ui.provision.models import AuszahlungenModel, STATUS_LABELS, STATUS_PILL_MAP
from services.statement_export import FILE_FILTERS, get_statement_filename, EXTENSIONS
from infrastructure.threading.worker_utils import run_worker
from i18n import de as texts
import logging

logger = logging.getLogger(__name__)


class AuszahlungenPanel(QWidget):
    """Auszahlungen & Reports mit Statement-Detail und CSV-Export.

    Implementiert IPayoutsView fuer den PayoutsPresenter.
    """

    navigate_to_panel = Signal(int)

    def __init__(self, api: ProvisionAPI = None):
        super().__init__()
        self._api = api
        self._presenter = None
        self._worker = None
        self._toast_manager = None
        self._setup_ui()
        if api:
            QTimer.singleShot(100, self._load_data)

    @property
    def _backend(self):
        return self._presenter or self._api

    def set_presenter(self, presenter) -> None:
        """Verbindet dieses Panel mit dem PayoutsPresenter."""
        self._presenter = presenter
        presenter.set_view(self)
        self._load_data()

    # ── IPayoutsView ──

    def show_abrechnungen(self, abrechnungen: list) -> None:
        """View-Interface: Abrechnungen anzeigen."""
        self._loading_overlay.setVisible(False)
        self._all_data = abrechnungen
        self._pagination.set_total(len(abrechnungen))
        self._paginate()
        self._resize_columns()
        self._status.setText("")

    def show_loading(self, loading: bool) -> None:
        """View-Interface: Ladezustand."""
        overlay = getattr(self, '_loading_overlay', None)
        if overlay:
            overlay.setVisible(loading)

    def show_error(self, message: str) -> None:
        """View-Interface: Fehler anzeigen."""
        logger.error(f"Auszahlungen-Fehler: {message}")

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(16)

        # Header
        header_row = QHBoxLayout()
        header = SectionHeader(texts.PROVISION_PAY_TITLE, texts.PROVISION_PAY_DESC)
        header_row.addWidget(header)
        header_row.addStretch()

        # Monat
        header_row.addWidget(QLabel(texts.PROVISION_DASH_MONAT))
        self._monat_combo = QComboBox()
        self._monat_combo.setFixedWidth(150)
        self._monat_combo.setStyleSheet(get_combo_style())
        now = datetime.now()
        for i in range(12):
            m = now.month - i
            y = now.year
            if m <= 0:
                m += 12
                y -= 1
            val = f"{y}-{m:02d}"
            self._monat_combo.addItem(f"{m:02d}/{y}", val)
        self._monat_combo.currentIndexChanged.connect(self._load_data)
        header_row.addWidget(self._monat_combo)
        layout.addLayout(header_row)

        # Toolbar
        toolbar = QHBoxLayout()

        gen_btn = QPushButton(texts.PROVISION_PAY_GENERATE)
        gen_btn.setStyleSheet(f"""
            QPushButton {{ background-color: {ACCENT_500}; color: white; border: none;
                border-radius: 6px; padding: 8px 16px; font-weight: 500; }}
            QPushButton:hover {{ background-color: #e88a2d; }}
        """)
        gen_btn.clicked.connect(self._generate)
        toolbar.addWidget(gen_btn)

        csv_btn = QPushButton(texts.PROVISION_PAY_EXPORT_CSV)
        csv_btn.setStyleSheet(get_secondary_button_style())
        csv_btn.clicked.connect(self._export_csv)
        toolbar.addWidget(csv_btn)

        xlsx_btn = QPushButton(texts.PROVISION_PAY_EXPORT_EXCEL)
        xlsx_btn.setStyleSheet(get_secondary_button_style())
        xlsx_btn.clicked.connect(self._export_xlsx)
        toolbar.addWidget(xlsx_btn)

        stmt_btn = QPushButton(texts.PM_STMT_EXPORT_MENU)
        stmt_btn.setStyleSheet(get_secondary_button_style())
        stmt_menu = QMenu(self)
        stmt_menu.addAction(texts.PM_STMT_EXPORT_PDF, lambda: self._export_statement_selected('pdf'))
        stmt_menu.addAction(texts.PM_STMT_EXPORT_XLSX, lambda: self._export_statement_selected('xlsx'))
        stmt_menu.addAction(texts.PM_STMT_EXPORT_DOCX, lambda: self._export_statement_selected('docx'))
        stmt_menu.addSeparator()
        all_menu = stmt_menu.addMenu(texts.PM_STMT_EXPORT_ALL_MENU)
        all_menu.addAction(texts.PM_STMT_EXPORT_PDF, lambda: self._export_all_statements('pdf'))
        all_menu.addAction(texts.PM_STMT_EXPORT_XLSX, lambda: self._export_all_statements('xlsx'))
        all_menu.addAction(texts.PM_STMT_EXPORT_DOCX, lambda: self._export_all_statements('docx'))
        stmt_btn.setMenu(stmt_menu)
        toolbar.addWidget(stmt_btn)

        email_btn = QPushButton(texts.PM_STMT_EMAIL_SEND)
        email_btn.setStyleSheet(get_secondary_button_style())
        email_menu = QMenu(self)
        email_menu.addAction(texts.PM_STMT_EMAIL_SEND_SELECTED, self._send_email_selected)
        email_menu.addSeparator()
        email_menu.addAction(texts.PM_STMT_EMAIL_SEND_ALL, self._send_email_all)
        email_btn.setMenu(email_menu)
        toolbar.addWidget(email_btn)

        toolbar.addStretch()
        layout.addLayout(toolbar)

        # Splitter: Tabelle + Detail
        self._splitter = QSplitter(Qt.Horizontal)

        # Tabelle
        table_w = QWidget()
        table_l = QVBoxLayout(table_w)
        table_l.setContentsMargins(0, 0, 0, 0)

        self._model = AuszahlungenModel()
        self._table = QTableView()
        self._table.setModel(self._model)
        self._table.setAlternatingRowColors(True)
        self._table.setSelectionBehavior(QTableView.SelectRows)
        self._table.setSelectionMode(QTableView.SingleSelection)
        self._table.verticalHeader().setVisible(False)
        self._table.verticalHeader().setDefaultSectionSize(52)
        self._table.horizontalHeader().setStretchLastSection(False)
        self._table.setStyleSheet(get_provision_table_style())
        self._table.setMinimumHeight(350)
        self._table.setMinimumWidth(1100)
        self._table.selectionModel().selectionChanged.connect(self._on_selection)

        # Delegates
        role_del = PillBadgeDelegate(ROLE_BADGE_COLORS, label_map={
            'consulter': texts.PROVISION_EMP_ROLE_CONSULTER,
            'teamleiter': texts.PROVISION_EMP_ROLE_TEAMLEITER,
            'backoffice': texts.PROVISION_EMP_ROLE_BACKOFFICE,
        })
        self._table.setItemDelegateForColumn(self._model.COL_ROLE, role_del)
        self._role_del = role_del

        status_del = PillBadgeDelegate(PILL_COLORS, label_map={
            'entwurf': texts.PROVISION_STATUS_ENTWURF,
            'geprueft': texts.PROVISION_STATUS_GEPRUEFT,
            'freigegeben': texts.PROVISION_STATUS_FREIGEGEBEN,
            'ausgezahlt': texts.PROVISION_STATUS_AUSGEZAHLT,
        })
        self._table.setItemDelegateForColumn(self._model.COL_STATUS, status_del)
        self._status_del = status_del

        menu_del = ThreeDotMenuDelegate(self._build_menu)
        self._table.setItemDelegateForColumn(self._model.COL_MENU, menu_del)
        self._menu_del = menu_del

        table_l.addWidget(self._table)

        self._pagination = PaginationBar(page_size=25)
        self._pagination.page_changed.connect(self._on_page_changed)
        table_l.addWidget(self._pagination)

        self._splitter.addWidget(table_w)

        # Detail: StatementCard
        self._detail = self._create_detail()
        self._detail.setVisible(False)
        self._splitter.addWidget(self._detail)
        self._splitter.setStretchFactor(0, 3)
        self._splitter.setStretchFactor(1, 1)

        layout.addWidget(self._splitter)

        self._status = QLabel("")
        self._status.setStyleSheet(f"color: {ERROR}; font-size: {FONT_SIZE_CAPTION};")
        layout.addWidget(self._status)
        self._loading_overlay = ProvisionLoadingOverlay(self)

    def _create_detail(self) -> QFrame:
        frame = QFrame()
        frame.setMinimumWidth(300)
        frame.setMaximumWidth(400)
        frame.setStyleSheet(f"background: white; border: 1.5px solid #b0c4d8; border-radius: 8px;")

        outer = QVBoxLayout(frame)
        outer.setContentsMargins(16, 16, 16, 16)
        outer.setSpacing(12)

        close_row = QHBoxLayout()
        close_row.addStretch()
        close_btn = QPushButton("\u2715")
        close_btn.setFixedSize(28, 28)
        close_btn.setStyleSheet(f"QPushButton {{ border: none; color: {PRIMARY_500}; font-size: 14pt; background: transparent; }}")
        close_btn.clicked.connect(lambda: self._detail.setVisible(False))
        close_row.addWidget(close_btn)
        outer.addLayout(close_row)

        lbl = QLabel(texts.PROVISION_PAY_DETAIL_OVERVIEW)
        lbl.setStyleSheet(f"font-weight: 600; font-size: 11pt; color: {PRIMARY_900};")
        outer.addWidget(lbl)

        self._statement = StatementCard()
        outer.addWidget(self._statement)

        self._det_name = QLabel("")
        self._det_name.setStyleSheet(f"color: {PRIMARY_900}; font-size: 13pt; font-weight: 600;")
        outer.insertWidget(1, self._det_name)

        # Positionsliste
        pos_lbl = QLabel(texts.PROVISION_PAY_DETAIL_POSITIONS)
        pos_lbl.setStyleSheet(f"font-weight: 600; font-size: 11pt; color: {PRIMARY_900}; margin-top: 8px;")
        outer.addWidget(pos_lbl)

        self._pos_scroll = QScrollArea()
        self._pos_scroll.setWidgetResizable(True)
        self._pos_scroll.setMaximumHeight(200)
        self._pos_scroll.setStyleSheet("QScrollArea { border: none; }")
        self._pos_container = QWidget()
        self._pos_layout = QVBoxLayout(self._pos_container)
        self._pos_layout.setContentsMargins(0, 0, 0, 0)
        self._pos_layout.setSpacing(4)
        self._pos_scroll.setWidget(self._pos_container)
        outer.addWidget(self._pos_scroll)

        outer.addStretch()
        return frame

    def _build_menu(self, index: QModelIndex) -> Optional[QMenu]:
        item = self._model.get_item(index.row())
        if not item:
            return None
        menu = QMenu(self)
        menu.addAction(texts.PROVISION_MENU_DETAILS, lambda: self._show_detail(item))
        allowed = {
            'berechnet':   ['geprueft'],
            'geprueft':    ['berechnet', 'freigegeben'],
            'freigegeben': ['geprueft', 'ausgezahlt'],
            'ausgezahlt':  [],
        }
        transitions = allowed.get(item.status, [])
        if transitions:
            status_menu = menu.addMenu(texts.PROVISION_MENU_STATUS)
            for s_key in transitions:
                s_label = STATUS_LABELS.get(s_key, s_key)
                status_menu.addAction(s_label, lambda sid=item.id, sk=s_key: self._change_status(sid, sk))

        export_menu = menu.addMenu(texts.PM_STMT_EXPORT_MENU)
        export_menu.addAction(texts.PM_STMT_EXPORT_PDF,
                              lambda b=item: self._export_statement_for_berater(b, 'pdf'))
        export_menu.addAction(texts.PM_STMT_EXPORT_XLSX,
                              lambda b=item: self._export_statement_for_berater(b, 'xlsx'))
        export_menu.addAction(texts.PM_STMT_EXPORT_DOCX,
                              lambda b=item: self._export_statement_for_berater(b, 'docx'))

        menu.addSeparator()
        email_action = menu.addAction(
            texts.PM_STMT_EMAIL_SEND,
            lambda b=item: self._send_email_for_berater(b))
        email_action.setEnabled(item.has_email)
        if not item.has_email:
            email_action.setToolTip(texts.PM_STMT_EMAIL_TOOLTIP_NO_ADDR)

        if item.email_status == 'failed':
            menu.addAction(
                texts.PM_STMT_EMAIL_RESEND,
                lambda b=item: self._send_email_for_berater(b))
        return menu

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._loading_overlay.setGeometry(self.rect())

    def refresh(self):
        self._load_data()

    def _load_data(self):
        monat = self._monat_combo.currentData() or datetime.now().strftime('%Y-%m')
        self._status.setText("")
        self._loading_overlay.setGeometry(self.rect())
        self._loading_overlay.setVisible(True)

        if self._presenter:
            self._presenter.load_abrechnungen(monat)
            return

        if self._worker and self._worker.isRunning():
            return
        self._worker = AuszahlungenLoadWorker(self._backend, monat)
        self._worker.finished.connect(self._on_loaded)
        self._worker.error.connect(self._on_error)
        self._worker.start()

    def _on_loaded(self, data: List[BeraterAbrechnung]):
        self._loading_overlay.setVisible(False)
        self._all_data = data
        self._pagination.set_total(len(data))
        self._paginate()
        self._resize_columns()
        self._status.setText("")

    def _on_page_changed(self, page: int):
        self._paginate()

    def _paginate(self):
        data = getattr(self, '_all_data', [])
        page = self._pagination.current_page
        ps = self._pagination._page_size
        start = page * ps
        end = start + ps
        self._model.set_data(data[start:end])

    def _on_error(self, msg: str):
        self._loading_overlay.setVisible(False)
        self._status.setText(texts.PROVISION_DASH_ERROR)
        logger.error(f"Auszahlungen-Ladefehler: {msg}")

    def _resize_columns(self):
        h = self._table.horizontalHeader()
        m = self._model
        fixed = {
            m.COL_ROLE: 130,
            m.COL_BRUTTO: 110,
            m.COL_TL: 100,
            m.COL_NETTO: 110,
            m.COL_RUECK: 110,
            m.COL_KORREKTUR: 100,
            m.COL_VU_ABZUG: 100,
            m.COL_AUSZAHLUNG: 115,
            m.COL_POS: 50,
            m.COL_STATUS: 140,
            m.COL_VERSION: 50,
            m.COL_EMAIL: 110,
            m.COL_MENU: 48,
        }
        for i in range(m.columnCount()):
            if i in fixed:
                h.setSectionResizeMode(i, QHeaderView.Fixed)
                self._table.setColumnWidth(i, fixed[i])
            else:
                h.setSectionResizeMode(i, QHeaderView.Stretch)

    def _on_selection(self, selected, deselected):
        indexes = self._table.selectionModel().selectedRows()
        if indexes:
            item = self._model.get_item(indexes[0].row())
            if item:
                self._show_detail(item)

    def _show_detail(self, item: BeraterAbrechnung):
        self._det_name.setText(item.berater_name)
        self._statement.clear_rows()
        self._statement.add_line(texts.PROVISION_PAY_DETAIL_BRUTTO, format_eur(item.brutto_provision))
        self._statement.add_line(texts.PROVISION_PAY_DETAIL_TL, format_eur(item.tl_abzug), color=ERROR if item.tl_abzug < 0 else "")
        self._statement.add_line(texts.PROVISION_PAY_DETAIL_RUECK, format_eur(item.rueckbelastungen), color=ERROR if item.rueckbelastungen < 0 else "")
        if item.has_korrektur:
            self._statement.add_line(
                texts.PM_KORREKTUR_DETAIL_HEADER,
                format_eur(item.korrektur_vormonat),
                color=WARNING,
            )
        self._statement.add_separator()
        self._statement.add_line(texts.PROVISION_PAY_DETAIL_NETTO, format_eur(item.auszahlung), bold=True)

        self._load_positions(item)
        self._detail.setVisible(True)

    def _load_positions(self, item: BeraterAbrechnung):
        while self._pos_layout.count():
            w = self._pos_layout.takeAt(0).widget()
            if w:
                w.deleteLater()

        monat = item.abrechnungsmonat
        try:
            year, month = int(monat[:4]), int(monat[5:7])
            _, last_day = calendar.monthrange(year, month)
            von = f"{monat}-01"
            bis = f"{monat}-{last_day:02d}"
        except (ValueError, IndexError):
            von, bis = None, None

        if not von or not bis:
            return

        self._current_positions_berater_id = item.berater_id

        if hasattr(self, '_pos_worker') and self._pos_worker and self._pos_worker.isRunning():
            return
        self._pos_worker = AuszahlungenPositionenWorker(self._backend, item.berater_id, von, bis)
        self._pos_worker.finished.connect(self._on_positions_loaded)
        self._pos_worker.error.connect(lambda msg: logger.warning(f"Positionen-Laden fehlgeschlagen: {msg}"))
        self._pos_worker.start()

    def _on_positions_loaded(self, berater_id: int, comms: list):
        if getattr(self, '_current_positions_berater_id', None) != berater_id:
            return

        while self._pos_layout.count():
            w = self._pos_layout.takeAt(0).widget()
            if w:
                w.deleteLater()

        if not comms:
            lbl = QLabel(texts.PROVISION_PAY_NO_POSITIONS)
            lbl.setStyleSheet(f"color: {PRIMARY_500}; font-style: italic;")
            self._pos_layout.addWidget(lbl)
            return

        for c in comms:
            row = QFrame()
            row.setStyleSheet(
                f"QFrame {{ background: {BG_SECONDARY}; border-radius: 4px; padding: 4px 8px; }}"
            )
            hl = QHBoxLayout(row)
            hl.setContentsMargins(0, 0, 0, 0)
            hl.setSpacing(4)

            desc = c.versicherer or c.vu_name or ""
            if c.vsnr:
                desc += f"  {c.vsnr}"
            left = QLabel(desc)
            left.setStyleSheet(f"font-size: {FONT_SIZE_CAPTION}; color: {PRIMARY_900};")
            hl.addWidget(left, 1)

            amount = QLabel(format_eur(c.betrag))
            color = ERROR if c.betrag < 0 else PRIMARY_900
            amount.setStyleSheet(f"font-size: {FONT_SIZE_CAPTION}; font-weight: 600; color: {color};")
            amount.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            hl.addWidget(amount)

            self._pos_layout.addWidget(row)

    def _generate(self):
        monat = self._monat_combo.currentData() or datetime.now().strftime('%Y-%m')
        result = QMessageBox.question(
            self,
            texts.PROVISION_PAY_GENERATE,
            texts.PROVISION_PAY_CONFIRM.format(monat=monat),
        )
        if result != QMessageBox.Yes:
            return
        if hasattr(self, '_gen_worker') and self._gen_worker and self._gen_worker.isRunning():
            return
        self._loading_overlay.setVisible(True)
        self._gen_worker = AbrechnungGenerateWorker(self._backend, monat, parent=self)
        self._gen_worker.finished.connect(self._on_generate_finished)
        self._gen_worker.start()

    def _on_generate_finished(self, resp, error_msg: str):
        self._loading_overlay.setVisible(False)
        if error_msg:
            logger.warning(f"Abrechnung-Generierung fehlgeschlagen: {error_msg}")
            if self._toast_manager:
                self._toast_manager.show_error(error_msg)
            return
        if resp:
            monat = self._monat_combo.currentData() or datetime.now().strftime('%Y-%m')
            if self._toast_manager:
                self._toast_manager.show_success(texts.PROVISION_TOAST_GENERATE_DONE.format(monat=monat))
            self._load_data()

    def _change_status(self, abrechnung_id: int, status: str):
        if hasattr(self, '_status_worker') and self._status_worker and self._status_worker.isRunning():
            return
        self._status_worker = AbrechnungStatusWorker(
            self._backend, abrechnung_id, status, parent=self)
        self._status_worker.finished.connect(self._on_status_changed)
        self._status_worker.start()

    def _on_status_changed(self, ok: bool, status: str, error_msg: str):
        if ok:
            if self._toast_manager:
                label = STATUS_LABELS.get(status, status)
                self._toast_manager.show_success(texts.PROVISION_TOAST_STATUS_CHANGED.format(status=label))
            self._load_data()
        else:
            msg = error_msg or texts.PROVISION_TOAST_STATUS_ERROR
            logger.warning(f"Statusaenderung fehlgeschlagen: {msg}")
            if self._toast_manager:
                self._toast_manager.show_error(msg)
            self._load_data()

    def _export_csv(self):
        path, _ = QFileDialog.getSaveFileName(self, texts.PROVISION_PAY_EXPORT_CSV_TITLE, "", "CSV (*.csv)")
        if not path:
            return
        headers = [self._model.COLUMNS[i] for i in range(self._model.columnCount() - 1)]
        data = []
        for row in range(self._model.rowCount()):
            row_data = []
            for col in range(self._model.columnCount() - 1):
                val = self._model.data(self._model.index(row, col))
                row_data.append(val or "")
            data.append(row_data)

        self._loading_overlay.setVisible(True)

        def _write_csv(w, p=path, h=headers, d=data):
            with open(p, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, delimiter=';')
                writer.writerow(h)
                writer.writerows(d)
            return p

        run_worker(
            self, _write_csv,
            lambda p: self._on_export_done(p),
            on_error=self._on_export_error,
        )

    def _on_export_done(self, path):
        self._loading_overlay.setVisible(False)
        if self._toast_manager:
            self._toast_manager.show_success(
                texts.PROVISION_TOAST_EXPORT_DONE.format(path=os.path.basename(path)),
                action_text=texts.PROVISION_TOAST_EXPORT_OPEN,
                action_callback=lambda: os.startfile(path),
            )

    def _on_export_error(self, error_msg):
        self._loading_overlay.setVisible(False)
        logger.error(f"Export-Fehler: {error_msg}")
        if self._toast_manager:
            self._toast_manager.show_error(error_msg)

    def _export_xlsx(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            if self._toast_manager:
                self._toast_manager.show_error(texts.PROVISION_PAY_OPENPYXL_MISSING)
            return

        path, _ = QFileDialog.getSaveFileName(self, texts.PROVISION_PAY_EXPORT_XLSX_TITLE, "", "Excel (*.xlsx)")
        if not path:
            return

        self._loading_overlay.setVisible(True)

        monat = self._monat_combo.currentData() or datetime.now().strftime('%Y-%m')
        headers = [self._model.COLUMNS[i] for i in range(self._model.columnCount() - 1)]
        eur_cols = {
            self._model.COL_BRUTTO, self._model.COL_TL,
            self._model.COL_NETTO, self._model.COL_RUECK,
            self._model.COL_VU_ABZUG, self._model.COL_AUSZAHLUNG,
        }
        items = [self._model.get_item(r) for r in range(self._model.rowCount())]
        items = [i for i in items if i is not None]

        def _write_xlsx(w, p=path, m=monat, h=headers, ec=eur_cols, it=items):
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = texts.PROVISION_PAY_EXCEL_SHEET.format(monat=m)

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")

            for col_idx, header in enumerate(h, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row_idx, item in enumerate(it):
                values = [
                    item.berater_name,
                    item.berater_role,
                    item.brutto_provision,
                    item.tl_abzug,
                    item.netto_provision,
                    item.rueckbelastungen,
                    item.vu_abzug_summe,
                    item.auszahlung,
                    item.anzahl_provisionen,
                    STATUS_LABELS.get(item.status, item.status),
                    item.revision,
                ]
                for col_idx, val in enumerate(values):
                    cell = ws.cell(row=row_idx + 2, column=col_idx + 1, value=val)
                    if col_idx in ec:
                        cell.number_format = '#,##0.00 €'
                        cell.alignment = Alignment(horizontal="right")

            for col_idx in range(1, len(h) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 16

            wb.save(p)
            return p

        run_worker(
            self, _write_xlsx,
            lambda p: self._on_export_done(p),
            on_error=self._on_export_error,
        )

    # ── Statement-Export (Einzelabrechnung PDF/Excel/Word) ──

    def _get_selected_berater(self) -> Optional['BeraterAbrechnung']:
        indexes = self._table.selectionModel().selectedRows()
        if not indexes:
            return None
        return self._model.get_item(indexes[0].row())

    def _export_statement_selected(self, fmt: str):
        item = self._get_selected_berater()
        if not item:
            if self._toast_manager:
                self._toast_manager.show_warning(texts.PM_STMT_NO_DATA)
            return
        self._export_statement_for_berater(item, fmt)

    def _export_statement_for_berater(self, berater: 'BeraterAbrechnung', fmt: str):
        if not self._presenter:
            if self._toast_manager:
                self._toast_manager.show_error(texts.PM_STMT_NO_DATA)
            return

        default_name = get_statement_filename(berater, EXTENSIONS[fmt])
        file_filter = FILE_FILTERS[fmt]
        path, _ = QFileDialog.getSaveFileName(
            self, texts.PM_STMT_EXPORT_MENU, default_name, file_filter)
        if not path:
            return

        self._loading_overlay.setVisible(True)
        self._stmt_worker = StatementExportWorker(
            self._presenter, berater, fmt, path)
        self._stmt_worker.finished.connect(self._on_stmt_export_done)
        self._stmt_worker.error.connect(self._on_stmt_export_error)
        self._stmt_worker.start()

    def _on_stmt_export_done(self, path: str):
        self._loading_overlay.setVisible(False)
        if self._toast_manager:
            self._toast_manager.show_success(
                texts.PM_STMT_TOAST_SUCCESS.format(path=os.path.basename(path)))

    def _on_stmt_export_error(self, error: str):
        self._loading_overlay.setVisible(False)
        logger.error(f"Statement-Export-Fehler: {error}")
        if self._toast_manager:
            self._toast_manager.show_error(
                texts.PM_STMT_TOAST_ERROR.format(error=error))

    def _export_all_statements(self, fmt: str):
        if not self._presenter:
            return
        all_data = getattr(self, '_all_data', [])
        if not all_data:
            if self._toast_manager:
                self._toast_manager.show_warning(texts.PM_STMT_NO_DATA)
            return

        folder = QFileDialog.getExistingDirectory(
            self, texts.PM_STMT_FOLDER_TITLE)
        if not folder:
            return

        self._loading_overlay.setVisible(True)
        self._batch_worker = StatementBatchExportWorker(
            self._presenter, all_data, fmt, folder)
        self._batch_worker.finished.connect(
            lambda count: self._on_batch_export_done(count, folder))
        self._batch_worker.progress.connect(self._on_batch_progress)
        self._batch_worker.error.connect(self._on_stmt_export_error)
        self._batch_worker.start()

    def _on_batch_export_done(self, count: int, folder: str):
        self._loading_overlay.setVisible(False)
        if self._toast_manager:
            self._toast_manager.show_success(
                texts.PM_STMT_TOAST_BATCH_SUCCESS.format(
                    count=count, folder=os.path.basename(folder)))

    def _on_batch_progress(self, current: int, total: int):
        pass

    # ── E-Mail-Versand ──

    def _send_email_selected(self):
        item = self._get_selected_berater()
        if not item:
            if self._toast_manager:
                self._toast_manager.show_warning(texts.PM_STMT_NO_DATA)
            return
        self._send_email_for_berater(item)

    def _send_email_for_berater(self, berater: 'BeraterAbrechnung'):
        if not self._presenter:
            return
        if not berater.has_email:
            if self._toast_manager:
                self._toast_manager.show_warning(
                    texts.PM_STMT_EMAIL_TOAST_NO_ADDR.format(name=berater.berater_name))
            return

        self._loading_overlay.setVisible(True)
        self._email_worker = StatementEmailWorker(self._presenter, berater)
        self._email_worker.finished.connect(
            lambda result, b=berater: self._on_email_done(result, b))
        self._email_worker.error.connect(self._on_email_error)
        self._email_worker.start()

    def _on_email_done(self, result: dict, berater: 'BeraterAbrechnung'):
        self._loading_overlay.setVisible(False)
        if result.get('success'):
            if self._toast_manager:
                self._toast_manager.show_success(
                    texts.PM_STMT_EMAIL_TOAST_SUCCESS.format(
                        email=result.get('recipient', berater.berater_email or '')))
            self._load_data()
        else:
            error = result.get('error', '')
            if self._toast_manager:
                self._toast_manager.show_error(
                    texts.PM_STMT_EMAIL_TOAST_ERROR.format(error=error))

    def _on_email_error(self, error: str):
        self._loading_overlay.setVisible(False)
        logger.error(f"E-Mail-Versand-Fehler: {error}")
        if self._toast_manager:
            self._toast_manager.show_error(
                texts.PM_STMT_EMAIL_TOAST_ERROR.format(error=error))

    def _send_email_all(self):
        if not self._presenter:
            return
        all_data = getattr(self, '_all_data', [])
        if not all_data:
            if self._toast_manager:
                self._toast_manager.show_warning(texts.PM_STMT_NO_DATA)
            return

        eligible = [b for b in all_data if b.has_email]
        if not eligible:
            if self._toast_manager:
                self._toast_manager.show_warning(texts.PM_STMT_EMAIL_NO_ELIGIBLE)
            return

        result = QMessageBox.question(
            self,
            texts.PM_STMT_EMAIL_SEND_ALL,
            texts.PM_STMT_EMAIL_CONFIRM_BATCH,
        )
        if result != QMessageBox.Yes:
            return

        self._loading_overlay.setVisible(True)
        self._batch_email_worker = StatementBatchEmailWorker(
            self._presenter, all_data)
        self._batch_email_worker.finished.connect(self._on_batch_email_done)
        self._batch_email_worker.progress.connect(self._on_batch_email_progress)
        self._batch_email_worker.error.connect(self._on_email_error)
        self._batch_email_worker.start()

    def _on_batch_email_done(self, sent: int, failed: int):
        self._loading_overlay.setVisible(False)
        if self._toast_manager:
            self._toast_manager.show_success(
                texts.PM_STMT_EMAIL_TOAST_BATCH.format(sent=sent, failed=failed))
        self._load_data()

    def _on_batch_email_progress(self, current: int, total: int):
        if self._toast_manager:
            self._toast_manager.show_info(
                texts.PM_STMT_EMAIL_TOAST_BATCH_PROGRESS.format(
                    current=current, total=total))

import os
import uuid
import json
from datetime import datetime
import openpyxl
import hashlib
import requests
import base64
from abc import ABC, abstractmethod
from threading import Lock
from collections import Counter
from flask import Flask, make_response, render_template, jsonify, request, redirect, url_for, flash, send_from_directory, session, Response
import traceback
from werkzeug.security import check_password_hash, generate_password_hash
from datetime import datetime, timezone
import logging

# ==============================================================================
# --- SECURITY FIX SV-002/SV-007: Credential-Verschlüsselung ---
# ==============================================================================
# Verwendet Fernet-Verschlüsselung aus dem cryptography-Paket
# Master-Key wird aus Umgebungsvariable ACENCIA_MASTER_KEY geladen
try:
    from cryptography.fernet import Fernet, InvalidToken
    _CRYPTO_AVAILABLE = True
except ImportError:
    _CRYPTO_AVAILABLE = False
    print("[WARNUNG] cryptography nicht installiert. Credential-Verschlüsselung deaktiviert.")

# Encryption Marker für verschlüsselte Werte
_ENCRYPTED_PREFIX = "ENC:"

def _get_fernet_key():
    """
    Holt oder generiert den Fernet-Verschlüsselungsschlüssel.
    
    Der Schlüssel wird aus ACENCIA_MASTER_KEY geladen.
    Falls nicht gesetzt, wird eine Warnung ausgegeben.
    
    Returns:
        bytes | None: Der Fernet-Schlüssel oder None
    """
    master_key = os.environ.get('ACENCIA_MASTER_KEY')
    if not master_key:
        return None
    # Fernet benötigt einen 32-Byte base64-codierten Schlüssel
    # Wir hashen den Master-Key und codieren ihn entsprechend
    key_hash = hashlib.sha256(master_key.encode()).digest()
    return base64.urlsafe_b64encode(key_hash)

def encrypt_credential(plaintext: str) -> str:
    """
    Verschlüsselt einen Credential-String.
    
    Args:
        plaintext (str): Der zu verschlüsselnde Klartext
    
    Returns:
        str: Der verschlüsselte String mit ENC: Präfix, oder Klartext wenn Verschlüsselung nicht verfügbar
    """
    if not plaintext:
        return plaintext
    if not _CRYPTO_AVAILABLE:
        return plaintext
    
    key = _get_fernet_key()
    if not key:
        print("[WARNUNG] ACENCIA_MASTER_KEY nicht gesetzt. Credential wird unverschlüsselt gespeichert.")
        return plaintext
    
    try:
        f = Fernet(key)
        encrypted = f.encrypt(plaintext.encode()).decode()
        return f"{_ENCRYPTED_PREFIX}{encrypted}"
    except Exception as e:
        print(f"[FEHLER] Verschlüsselung fehlgeschlagen: {e}")
        return plaintext

def decrypt_credential(ciphertext: str) -> str:
    """
    Entschlüsselt einen Credential-String.
    
    Args:
        ciphertext (str): Der verschlüsselte String (mit ENC: Präfix)
    
    Returns:
        str: Der entschlüsselte Klartext, oder der Original-String wenn nicht verschlüsselt
    """
    if not ciphertext:
        return ciphertext
    if not ciphertext.startswith(_ENCRYPTED_PREFIX):
        return ciphertext  # Nicht verschlüsselt
    if not _CRYPTO_AVAILABLE:
        print("[FEHLER] Verschlüsselter Wert gefunden, aber cryptography nicht installiert!")
        return ""
    
    key = _get_fernet_key()
    if not key:
        print("[FEHLER] Verschlüsselter Wert gefunden, aber ACENCIA_MASTER_KEY nicht gesetzt!")
        return ""
    
    try:
        f = Fernet(key)
        encrypted_data = ciphertext[len(_ENCRYPTED_PREFIX):]
        return f.decrypt(encrypted_data.encode()).decode()
    except InvalidToken:
        print("[FEHLER] Entschlüsselung fehlgeschlagen - falscher Master-Key oder korrupte Daten!")
        return ""
    except Exception as e:
        print(f"[FEHLER] Entschlüsselung fehlgeschlagen: {e}")
        return ""

def is_encrypted(value: str) -> bool:
    """Prüft ob ein Wert bereits verschlüsselt ist."""
    return value and value.startswith(_ENCRYPTED_PREFIX)

# --- Robust Logging Setup ---

# 1. Get Project Root & Log File Path
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
LOG_FILE_PATH = os.path.join(PROJECT_ROOT, 'server.log')

# 2. Configure File Logger
# This logger is responsible ONLY for writing to server.log
file_logger = logging.getLogger('file_logger')
file_logger.setLevel(logging.INFO)
# Prevent logs from propagating to the root logger
file_logger.propagate = False

# ==============================================================================
# --- SECURITY FIX SV-010: Log-Rotation mit RotatingFileHandler ---
# ==============================================================================
# Verwendet RotatingFileHandler statt FileHandler um Disk-Overflow zu verhindern.
# Konfiguration: maxBytes=10MB, backupCount=5 (behält 5 alte Logdateien)
from logging.handlers import RotatingFileHandler

# Make sure we don't add handlers again if this code is re-run (e.g., by a reloader)
if not file_logger.handlers:
    file_handler = RotatingFileHandler(
        LOG_FILE_PATH,
        mode='a',
        maxBytes=10*1024*1024,  # 10 MB pro Datei
        backupCount=5,          # Behalte 5 Backup-Dateien
        encoding='utf-8'
    )
    # The custom_log function will provide the full formatted string, so the formatter is simple.
    file_formatter = logging.Formatter('%(message)s')
    file_handler.setFormatter(file_formatter)
    file_logger.addHandler(file_handler)

# ==============================================================================
# --- SECURITY FIX SV-016: Audit-Logger für administrative Aktionen ---
# ==============================================================================
# Separater Logger für Sicherheits-relevante administrative Aktionen
AUDIT_LOG_PATH = os.path.join(PROJECT_ROOT, 'audit.log')
audit_logger = logging.getLogger('audit')
audit_logger.setLevel(logging.INFO)
audit_logger.propagate = False

if not audit_logger.handlers:
    audit_handler = RotatingFileHandler(
        AUDIT_LOG_PATH,
        mode='a',
        maxBytes=10*1024*1024,  # 10 MB pro Datei
        backupCount=10,          # Behalte 10 Backup-Dateien (mehr für Audit)
        encoding='utf-8'
    )
    audit_formatter = logging.Formatter('%(asctime)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
    audit_handler.setFormatter(audit_formatter)
    audit_logger.addHandler(audit_handler)

def audit_log(user: str, action: str, target: str, details: str = "", ip: str = ""):
    """
    Protokolliert administrative Aktionen für Audit-Zwecke (SV-016).
    
    Args:
        user (str): Der Benutzer, der die Aktion ausführt
        action (str): Die Art der Aktion (z.B. CREATE_USER, DELETE_USER)
        target (str): Das Ziel der Aktion (z.B. Benutzername, Arbeitgeber-ID)
        details (str): Zusätzliche Details zur Aktion
        ip (str): Die IP-Adresse des Benutzers
    """
    msg = f"USER={user} ACTION={action} TARGET={target}"
    if details:
        msg += f" DETAILS={details}"
    if ip:
        msg += f" IP={ip}"
    audit_logger.info(msg)

# 3. Configure Werkzeug logger to use our file handler
# This captures server access logs (GET, POST, etc.) into our file.
werkzeug_log = logging.getLogger('werkzeug')
# Set level to INFO to capture standard access log messages like "GET / HTTP/1.1" 200 -
werkzeug_log.setLevel(logging.INFO)
werkzeug_log.handlers.clear()
werkzeug_log.addHandler(file_handler)


# ==============================================================================
# --- SECURITY FIX SV-014: PII-Anonymisierung in Logs ---
# ==============================================================================
# Konfiguration: Wenn True, werden Personennamen in Logs anonymisiert
_ANONYMIZE_PII_IN_LOGS = os.environ.get('ANONYMIZE_PII_LOGS', 'true').lower() == 'true'

def anonymize_name(name: str) -> str:
    """
    Anonymisiert einen Personennamen für das Logging.
    
    Args:
        name (str): Der vollständige Name
    
    Returns:
        str: Der anonymisierte Name (z.B. "Max M." statt "Max Mustermann")
    """
    if not name or not _ANONYMIZE_PII_IN_LOGS:
        return name
    parts = name.strip().split()
    if len(parts) == 0:
        return name
    if len(parts) == 1:
        return parts[0][:3] + "***" if len(parts[0]) > 3 else parts[0]
    # Vorname + erster Buchstabe des Nachnamens
    return f"{parts[0]} {parts[-1][0]}."

def anonymize_log_message(message: str) -> str:
    """
    Versucht, PII in einer Log-Nachricht zu anonymisieren.
    Erkennt Muster wie "Mitarbeiter 'Name'" oder "'Name' wurde...".
    
    Args:
        message (str): Die Original-Nachricht
    
    Returns:
        str: Die anonymisierte Nachricht
    """
    if not _ANONYMIZE_PII_IN_LOGS:
        return message
    import re
    # Einfache Anonymisierung: Ersetze Namen in Anführungszeichen durch anonymisierte Version
    def replace_quoted_name(match):
        name = match.group(1)
        # Prüfe ob es wie ein Name aussieht (enthält Leerzeichen und Buchstaben)
        if ' ' in name and name.replace(' ', '').isalpha():
            return f"'{anonymize_name(name)}'"
        return match.group(0)
    
    return re.sub(r"'([^']+)'", replace_quoted_name, message)

# ANSI colors for the custom logger
COLORS = {
    "red": "\033[31m",
    "green": "\033[32m",
    "yellow": "\033[33m",
    "blue": "\033[34m",
    "magenta": "\033[35m",
    "cyan": "\033[36m",
    "white": "\033[37m",
    "reset": "\033[0m"
}

def custom_log(kuerzel, message, color_name=None, ip=None):
    """
    Protokolliert eine benutzerdefinierte Nachricht sowohl in der Konsole (mit Farbe) als auch in der Log-Datei (ohne Farbe).
    
    SECURITY FIX SV-014: Nachrichten werden automatisch auf PII geprüft und anonymisiert.
    
    Args:
        kuerzel (str): Das Kürzel des Benutzers für die Protokollierung
        message (str): Die zu protokollierende Nachricht
        color_name (str, optional): Der Name der Farbe für die Konsolenausgabe
        ip (str, optional): Die IP-Adresse für unauthentifizierte Zugriffe
    
    Returns:
        None
    """
    now = datetime.now().strftime("%d/%b/%Y %H:%M:%S")
    
    # SV-014: PII-Anonymisierung anwenden
    message = anonymize_log_message(message)

    if ip:
        # Format for unauthenticated access, e.g., ("127.0.0.1" - LOGIN)
        log_message_plain = f'("{ip}" - {message})'
        log_message_color = log_message_plain # No color for IP logs
    else:
        color = COLORS.get(color_name, COLORS["white"])
        colored_kuerzel = f"{color}{kuerzel}{COLORS['reset']}"
        log_message_plain = f'("{kuerzel}" - {message})'
        log_message_color = f'("{colored_kuerzel}" - {message})'

    # Construct the final, full log lines
    full_log_plain = f'[{now}] {log_message_plain}'
    full_log_color = f'[{now}] {log_message_color}'

    # Log plain message to file
    file_logger.info(full_log_plain)
    # Print colored message to console
    print(full_log_color)


# ==============================================================================
# --- SECTION 1: HELPER FUNCTIONS ---
# ==============================================================================

# --- SCS: fixe Ziel-Header (stabil & in Import-Reihenfolge) ---
SCS_HEADERS = [
    "Name","Vorname","Geschlecht","Titel","Geburtsdatum",
    "Strasse","Hausnummer","PLZ","Ort","Land","Kommentar",
    "Email","Telefon","Personalnummer","Position","Firmeneintritt",
    "Bruttogehalt","VWL","geldwerterVorteil","SteuerfreibetragJahr","SteuerfreibetragMonat",
    "SV_Brutto","Steuerklasse","Religion","Kinder","Abteilung","Arbeitsplatz","Arbeitgeber",
    "Status"
]

def _get_from_path(obj: dict, *paths, default=None):
    """
    Greift sicher auf verschachtelte Schlüssel in einem Dictionary zu.
    
    Args:
        obj (dict): Das Dictionary, aus dem Werte extrahiert werden sollen
        *paths: Variable Anzahl von Pfaden, die als Strings ('a.b.c') oder Tupel ('a', 'b', 'c') übergeben werden können
        default: Der Standardwert, der zurückgegeben wird, wenn kein Pfad gefunden wird
    
    Returns:
        Der gefundene Wert oder der Standardwert, wenn kein Pfad existiert
    """
    for p in paths:
        if not p:
            continue
        keys = p if isinstance(p, (list, tuple)) else str(p).split(".")
        cur = obj
        ok = True
        for k in keys:
            if not isinstance(cur, dict):
                ok = False
                break
            cur = cur.get(k)
            if cur is None:
                ok = False
                break
        if ok and cur not in (None, ""):
            return cur
    return default

def _getv(e: dict, details: dict | None, label: str | None, *flat_paths, default: str = "") -> str:
    """
    Extrahiert einen Wert aus einem Mitarbeiter-Dictionary mit bevorzugter Suche in Details-Labels.
    
    Args:
        e (dict): Das Mitarbeiter-Dictionary
        details (dict | None): Das Details-Dictionary mit strukturierten Labels
        label (str | None): Das bevorzugte Label für die Suche
        *flat_paths: Variable Anzahl von flachen Pfaden als Fallback
        default (str): Der Standardwert, der zurückgegeben wird
    
    Returns:
        str: Der gefundene Wert als String oder der Standardwert
    """
    v = None
    if details and label:
        v = _get_value_from_details(details, label, default=None)
    if v in (None, ""):
        v = _get_from_path(e, *flat_paths, default=None)
    return "" if v is None else str(v).strip()

def _get_safe_employer_name(name):
    """
    Bereinigt einen Arbeitgeber-Namen für die Verwendung in Dateinamen.
    
    Args:
        name (str): Der ursprüngliche Arbeitgeber-Name
    
    Returns:
        str: Der bereinigte Name, der sicher in Dateinamen verwendet werden kann
    """
    return "".join(c for c in name if c.isalnum() or c in (' ', '_')).rstrip().replace(' ', '_')

def _get_value_from_details(details: dict, target_label: str, default=None):
    """
    Extrahiert einen Wert aus dem Details-Dictionary basierend auf dem Label.
    
    Args:
        details (dict): Das Details-Dictionary mit strukturierten Gruppen
        target_label (str): Das gesuchte Label (case-insensitive)
        default: Der Standardwert, der zurückgegeben wird, wenn das Label nicht gefunden wird
    
    Returns:
        Der gefundene Wert oder der Standardwert
    """
    target_label = target_label.lower()
    for group_items in details.values():
        for item in group_items:
            if item.get('label', '').lower() == target_label:
                return item.get('value', default)
    return default

def _parse_date(date_str: str) -> datetime | None:
    """
    Parst einen Datumsstring aus gängigen Formaten in ein datetime-Objekt.
    
    Args:
        date_str (str): Der zu parsende Datumsstring
    
    Returns:
        datetime | None: Das geparste datetime-Objekt oder None bei Fehlern
    """
    if not date_str or not isinstance(date_str, str): return None
    for fmt in ('%d.%m.%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(date_str.split('T')[0], fmt)
        except ValueError:
            pass
    return None

def _format_date_for_display(date_str: str | None) -> str | None:
    """
    Formatiert einen Datumsstring in das DD.MM.YYYY Format für die Anzeige.
    
    Args:
        date_str (str | None): Der zu formatierende Datumsstring
    
    Returns:
        str | None: Der formatierte Datumsstring oder None bei Fehlern
    """
    dt = _parse_date(date_str)
    return dt.strftime('%d.%m.%Y') if dt else date_str

def save_history_entry(history_dir: str, employer_cfg: dict, provider_response: dict | list):
    """
    Speichert die rohe Provider-API-Antwort im Verzeichnis für die Historie.
    
    Args:
        history_dir (str): Das Verzeichnis, in dem die Historie gespeichert werden soll
        employer_cfg (dict): Die Arbeitgeber-Konfiguration mit Name und Provider-Key
        provider_response (dict | list): Die rohe API-Antwort des Providers
    
    Returns:
        None
    """
    employer_name = employer_cfg.get('name', 'unknown_employer')
    provider_key = employer_cfg.get('provider_key', 'unknown_provider')
    safe_emp_name = _get_safe_employer_name(employer_name)
    # Use a more unique timestamp to avoid collisions on rapid calls
    timestamp = datetime.now().strftime('%Y%m%d-%H%M%S-%f')
    filename = f"{safe_emp_name}-{provider_key}-history-{timestamp}.json"
    filepath = os.path.join(history_dir, filename)
    try:
        # Ensure the directory exists, as it might be the first run
        os.makedirs(history_dir, exist_ok=True)
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(provider_response, f, ensure_ascii=False, indent=2)
    except Exception as e:
        # Log the error but don't crash the main process
        print(f"ERROR: Could not write history file to {filepath}: {e}")

# ==============================================================================
# --- SECTION 2: CORE CLASSES (DATA STORE & PROVIDERS) ---
# ==============================================================================

class EmployerStore:
    """
    Singleton-Klasse zur Verwaltung der Arbeitgeber-Datenpersistenz in einer JSON-Datei.
    
    Diese Klasse implementiert das Singleton-Pattern und verwaltet alle CRUD-Operationen
    für Arbeitgeber-Daten in einer JSON-Datei.
    """
    _instance, _lock = None, Lock()
    def __new__(cls, *args, **kwargs):
        """
        Implementiert das Singleton-Pattern für die EmployerStore-Klasse.
        
        Args:
            *args: Variable Argumente
            **kwargs: Variable Schlüsselwort-Argumente
        
        Returns:
            EmployerStore: Die einzige Instanz der Klasse
        """
        if not cls._instance:
            with cls._lock:
                if not cls._instance: cls._instance = super(EmployerStore, cls).__new__(cls)
        return cls._instance
    def __init__(self, filepath='employers.json'):
        """
        Initialisiert die EmployerStore-Instanz.
        
        Args:
            filepath (str): Der Pfad zur JSON-Datei für die Arbeitgeber-Daten
        """
        if not hasattr(self, 'initialized'):
            self.filepath = filepath
            self.initialized = True
            if not os.path.exists(self.filepath): self._write_data([])
    def _read_data(self):
        """
        Liest die Arbeitgeber-Daten aus der JSON-Datei.
        
        SECURITY FIX SV-002: Credentials werden automatisch entschlüsselt.
        
        Returns:
            list: Liste der Arbeitgeber-Daten mit entschlüsselten Credentials
        """
        with self._lock:
            try:
                with open(self.filepath, 'r', encoding='utf-8') as f:
                    employers = json.load(f)
                # SV-002: Credentials entschlüsseln
                for employer in employers:
                    if 'access_key' in employer and employer['access_key']:
                        employer['access_key'] = decrypt_credential(employer['access_key'])
                    if 'secret_key' in employer and employer['secret_key']:
                        employer['secret_key'] = decrypt_credential(employer['secret_key'])
                return employers
            except (FileNotFoundError, json.JSONDecodeError): return []
    
    def _write_data(self, data):
        """
        Schreibt die Arbeitgeber-Daten in die JSON-Datei.
        
        SECURITY FIX SV-002: Credentials werden automatisch verschlüsselt.
        
        Args:
            data: Die zu schreibenden Daten (mit Klartext-Credentials)
        
        Returns:
            None
        """
        import copy
        # Deep-Copy erstellen um Original-Daten nicht zu ändern
        data_to_save = copy.deepcopy(data)
        # SV-002: Credentials verschlüsseln
        for employer in data_to_save:
            if 'access_key' in employer and employer['access_key']:
                if not is_encrypted(employer['access_key']):
                    employer['access_key'] = encrypt_credential(employer['access_key'])
            if 'secret_key' in employer and employer['secret_key']:
                if not is_encrypted(employer['secret_key']):
                    employer['secret_key'] = encrypt_credential(employer['secret_key'])
        with self._lock:
            with open(self.filepath, 'w', encoding='utf-8') as f:
                json.dump(data_to_save, f, indent=4, ensure_ascii=False)
    def get_all(self):
        """
        Ruft alle Arbeitgeber-Daten ab.
        
        Returns:
            list: Liste aller Arbeitgeber
        """
        return self._read_data()
    def get_by_id(self, employer_id):
        """
        Ruft einen Arbeitgeber anhand seiner ID ab.
        
        Args:
            employer_id (str): Die ID des gesuchten Arbeitgebers
        
        Returns:
            dict | None: Die Arbeitgeber-Daten oder None, wenn nicht gefunden
        """
        for employer in self._read_data():
            if employer.get('id') == employer_id: return employer
        return None
    def add(self, employer_data):
        """
        Fügt einen neuen Arbeitgeber hinzu.
        
        Args:
            employer_data (dict): Die Daten des neuen Arbeitgebers
        
        Returns:
            dict: Die hinzugefügten Arbeitgeber-Daten
        """
        employers = self._read_data()
        employers.append(employer_data)
        self._write_data(employers)
        return employer_data
    def delete(self, employer_id):
        """
        Löscht einen Arbeitgeber anhand seiner ID.
        
        Args:
            employer_id (str): Die ID des zu löschenden Arbeitgebers
        
        Returns:
            bool: True, wenn der Arbeitgeber gelöscht wurde, False wenn nicht gefunden
        """
        employers = self._read_data()
        original_length = len(employers)
        employers = [emp for emp in employers if emp.get('id') != employer_id]
        if len(employers) < original_length:
            self._write_data(employers)
            return True
        return False

    def update(self, employer_id, updated_data):
        """
        Aktualisiert die Daten eines Arbeitgebers.
        
        Args:
            employer_id (str): Die ID des zu aktualisierenden Arbeitgebers
            updated_data (dict): Die neuen Daten, die aktualisiert werden sollen
        
        Returns:
            bool: True, wenn der Arbeitgeber aktualisiert wurde, False wenn nicht gefunden
        """
        employers = self._read_data()
        updated = False
        for i, employer in enumerate(employers):
            if employer.get('id') == employer_id:
                employers[i].update(updated_data)
                updated = True
                break
        if updated:
            self._write_data(employers)
        return updated


# ==============================================================================
# --- TRIGGER SYSTEM: Data Store Classes ---
# ==============================================================================
# Diese Klassen verwalten die Trigger-Konfiguration und das Ausführungsprotokoll.
# Trigger werden bei Delta-Exporten ausgewertet und können E-Mails oder API-Calls auslösen.

class TriggerStore:
    """
    Singleton-Klasse zur Verwaltung der Trigger-Konfiguration und SMTP-Einstellungen.
    
    Trigger sind regelbasierte Automatisierungen, die bei Änderungen an Mitarbeiterdaten
    (erkannt im Delta-Export) ausgelöst werden. Sie basieren auf den normalisierten
    SCS-Feldern und funktionieren daher provider-unabhängig.
    
    Datenstruktur:
    - smtp_config: Globale SMTP-Konfiguration für E-Mail-Versand
    - triggers: Liste der Trigger-Definitionen
    
    Trigger-Events:
    - employee_changed: Mitarbeiterdaten wurden geändert
    - employee_added: Neuer Mitarbeiter wurde hinzugefügt
    - employee_removed: Mitarbeiter wurde entfernt
    
    Bedingungsoperatoren (für employee_changed):
    - changed: Feld hat sich geändert (beliebig)
    - changed_to: Feld wurde zu bestimmtem Wert geändert
    - changed_from: Feld wurde von bestimmtem Wert geändert
    - changed_from_to: Feld wurde von Wert A zu Wert B geändert
    - is_empty: Feld ist jetzt leer
    - is_not_empty: Feld ist jetzt befüllt
    - contains: Feld enthält Substring
    
    Erweiterbarkeit:
    - Neue Trigger-Events können durch Erweiterung der TriggerEngine hinzugefügt werden
    - Neue Aktionen können durch das ActionRegistry-Pattern registriert werden
    """
    _instance, _lock = None, Lock()
    
    # Verfügbare Trigger-Events
    TRIGGER_EVENTS = ['employee_changed', 'employee_added', 'employee_removed']
    
    # Verfügbare Bedingungsoperatoren
    CONDITION_OPERATORS = ['changed', 'changed_to', 'changed_from', 'changed_from_to', 'is_empty', 'is_not_empty', 'contains']
    
    # Verfügbare Aktionstypen
    ACTION_TYPES = ['email', 'api']
    
    def __new__(cls, *args, **kwargs):
        """Implementiert das Singleton-Pattern für die TriggerStore-Klasse."""
        if not cls._instance:
            with cls._lock:
                if not cls._instance:
                    cls._instance = super(TriggerStore, cls).__new__(cls)
        return cls._instance
    
    def __init__(self, filepath=None):
        """
        Initialisiert die TriggerStore-Instanz.
        
        Args:
            filepath (str): Der Pfad zur JSON-Datei für die Trigger-Daten
        """
        if not hasattr(self, 'initialized'):
            if filepath is None:
                # Default path relative to app.py
                filepath = os.path.join(os.path.dirname(__file__), 'data', 'triggers.json')
            self.filepath = filepath
            self.initialized = True
            if not os.path.exists(self.filepath):
                self._write_data({'smtp_config': {
                    'host': '', 'port': 587, 'username': '', 'password': '',
                    'from_email': '', 'use_tls': True
                }, 'triggers': []})
    
    def _read_data(self):
        """
        Liest die Trigger-Daten aus der JSON-Datei.
        
        SMTP-Credentials werden automatisch entschlüsselt.
        
        Returns:
            dict: Die Trigger-Konfiguration
        """
        with self._lock:
            try:
                with open(self.filepath, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                # Decrypt SMTP credentials
                if 'smtp_config' in data:
                    if data['smtp_config'].get('username'):
                        data['smtp_config']['username'] = decrypt_credential(data['smtp_config']['username'])
                    if data['smtp_config'].get('password'):
                        data['smtp_config']['password'] = decrypt_credential(data['smtp_config']['password'])
                # Decrypt API credentials in triggers
                for trigger in data.get('triggers', []):
                    if trigger.get('action', {}).get('type') == 'api':
                        auth = trigger['action'].get('config', {}).get('auth', {})
                        if auth.get('token'):
                            auth['token'] = decrypt_credential(auth['token'])
                        if auth.get('password'):
                            auth['password'] = decrypt_credential(auth['password'])
                        if auth.get('api_key'):
                            auth['api_key'] = decrypt_credential(auth['api_key'])
                return data
            except (FileNotFoundError, json.JSONDecodeError):
                return {'smtp_config': {}, 'triggers': []}
    
    def _write_data(self, data):
        """
        Schreibt die Trigger-Daten in die JSON-Datei.
        
        Credentials werden automatisch verschlüsselt.
        
        Args:
            data: Die zu schreibenden Daten
        """
        import copy
        data_to_save = copy.deepcopy(data)
        # Encrypt SMTP credentials
        if 'smtp_config' in data_to_save:
            if data_to_save['smtp_config'].get('username') and not is_encrypted(data_to_save['smtp_config']['username']):
                data_to_save['smtp_config']['username'] = encrypt_credential(data_to_save['smtp_config']['username'])
            if data_to_save['smtp_config'].get('password') and not is_encrypted(data_to_save['smtp_config']['password']):
                data_to_save['smtp_config']['password'] = encrypt_credential(data_to_save['smtp_config']['password'])
        # Encrypt API credentials in triggers
        for trigger in data_to_save.get('triggers', []):
            if trigger.get('action', {}).get('type') == 'api':
                auth = trigger['action'].get('config', {}).get('auth', {})
                if auth.get('token') and not is_encrypted(auth['token']):
                    auth['token'] = encrypt_credential(auth['token'])
                if auth.get('password') and not is_encrypted(auth['password']):
                    auth['password'] = encrypt_credential(auth['password'])
                if auth.get('api_key') and not is_encrypted(auth['api_key']):
                    auth['api_key'] = encrypt_credential(auth['api_key'])
        with self._lock:
            # Ensure directory exists
            os.makedirs(os.path.dirname(self.filepath), exist_ok=True)
            with open(self.filepath, 'w', encoding='utf-8') as f:
                json.dump(data_to_save, f, indent=2, ensure_ascii=False)
    
    # --- SMTP Configuration ---
    def get_smtp_config(self):
        """Ruft die SMTP-Konfiguration ab."""
        return self._read_data().get('smtp_config', {})
    
    def update_smtp_config(self, smtp_config):
        """
        Aktualisiert die SMTP-Konfiguration.
        
        Args:
            smtp_config (dict): Die neue SMTP-Konfiguration
        """
        data = self._read_data()
        data['smtp_config'] = smtp_config
        self._write_data(data)
    
    # --- Trigger CRUD Operations ---
    def get_all_triggers(self):
        """Ruft alle Trigger ab."""
        return self._read_data().get('triggers', [])
    
    def get_active_triggers(self):
        """Ruft alle aktivierten Trigger ab."""
        return [t for t in self.get_all_triggers() if t.get('enabled', True)]
    
    def get_trigger_by_id(self, trigger_id):
        """
        Ruft einen Trigger anhand seiner ID ab.
        
        Args:
            trigger_id (str): Die Trigger-ID
        
        Returns:
            dict | None: Der Trigger oder None
        """
        for trigger in self.get_all_triggers():
            if trigger.get('id') == trigger_id:
                return trigger
        return None
    
    def add_trigger(self, trigger_data, created_by='system'):
        """
        Fügt einen neuen Trigger hinzu.
        
        Args:
            trigger_data (dict): Die Trigger-Konfiguration
            created_by (str): Der Benutzer, der den Trigger erstellt
        
        Returns:
            dict: Der erstellte Trigger mit generierter ID
        """
        data = self._read_data()
        trigger_data['id'] = str(uuid.uuid4())
        trigger_data['created_at'] = datetime.now(timezone.utc).isoformat()
        trigger_data['created_by'] = created_by
        trigger_data.setdefault('enabled', True)
        trigger_data.setdefault('excluded_employers', [])
        trigger_data.setdefault('statistics', {
            'total_executions': 0,
            'last_execution': None,
            'success_count': 0,
            'error_count': 0
        })
        data['triggers'].append(trigger_data)
        self._write_data(data)
        return trigger_data
    
    def update_trigger(self, trigger_id, updated_data):
        """
        Aktualisiert einen Trigger.
        
        Args:
            trigger_id (str): Die Trigger-ID
            updated_data (dict): Die zu aktualisierenden Daten
        
        Returns:
            bool: True wenn erfolgreich, False wenn nicht gefunden
        """
        data = self._read_data()
        updated = False
        for i, trigger in enumerate(data['triggers']):
            if trigger.get('id') == trigger_id:
                # Preserve certain fields
                updated_data['id'] = trigger_id
                updated_data['created_at'] = trigger.get('created_at')
                updated_data['created_by'] = trigger.get('created_by')
                updated_data['statistics'] = trigger.get('statistics', {})
                data['triggers'][i] = updated_data
                updated = True
                break
        if updated:
            self._write_data(data)
        return updated
    
    def delete_trigger(self, trigger_id):
        """
        Löscht einen Trigger.
        
        Args:
            trigger_id (str): Die Trigger-ID
        
        Returns:
            bool: True wenn gelöscht, False wenn nicht gefunden
        """
        data = self._read_data()
        original_length = len(data['triggers'])
        data['triggers'] = [t for t in data['triggers'] if t.get('id') != trigger_id]
        if len(data['triggers']) < original_length:
            self._write_data(data)
            return True
        return False
    
    def toggle_trigger(self, trigger_id, enabled=None):
        """
        Aktiviert oder deaktiviert einen Trigger.
        
        Args:
            trigger_id (str): Die Trigger-ID
            enabled (bool | None): True/False oder None zum Umschalten
        
        Returns:
            bool | None: Der neue Status oder None wenn nicht gefunden
        """
        data = self._read_data()
        for trigger in data['triggers']:
            if trigger.get('id') == trigger_id:
                if enabled is None:
                    trigger['enabled'] = not trigger.get('enabled', True)
                else:
                    trigger['enabled'] = enabled
                self._write_data(data)
                return trigger['enabled']
        return None
    
    def exclude_employer(self, trigger_id, employer_id):
        """
        Schließt einen Arbeitgeber von einem Trigger aus.
        
        Args:
            trigger_id (str): Die Trigger-ID
            employer_id (str): Die Arbeitgeber-ID
        
        Returns:
            bool: True wenn erfolgreich
        """
        data = self._read_data()
        for trigger in data['triggers']:
            if trigger.get('id') == trigger_id:
                excluded = trigger.setdefault('excluded_employers', [])
                if employer_id not in excluded:
                    excluded.append(employer_id)
                    self._write_data(data)
                return True
        return False
    
    def include_employer(self, trigger_id, employer_id):
        """
        Entfernt einen Arbeitgeber aus der Ausschlussliste eines Triggers.
        
        Args:
            trigger_id (str): Die Trigger-ID
            employer_id (str): Die Arbeitgeber-ID
        
        Returns:
            bool: True wenn erfolgreich
        """
        data = self._read_data()
        for trigger in data['triggers']:
            if trigger.get('id') == trigger_id:
                excluded = trigger.get('excluded_employers', [])
                if employer_id in excluded:
                    excluded.remove(employer_id)
                    self._write_data(data)
                return True
        return False
    
    def update_statistics(self, trigger_id, success=True):
        """
        Aktualisiert die Statistiken eines Triggers nach einer Ausführung.
        
        Args:
            trigger_id (str): Die Trigger-ID
            success (bool): Ob die Ausführung erfolgreich war
        """
        data = self._read_data()
        for trigger in data['triggers']:
            if trigger.get('id') == trigger_id:
                stats = trigger.setdefault('statistics', {
                    'total_executions': 0, 'last_execution': None,
                    'success_count': 0, 'error_count': 0
                })
                stats['total_executions'] = stats.get('total_executions', 0) + 1
                stats['last_execution'] = datetime.now(timezone.utc).isoformat()
                if success:
                    stats['success_count'] = stats.get('success_count', 0) + 1
                else:
                    stats['error_count'] = stats.get('error_count', 0) + 1
                self._write_data(data)
                return
    
    def get_triggers_for_employer(self, employer_id):
        """
        Ruft alle Trigger ab, die für einen bestimmten Arbeitgeber aktiv sind.
        
        Args:
            employer_id (str): Die Arbeitgeber-ID
        
        Returns:
            list: Liste der aktiven Trigger für diesen Arbeitgeber
        """
        triggers = []
        for trigger in self.get_active_triggers():
            if employer_id not in trigger.get('excluded_employers', []):
                triggers.append(trigger)
        return triggers


class TriggerLogStore:
    """
    Singleton-Klasse zur Verwaltung des Trigger-Ausführungsprotokolls.
    
    Speichert alle Trigger-Ausführungen mit Details zu:
    - Welcher Trigger wurde ausgelöst
    - Welcher Arbeitgeber war betroffen
    - Welche Mitarbeiter waren betroffen (mit Änderungsdetails)
    - Welche Aktion wurde ausgeführt (E-Mail, API)
    - Status der Ausführung (success, error, skipped)
    
    Unterstützt Filterung nach:
    - Arbeitgeber
    - Trigger
    - Status
    - Zeitraum
    
    Ermöglicht Wiederholung fehlgeschlagener Ausführungen.
    """
    _instance, _lock = None, Lock()
    
    def __new__(cls, *args, **kwargs):
        """Implementiert das Singleton-Pattern."""
        if not cls._instance:
            with cls._lock:
                if not cls._instance:
                    cls._instance = super(TriggerLogStore, cls).__new__(cls)
        return cls._instance
    
    def __init__(self, filepath=None):
        """
        Initialisiert die TriggerLogStore-Instanz.
        
        Args:
            filepath (str): Der Pfad zur JSON-Datei für das Ausführungsprotokoll
        """
        if not hasattr(self, 'initialized'):
            if filepath is None:
                filepath = os.path.join(os.path.dirname(__file__), 'data', 'trigger_log.json')
            self.filepath = filepath
            self.initialized = True
            if not os.path.exists(self.filepath):
                self._write_data({'executions': []})
    
    def _read_data(self):
        """Liest das Ausführungsprotokoll aus der JSON-Datei."""
        with self._lock:
            try:
                with open(self.filepath, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except (FileNotFoundError, json.JSONDecodeError):
                return {'executions': []}
    
    def _write_data(self, data):
        """Schreibt das Ausführungsprotokoll in die JSON-Datei."""
        with self._lock:
            os.makedirs(os.path.dirname(self.filepath), exist_ok=True)
            with open(self.filepath, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
    
    def log_execution(self, trigger_id, trigger_name, event, employer_id, employer_name,
                      affected_employees, action_type, action_details, status,
                      error_message=None, executed_by='system', retry_of=None):
        """
        Protokolliert eine Trigger-Ausführung.
        
        Args:
            trigger_id (str): Die Trigger-ID
            trigger_name (str): Der Trigger-Name
            event (str): Das auslösende Event (employee_changed, employee_added, employee_removed)
            employer_id (str): Die Arbeitgeber-ID
            employer_name (str): Der Arbeitgeber-Name
            affected_employees (list): Liste der betroffenen Mitarbeiter mit Änderungsdetails
            action_type (str): Der Aktionstyp (email, api)
            action_details (dict): Details zur ausgeführten Aktion
            status (str): Der Status (success, error, skipped)
            error_message (str): Fehlermeldung bei Status 'error'
            executed_by (str): Der ausführende Benutzer
            retry_of (str): ID einer vorherigen Ausführung (bei Wiederholung)
        
        Returns:
            dict: Der erstellte Log-Eintrag
        """
        data = self._read_data()
        log_entry = {
            'id': str(uuid.uuid4()),
            'trigger_id': trigger_id,
            'trigger_name': trigger_name,
            'event': event,
            'employer_id': employer_id,
            'employer_name': employer_name,
            'executed_at': datetime.now(timezone.utc).isoformat(),
            'executed_by': executed_by,
            'affected_employees': affected_employees,
            'action_type': action_type,
            'action_details': action_details,
            'status': status,
            'error_message': error_message,
            'can_retry': status == 'error',
            'retry_of': retry_of
        }
        data['executions'].insert(0, log_entry)  # Newest first
        
        # Limit to 10000 entries to prevent unbounded growth
        if len(data['executions']) > 10000:
            data['executions'] = data['executions'][:10000]
        
        self._write_data(data)
        return log_entry
    
    def get_executions(self, employer_id=None, trigger_id=None, status=None,
                       from_date=None, to_date=None, limit=100, offset=0):
        """
        Ruft Ausführungen mit optionalen Filtern ab.
        
        Args:
            employer_id (str): Filter nach Arbeitgeber
            trigger_id (str): Filter nach Trigger
            status (str): Filter nach Status (success, error, skipped)
            from_date (str): Filter ab Datum (ISO-Format)
            to_date (str): Filter bis Datum (ISO-Format)
            limit (int): Maximale Anzahl der Ergebnisse
            offset (int): Offset für Pagination
        
        Returns:
            dict: Dictionary mit 'executions', 'total', 'limit', 'offset'
        """
        data = self._read_data()
        executions = data.get('executions', [])
        
        # Apply filters
        if employer_id:
            executions = [e for e in executions if e.get('employer_id') == employer_id]
        if trigger_id:
            executions = [e for e in executions if e.get('trigger_id') == trigger_id]
        if status:
            executions = [e for e in executions if e.get('status') == status]
        if from_date:
            executions = [e for e in executions if e.get('executed_at', '') >= from_date]
        if to_date:
            executions = [e for e in executions if e.get('executed_at', '') <= to_date]
        
        total = len(executions)
        executions = executions[offset:offset + limit]
        
        return {
            'executions': executions,
            'total': total,
            'limit': limit,
            'offset': offset
        }
    
    def get_execution_by_id(self, execution_id):
        """
        Ruft eine einzelne Ausführung anhand ihrer ID ab.
        
        Args:
            execution_id (str): Die Ausführungs-ID
        
        Returns:
            dict | None: Der Log-Eintrag oder None
        """
        data = self._read_data()
        for execution in data.get('executions', []):
            if execution.get('id') == execution_id:
                return execution
        return None
    
    def mark_as_retried(self, execution_id):
        """
        Markiert eine Ausführung als wiederholt (can_retry = False).
        
        Args:
            execution_id (str): Die Ausführungs-ID
        """
        data = self._read_data()
        for execution in data.get('executions', []):
            if execution.get('id') == execution_id:
                execution['can_retry'] = False
                self._write_data(data)
                return


# ==============================================================================
# --- TRIGGER SYSTEM: Engine & Actions ---
# ==============================================================================
# TriggerEngine wertet Trigger aus und führt Aktionen aus.
# Actions sind modulare Handler für verschiedene Aktionstypen (E-Mail, API).

class TriggerEngine:
    """
    Engine zur Auswertung von Triggern und Ausführung von Aktionen.
    
    Die TriggerEngine wird beim Delta-Export aufgerufen und prüft:
    1. Welche Trigger für den Arbeitgeber aktiv sind
    2. Welche Events eingetreten sind (employee_changed, employee_added, employee_removed)
    3. Ob die Bedingungen für jeden Trigger erfüllt sind
    4. Führt die konfigurierten Aktionen aus (E-Mail, API)
    5. Protokolliert alle Ausführungen
    
    Erweiterbar durch neue Aktionstypen via ActionRegistry.
    """
    
    def __init__(self, trigger_store=None, log_store=None):
        """
        Initialisiert die TriggerEngine.
        
        Args:
            trigger_store: Optional: TriggerStore-Instanz
            log_store: Optional: TriggerLogStore-Instanz
        """
        self._trigger_store = trigger_store
        self._log_store = log_store
    
    @property
    def trigger_store(self):
        if self._trigger_store is None:
            self._trigger_store = TriggerStore()
        return self._trigger_store
    
    @property
    def log_store(self):
        if self._log_store is None:
            self._log_store = TriggerLogStore()
        return self._log_store
    
    def evaluate_and_execute(self, employer_cfg: dict, diff: dict, current_data: dict, executed_by: str = 'system'):
        """
        Wertet alle aktiven Trigger aus und führt passende Aktionen aus.
        
        Args:
            employer_cfg (dict): Arbeitgeber-Konfiguration
            diff (dict): Das Diff-Objekt aus dem Delta-Export mit 'added', 'removed', 'changed'
            current_data (dict): Die aktuellen Mitarbeiter-Daten (pid -> data)
            executed_by (str): Der ausführende Benutzer
        
        Returns:
            list: Liste der Ausführungsergebnisse
        """
        employer_id = employer_cfg.get('id')
        employer_name = employer_cfg.get('name', '')
        results = []
        
        # Hole aktive Trigger für diesen Arbeitgeber
        triggers = self.trigger_store.get_triggers_for_employer(employer_id)
        
        if not triggers:
            return results
        
        for trigger in triggers:
            try:
                event = trigger.get('event', 'employee_changed')
                
                # Bestimme betroffene Mitarbeiter basierend auf Event
                affected_employees = []
                
                if event == 'employee_added' and diff.get('added'):
                    affected_employees = self._process_added_employees(diff['added'], current_data, trigger)
                
                elif event == 'employee_removed' and diff.get('removed'):
                    affected_employees = self._process_removed_employees(diff['removed'], trigger)
                
                elif event == 'employee_changed' and diff.get('changed'):
                    affected_employees = self._process_changed_employees(diff['changed'], current_data, trigger)
                
                if not affected_employees:
                    continue
                
                # Führe Aktion aus
                action_type = trigger.get('action', {}).get('type', 'email')
                action_config = trigger.get('action', {}).get('config', {})
                
                # Option: Einzelne Aktion pro Mitarbeiter (Standard: True für E-Mail)
                send_individual = action_config.get('send_individual', True)
                
                if send_individual and len(affected_employees) > 1:
                    # Führe für jeden Mitarbeiter eine separate Aktion aus
                    for emp in affected_employees:
                        single_employee_list = [emp]
                        context = self._build_context(
                            employer_cfg, trigger, single_employee_list, current_data
                        )
                        
                        success, action_details, error_message = self._execute_action(
                            action_type, action_config, context
                        )
                        
                        # Protokolliere einzelne Ausführung
                        log_entry = self.log_store.log_execution(
                            trigger_id=trigger['id'],
                            trigger_name=trigger['name'],
                            event=event,
                            employer_id=employer_id,
                            employer_name=employer_name,
                            affected_employees=single_employee_list,
                            action_type=action_type,
                            action_details=action_details,
                            status='success' if success else 'error',
                            error_message=error_message,
                            executed_by=executed_by
                        )
                        
                        self.trigger_store.update_statistics(trigger['id'], success=success)
                        
                        results.append({
                            'trigger_id': trigger['id'],
                            'trigger_name': trigger['name'],
                            'success': success,
                            'affected_count': 1,
                            'employee': f"{emp.get('firstName', '')} {emp.get('lastName', '')}".strip(),
                            'log_id': log_entry['id']
                        })
                else:
                    # Sammel-Aktion für alle Mitarbeiter zusammen
                    context = self._build_context(
                        employer_cfg, trigger, affected_employees, current_data
                    )
                    
                    success, action_details, error_message = self._execute_action(
                        action_type, action_config, context
                    )
                    
                    # Protokolliere Ausführung
                    log_entry = self.log_store.log_execution(
                        trigger_id=trigger['id'],
                        trigger_name=trigger['name'],
                        event=event,
                        employer_id=employer_id,
                        employer_name=employer_name,
                        affected_employees=affected_employees,
                        action_type=action_type,
                        action_details=action_details,
                        status='success' if success else 'error',
                        error_message=error_message,
                        executed_by=executed_by
                    )
                    
                    # Aktualisiere Trigger-Statistiken
                    self.trigger_store.update_statistics(trigger['id'], success=success)
                    
                    results.append({
                        'trigger_id': trigger['id'],
                        'trigger_name': trigger['name'],
                        'success': success,
                        'affected_count': len(affected_employees),
                        'log_id': log_entry['id']
                    })
                
            except Exception as e:
                # Fehler bei Trigger-Auswertung protokollieren
                custom_log("TRIGGER", f"Fehler bei Trigger '{trigger.get('name')}': {str(e)}", "red")
                
                self.log_store.log_execution(
                    trigger_id=trigger.get('id'),
                    trigger_name=trigger.get('name'),
                    event=trigger.get('event', 'unknown'),
                    employer_id=employer_id,
                    employer_name=employer_name,
                    affected_employees=[],
                    action_type=trigger.get('action', {}).get('type', 'unknown'),
                    action_details={},
                    status='error',
                    error_message=str(e),
                    executed_by=executed_by
                )
        
        return results
    
    def _process_added_employees(self, added_list, current_data, trigger):
        """Verarbeitet hinzugefügte Mitarbeiter."""
        affected = []
        for emp in added_list:
            pid = emp.get('pid')
            emp_data = current_data.get(pid, {}).get('core', {})
            affected.append({
                'personId': pid,
                'firstName': emp_data.get('Vorname', emp.get('name', '').split()[0] if emp.get('name') else ''),
                'lastName': emp_data.get('Name', emp.get('name', '').split()[-1] if emp.get('name') else ''),
                'data': emp_data,
                'changes': []
            })
        return affected
    
    def _process_removed_employees(self, removed_list, trigger):
        """Verarbeitet entfernte Mitarbeiter."""
        affected = []
        for emp in removed_list:
            name_parts = emp.get('name', '').split()
            affected.append({
                'personId': emp.get('pid'),
                'firstName': name_parts[0] if name_parts else '',
                'lastName': name_parts[-1] if len(name_parts) > 1 else '',
                'data': {},
                'changes': []
            })
        return affected
    
    def _process_changed_employees(self, changed_list, current_data, trigger):
        """Verarbeitet geänderte Mitarbeiter und prüft Bedingungen."""
        conditions = trigger.get('conditions', [])
        condition_logic = trigger.get('condition_logic', 'AND')
        affected = []
        
        for emp in changed_list:
            pid = emp.get('pid')
            changes = emp.get('changes', [])
            
            if not changes:
                continue
            
            # Prüfe ob Bedingungen erfüllt sind
            if conditions:
                matches = []
                for condition in conditions:
                    match = self._check_condition(condition, changes)
                    matches.append(match)
                
                if condition_logic == 'AND' and not all(matches):
                    continue
                elif condition_logic == 'OR' and not any(matches):
                    continue
            
            emp_data = current_data.get(pid, {}).get('core', {})
            affected.append({
                'personId': pid,
                'firstName': emp_data.get('Vorname', ''),
                'lastName': emp_data.get('Name', ''),
                'data': emp_data,
                'changes': changes
            })
        
        return affected
    
    def _check_condition(self, condition, changes):
        """
        Prüft ob eine Bedingung durch die Änderungen erfüllt ist.
        
        Args:
            condition (dict): Die Bedingungsdefinition
            changes (list): Liste der Änderungen [{field, old, new}, ...]
        
        Returns:
            bool: True wenn Bedingung erfüllt
        """
        field = condition.get('field')
        operator = condition.get('operator')
        from_value = condition.get('from_value') or ''
        to_value = condition.get('to_value') or ''
        
        # Finde relevante Änderung für das Feld
        change = None
        for c in changes:
            if c.get('field') == field:
                change = c
                break
        
        if not change and operator != 'is_empty' and operator != 'is_not_empty':
            return False
        
        # Helper: Sichere String-Konvertierung (None -> '')
        def safe_str(val):
            return str(val) if val is not None else ''
        
        if operator == 'changed':
            return change is not None
        
        elif operator == 'changed_to':
            return change is not None and safe_str(change.get('new', '')).lower() == safe_str(to_value).lower()
        
        elif operator == 'changed_from':
            return change is not None and safe_str(change.get('old', '')).lower() == safe_str(from_value).lower()
        
        elif operator == 'changed_from_to':
            return (change is not None and 
                    safe_str(change.get('old', '')).lower() == safe_str(from_value).lower() and
                    safe_str(change.get('new', '')).lower() == safe_str(to_value).lower())
        
        elif operator == 'is_empty':
            return change is not None and not change.get('new')
        
        elif operator == 'is_not_empty':
            return change is not None and change.get('new')
        
        elif operator == 'contains':
            return change is not None and to_value and safe_str(to_value).lower() in safe_str(change.get('new', '')).lower()
        
        return False
    
    def _build_context(self, employer_cfg, trigger, affected_employees, current_data):
        """
        Erstellt den Kontext für Template-Rendering.
        
        Args:
            employer_cfg (dict): Arbeitgeber-Konfiguration
            trigger (dict): Trigger-Konfiguration
            affected_employees (list): Betroffene Mitarbeiter
            current_data (dict): Aktuelle Mitarbeiterdaten
        
        Returns:
            dict: Kontext-Dictionary für Templates
        """
        context = {
            '_employerId': employer_cfg.get('id'),
            '_employerName': employer_cfg.get('name'),
            '_triggerName': trigger.get('name'),
            '_timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            '_employeeCount': len(affected_employees),
            '_employees': affected_employees
        }
        
        # Wenn nur ein Mitarbeiter betroffen, füge seine Felder direkt hinzu
        if len(affected_employees) == 1:
            emp = affected_employees[0]
            context.update(emp.get('data', {}))
            
            # Füge Änderungsinformationen hinzu
            if emp.get('changes'):
                first_change = emp['changes'][0]
                context['_changedField'] = first_change.get('field', '')
                context['_oldValue'] = first_change.get('old', '')
                context['_newValue'] = first_change.get('new', '')
            
            # Alle Änderungen als formatierter String
            all_changes = []
            for c in emp.get('changes', []):
                all_changes.append(f"{c.get('field')}: {c.get('old', '-')} → {c.get('new', '-')}")
            context['_allChanges'] = '\n'.join(all_changes)
        
        return context
    
    def _execute_action(self, action_type, action_config, context):
        """
        Führt eine Aktion aus.
        
        Args:
            action_type (str): Der Aktionstyp ('email' oder 'api')
            action_config (dict): Die Aktionskonfiguration
            context (dict): Der Kontext für Template-Rendering
        
        Returns:
            tuple: (success: bool, details: dict, error_message: str | None)
        """
        try:
            if action_type == 'email':
                return EmailAction().execute(action_config, context, self.trigger_store.get_smtp_config())
            elif action_type == 'api':
                return APIAction().execute(action_config, context)
            else:
                return False, {}, f"Unbekannter Aktionstyp: {action_type}"
        except Exception as e:
            return False, {}, str(e)
    
    def retry_execution(self, execution_id, executed_by='system'):
        """
        Wiederholt eine fehlgeschlagene Ausführung.
        
        Args:
            execution_id (str): Die ID der ursprünglichen Ausführung
            executed_by (str): Der ausführende Benutzer
        
        Returns:
            dict | None: Das neue Log-Entry oder None bei Fehler
        """
        execution = self.log_store.get_execution_by_id(execution_id)
        if not execution:
            return None
        
        if not execution.get('can_retry'):
            return None
        
        trigger = self.trigger_store.get_trigger_by_id(execution['trigger_id'])
        if not trigger:
            return None
        
        # Markiere alte Ausführung als wiederholt
        self.log_store.mark_as_retried(execution_id)
        
        # Erstelle Kontext aus gespeicherten Daten
        context = {
            '_employerId': execution['employer_id'],
            '_employerName': execution['employer_name'],
            '_triggerName': execution['trigger_name'],
            '_timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            '_employeeCount': len(execution.get('affected_employees', [])),
            '_employees': execution.get('affected_employees', [])
        }
        
        # Wenn Mitarbeiter vorhanden, füge deren Daten hinzu
        if execution.get('affected_employees'):
            emp = execution['affected_employees'][0]
            for key, value in emp.get('data', {}).items():
                context[key] = value
            if emp.get('changes'):
                context['_changedField'] = emp['changes'][0].get('field', '')
                context['_oldValue'] = emp['changes'][0].get('old', '')
                context['_newValue'] = emp['changes'][0].get('new', '')
        
        # Führe Aktion erneut aus
        action_type = trigger.get('action', {}).get('type', 'email')
        action_config = trigger.get('action', {}).get('config', {})
        
        success, action_details, error_message = self._execute_action(
            action_type, action_config, context
        )
        
        # Protokolliere neue Ausführung
        log_entry = self.log_store.log_execution(
            trigger_id=execution['trigger_id'],
            trigger_name=execution['trigger_name'],
            event=execution['event'],
            employer_id=execution['employer_id'],
            employer_name=execution['employer_name'],
            affected_employees=execution.get('affected_employees', []),
            action_type=action_type,
            action_details=action_details,
            status='success' if success else 'error',
            error_message=error_message,
            executed_by=executed_by,
            retry_of=execution_id
        )
        
        # Aktualisiere Statistiken
        self.trigger_store.update_statistics(trigger['id'], success=success)
        
        return log_entry


class EmailAction:
    """
    Handler für E-Mail-Aktionen.
    
    Sendet E-Mails mit Mustache-Template-Rendering für dynamische Inhalte.
    Unterstützt mehrere Empfänger und mehrere betroffene Mitarbeiter.
    """
    
    def execute(self, config, context, smtp_config):
        """
        Führt die E-Mail-Aktion aus.
        
        Args:
            config (dict): Aktionskonfiguration (recipients, subject, body)
            context (dict): Kontext für Template-Rendering
            smtp_config (dict): SMTP-Server-Konfiguration
        
        Returns:
            tuple: (success: bool, details: dict, error_message: str | None)
        """
        if not smtp_config or not smtp_config.get('host'):
            return False, {}, "SMTP nicht konfiguriert"
        
        recipients = config.get('recipients', [])
        if not recipients:
            return False, {}, "Keine Empfänger angegeben"
        
        subject = self._render_template(config.get('subject', ''), context)
        body = self._render_template(config.get('body', ''), context)
        
        try:
            import smtplib
            import socket
            from email.message import EmailMessage
            
            msg = EmailMessage()
            
            from_email = smtp_config.get('from_email', smtp_config.get('username', ''))
            msg['From'] = from_email
            msg['To'] = ', '.join(recipients)
            msg['Subject'] = subject
            msg.set_content(body)
            
            # Explizit ASCII-sicheren lokalen Hostnamen setzen (vermeidet Umlaute im Computernamen)
            local_hostname = 'localhost'
            try:
                hostname = socket.gethostname()
                hostname.encode('ascii')
                local_hostname = hostname
            except (UnicodeEncodeError, socket.error):
                local_hostname = 'localhost'
            
            if smtp_config.get('use_tls', True):
                server = smtplib.SMTP(smtp_config['host'], smtp_config.get('port', 587), local_hostname=local_hostname)
                server.starttls()
            else:
                server = smtplib.SMTP(smtp_config['host'], smtp_config.get('port', 25), local_hostname=local_hostname)
            
            if smtp_config.get('username') and smtp_config.get('password'):
                server.login(smtp_config['username'], smtp_config['password'])
            
            server.send_message(msg)
            server.quit()
            
            return True, {
                'recipients': recipients,
                'subject': subject
            }, None
            
        except Exception as e:
            return False, {
                'recipients': recipients,
                'subject': subject
            }, str(e)
    
    def _render_template(self, template, context):
        """
        Rendert ein Template mit Mustache-Syntax.
        
        Unterstützt:
        - {{variable}} - Einfache Variablen
        - {{#_employees}}...{{/_employees}} - Listen-Iteration
        
        Args:
            template (str): Das Template
            context (dict): Der Kontext
        
        Returns:
            str: Das gerenderte Template
        """
        if not template:
            return ''
        
        result = template
        
        # Versuche chevron für Mustache-Rendering zu verwenden
        try:
            import chevron
            return chevron.render(template, context)
        except ImportError:
            pass
        
        # Fallback: Einfaches Variablen-Ersetzen
        for key, value in context.items():
            if not key.startswith('_employees'):
                placeholder = '{{' + key + '}}'
                result = result.replace(placeholder, str(value or ''))
        
        # Einfache Listen-Verarbeitung für _employees
        if '{{#_employees}}' in result and '{{/_employees}}' in result:
            start = result.find('{{#_employees}}')
            end = result.find('{{/_employees}}')
            if start != -1 and end != -1:
                before = result[:start]
                template_part = result[start + 15:end]
                after = result[end + 15:]
                
                employees = context.get('_employees', [])
                rendered_parts = []
                for emp in employees:
                    emp_context = dict(context)
                    emp_context.update(emp.get('data', {}))
                    emp_context['firstName'] = emp.get('firstName', '')
                    emp_context['lastName'] = emp.get('lastName', '')
                    emp_context['personId'] = emp.get('personId', '')
                    if emp.get('changes'):
                        emp_context['_changedField'] = emp['changes'][0].get('field', '')
                        emp_context['_oldValue'] = emp['changes'][0].get('old', '')
                        emp_context['_newValue'] = emp['changes'][0].get('new', '')
                    
                    part = template_part
                    for k, v in emp_context.items():
                        part = part.replace('{{' + k + '}}', str(v or ''))
                    rendered_parts.append(part)
                
                result = before + ''.join(rendered_parts) + after
        
        return result


class APIAction:
    """
    Handler für API-Aktionen.
    
    Führt HTTP-Requests mit dynamischen Feldern aus.
    Unterstützt verschiedene Authentifizierungsmethoden.
    """
    
    def execute(self, config, context):
        """
        Führt die API-Aktion aus.
        
        Args:
            config (dict): Aktionskonfiguration (url, method, headers, auth, body)
            context (dict): Kontext für Template-Rendering
        
        Returns:
            tuple: (success: bool, details: dict, error_message: str | None)
        """
        url = self._render_template(config.get('url', ''), context)
        method = config.get('method', 'POST').upper()
        timeout = config.get('timeout_seconds', 30)
        
        if not url:
            return False, {}, "Keine URL angegeben"
        
        # Baue Headers
        headers = {'Content-Type': 'application/json'}
        for key, value in config.get('headers', {}).items():
            headers[key] = self._render_template(value, context)
        
        # Authentifizierung
        auth = config.get('auth', {})
        auth_type = auth.get('type', 'none')
        
        if auth_type == 'bearer' and auth.get('token'):
            headers['Authorization'] = f"Bearer {auth['token']}"
        elif auth_type == 'api_key' and auth.get('api_key'):
            header_name = auth.get('api_key_header', 'X-API-Key')
            headers[header_name] = auth['api_key']
        
        # Basic Auth
        auth_tuple = None
        if auth_type == 'basic' and auth.get('username'):
            auth_tuple = (auth['username'], auth.get('password', ''))
        
        # Body rendern
        body = None
        if method in ['POST', 'PUT', 'PATCH']:
            body_template = config.get('body', '')
            if body_template:
                body_str = self._render_template(body_template, context)
                try:
                    body = json.loads(body_str)
                except json.JSONDecodeError:
                    body = body_str
        
        try:
            response = requests.request(
                method=method,
                url=url,
                headers=headers,
                json=body if isinstance(body, dict) else None,
                data=body if isinstance(body, str) else None,
                auth=auth_tuple,
                timeout=timeout
            )
            
            success = response.status_code < 400
            
            return success, {
                'url': url,
                'method': method,
                'status_code': response.status_code,
                'response_text': response.text[:500] if response.text else ''
            }, None if success else f"HTTP {response.status_code}: {response.text[:200]}"
            
        except requests.Timeout:
            return False, {'url': url, 'method': method}, f"Timeout nach {timeout} Sekunden"
        except Exception as e:
            return False, {'url': url, 'method': method}, str(e)
    
    def _render_template(self, template, context):
        """Rendert ein Template mit einfacher Variablen-Ersetzung."""
        if not template:
            return ''
        
        result = template
        for key, value in context.items():
            if not isinstance(value, (list, dict)):
                result = result.replace('{{' + key + '}}', str(value or ''))
        
        return result


class BaseProvider(ABC):
    """
    Abstrakte Basisklasse für alle HR-Datenprovider.
    
    Diese Klasse definiert die gemeinsame Schnittstelle für alle HR-Provider
    und stellt sicher, dass alle Provider die erforderlichen Methoden implementieren.
    """
    def __init__(self, access_key, secret_key=None, slug=None, **kwargs):
        """
        Initialisiert einen HR-Provider.
        
        Args:
            access_key (str): Der Zugangsschlüssel für den Provider
            secret_key (str, optional): Der geheime Schlüssel für den Provider
            slug (str, optional): Der Slug für den Provider
            **kwargs: Weitere Provider-spezifische Parameter
        """
        self.access_key, self.secret_key, self.slug = access_key, secret_key, slug
    @abstractmethod
    def list_employees(self, only_active: bool = True) -> tuple[list[dict], list]:
        """
        Ruft eine Liste aller Mitarbeiter ab.
        
        Args:
            only_active (bool): Ob nur aktive Mitarbeiter zurückgegeben werden sollen
        
        Returns:
            tuple[list[dict], list]: Tupel aus (verarbeitete Mitarbeiterdaten, rohe API-Antwort)
        """
        pass
    @abstractmethod
    def get_employee_details(self, employee_id: str, return_history: bool = True) -> tuple[dict, dict | list]:
        """
        Ruft detaillierte Informationen für einen einzelnen Mitarbeiter ab.
        
        Args:
            employee_id (str): Die ID des Mitarbeiters
            return_history (bool): Ob die rohe API-Antwort zurückgegeben werden soll
        
        Returns:
            tuple[dict, dict | list]: Tupel aus (verarbeitete Mitarbeiterdaten, rohe API-Antwort)
        """
        pass
    def __repr__(self):
        """
        Gibt eine String-Darstellung des Providers zurück.
        
        Returns:
            str: String-Darstellung des Providers
        """
        return f"<{self.__class__.__name__}>"

class HRworksProvider(BaseProvider):
    """
    Provider für das Abrufen von Mitarbeiterdaten aus der HRworks API.
    
    Diese Klasse implementiert die HRworks-spezifische Logik für Authentifizierung,
    Datenabruf und Normalisierung von Mitarbeiterdaten.
    """
    API_BASE_URL = "https://api.hrworks.de/v2"
    DEMO_API_BASE_URL = "https://api.demo-hrworks.de/v2"

    def __init__(self, access_key, secret_key, is_demo=False, **kwargs):
        """
        Initialisiert den HRworks-Provider.
        
        Args:
            access_key (str): Der Zugangsschlüssel für HRworks
            secret_key (str): Der geheime Schlüssel für HRworks
            is_demo (bool): Ob die Demo-Umgebung verwendet werden soll
            **kwargs: Weitere Provider-spezifische Parameter
        """
        super().__init__(access_key=access_key, secret_key=secret_key, **kwargs)
        self.base_url = self.DEMO_API_BASE_URL if is_demo else self.API_BASE_URL
        self.bearer_token = None
        self.auth_header = None
        self._persons_cache = None
        self._authenticate()

    def _fmt_date(self, v: str | None) -> str | None:
        """
        Formatiert sicher einen Datumsstring von YYYY-MM-DD zu DD.MM.YYYY.
        
        Args:
            v (str | None): Der zu formatierende Datumsstring
        
        Returns:
            str | None: Der formatierte Datumsstring oder None bei Fehlern
        """
        if not v:
            return None
        s = str(v)
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d").strftime("%d.%m.%Y")
        except Exception:
            return s[:10] if len(s) >= 10 else s

    def _authenticate(self):
        """
        Authentifiziert sich bei der HRworks API, um einen Bearer-Token zu erhalten.
        
        Raises:
            ConnectionError: Bei Authentifizierungsfehlern
        
        Returns:
            None
        """
        try:
            url = f"{self.base_url}/authentication"
            payload = {"accessKey": self.access_key, "secretAccessKey": self.secret_key}
            response = requests.post(url, json=payload, headers={"Content-Type": "application/json", "Accept": "application/json"}, timeout=20)
            response.raise_for_status()
            token_data = response.json() or {}
            token = token_data.get("token") or token_data.get("accessToken")
            if not token:
                raise ValueError("Kein Token in der HRworks-Antwort gefunden.")
            self.bearer_token = token
            self.auth_header = {"Authorization": f"Bearer {self.bearer_token}", "Accept": "application/json"}
        except Exception as e:
            raise ConnectionError(f"HRworks-Authentifizierung fehlgeschlagen: {e}")

    def _get_all_persons(self, only_active=True):
        """
        Ruft alle Personen-Stammdaten ab, behandelt Paginierung und gibt rohe Antworten zurück.
        
        Args:
            only_active (bool): Ob nur aktive Personen abgerufen werden sollen
        
        Returns:
            tuple[list, list]: Tupel aus (alle Personen, rohe API-Antworten)
        
        Raises:
            ConnectionError: Bei API-Fehlern
        """
        if not self.auth_header:
            raise ConnectionError("Nicht authentifiziert.")

        all_persons = []
        raw_responses = [] # Capture raw responses
        url = f"{self.base_url}/persons/master-data"
        params = {"onlyActive": str(only_active).lower(), "page": 1}

        while url:
            resp = requests.get(url, headers=self.auth_header, params=params, timeout=20)
            try:
                resp.raise_for_status()
            except Exception as e:
                raise ConnectionError(f"Fehler bei GET {url}: {resp.status_code} {resp.text}") from e

            data = resp.json() or {}
            raw_responses.append(data) # Save raw page data
            items = None
            if isinstance(data, dict):
                for k in ("data", "items", "persons", "results"):
                    if k in data and isinstance(data[k], list):
                        items = data[k]
                        break
            if items is None and isinstance(data, list):
                items = data

            if not items:
                break
            all_persons.extend(items)

            next_url = None
            link_header = resp.headers.get("Link", "")
            for part in link_header.split(","):
                p = part.strip()
                if p.startswith("<") and ">;" in p and 'rel="next"' in p:
                    next_url = p[1:p.index(">;")].strip()
                    break
            url = next_url
            params = None
        return all_persons, raw_responses

    def list_employees(self, only_active: bool = True) -> tuple[list[dict], list]:
        """
        Ruft eine Liste aller Mitarbeiter mit ihren vollständigen, normalisierten Details ab.
        
        Args:
            only_active (bool): Ob nur aktive Mitarbeiter zurückgegeben werden sollen
        
        Returns:
            tuple[list[dict], list]: Tupel aus (normalisierte Mitarbeiterdaten, rohe API-Antworten)
        """
        persons, raw_responses = self._get_all_persons(only_active=only_active)

        # Normalize the full details for each employee in the list
        employees = [self._normalize_employee_details(p) for p in persons]

        employees.sort(key=lambda x: ((x.get("lastName") or "").lower(), (x.get("firstName") or "").lower()))
        return employees, raw_responses

    def get_employee_details(self, employee_id: str, return_history: bool = True) -> tuple[dict, dict | list]:
        """
        Ruft detaillierte Informationen für einen einzelnen Mitarbeiter ab.
        
        Args:
            employee_id (str): Die ID des Mitarbeiters
            return_history (bool): Ob die rohe API-Antwort zurückgegeben werden soll
        
        Returns:
            tuple[dict, dict | list]: Tupel aus (normalisierte Mitarbeiterdaten, rohe API-Antwort)
        
        Raises:
            ValueError: Wenn der Mitarbeiter nicht gefunden wird
        """
        url = f"{self.base_url}/persons/master-data/{employee_id}"
        r = requests.get(url, headers=self.auth_header, timeout=20)
        if r.status_code == 200:
            raw = r.json()
            details = self._normalize_employee_details(raw.get("data", raw) if isinstance(raw, dict) else {})
            return (details, raw) if return_history else (details, [])

        # Fallback to searching the full list
        persons, raw_responses = self._get_all_persons(only_active=False)
        for p in persons:
            if any(str(employee_id) == str(v) for v in (p.get("uuid"), p.get("personId"), p.get("personnelNumber"), p.get("id"))):
                details = self._normalize_employee_details(p)
                return (details, raw_responses) if return_history else (details, [])

        raise ValueError(f"Mitarbeiter nicht gefunden: {employee_id}")

    def _normalize_employee_details(self, raw: dict) -> dict:
        """
        Normalisiert rohe API-Daten in ein konsistentes internes Format.
        
        Args:
            raw (dict): Die rohen API-Daten eines Mitarbeiters
        
        Returns:
            dict: Die normalisierten Mitarbeiterdaten
        """
        pid = raw.get("personId") or raw.get("uuid") or raw.get("personnelNumber") or raw.get("id")
        first, last = raw.get("firstName"), raw.get("lastName")
        status = raw.get("status") or ("active" if raw.get("isActive") else "inactive")
        is_active = bool(raw.get("isActive"))
        if not is_active and isinstance(status, str):
            is_active = (status.strip().lower() == "active")

        join = raw.get("hireDate") or raw.get("joinDate") or raw.get("startDate")
        leave = raw.get("terminationDate") or raw.get("leaveDate") or raw.get("endDate") or raw.get("contractEndDate")

        # Organization Unit für Abteilung extrahieren
        org_unit = raw.get("organizationUnit") or {}
        department_name = org_unit.get("name") if isinstance(org_unit, dict) else None
        
        detail = {
            "personId": str(pid) if pid is not None else None,
            "personnelNumber": raw.get("personnelNumber"),
            "firstName": first,
            "lastName": last,
            "birthday": self._fmt_date(raw.get("birthday")) or "",
            "email": raw.get("email") or raw.get("workEmail") or raw.get("businessEmail"),
            "position": raw.get("position") or raw.get("jobTitle"),
            "department": department_name,  # Abteilung direkt hinzufügen
            "gender": raw.get("gender"),
            "employmentType": raw.get("employmentType") or raw.get("typeOfEmployment"),
            "status": status,
            "isActive": is_active,
            "joinDate": self._fmt_date(join) or "",
            "leaveDate": self._fmt_date(leave) or "",
            "organizationUnit": org_unit,
            "costCenter": (raw.get("costCenter") or {}),
            "address": (raw.get("address") or {}),
            "bankAccount": (raw.get("bankAccount") or {}),
            "salary": (raw.get("salary") or {}),
            "superior": (raw.get("superior") or {}),
        }

        details_groups = {
            "Persönliche Informationen": [
                {"label": "Vorname", "value": first},
                {"label": "Nachname", "value": last},
                {"label": "Geburtsdatum", "value": detail["birthday"]},
                {"label": "Geschlecht", "value": detail["gender"]},
            ],
            "Kontaktdaten": [
                {"label": "E-Mail", "value": detail["email"]},
                {"label": "Straße", "value": detail["address"].get("street")},
                {"label": "Hausnummer", "value": detail["address"].get("streetNumber")},
                {"label": "PLZ", "value": detail["address"].get("zipCode")},
                {"label": "Stadt", "value": detail["address"].get("city")},
            ],
            "Anstellung": [
                {"label": "Position", "value": detail["position"]},
                {"label": "Beschäftigungsart", "value": detail["employmentType"]},
                {"label": "Eintrittsdatum", "value": detail["joinDate"]},
                {"label": "Kündigungsdatum", "value": detail["leaveDate"]},
            ],
            "Organisation": [
                {"label": "Abteilung", "value": detail["organizationUnit"].get("name")},
                {"label": "Kostenstelle", "value": detail["costCenter"].get("name")},
            ]
        }

        final_details = {}
        for group, items in details_groups.items():
            filtered_items = [item for item in items if item.get("value")]
            if filtered_items:
                final_details[group] = filtered_items

        detail["details"] = final_details
        return detail

class SageHrProvider(BaseProvider):
    """
    Mock-Provider für SageHR, der statische Daten zurückgibt.
    
    Diese Klasse dient als Platzhalter für die SageHR-Integration
    und gibt vordefinierte Testdaten zurück.
    """
    def __init__(self, access_key, slug, **kwargs):
        """
        Initialisiert den SageHR-Mock-Provider.
        
        Args:
            access_key (str): Der Zugangsschlüssel für SageHR
            slug (str): Der Slug für SageHR
            **kwargs: Weitere Provider-spezifische Parameter
        """
        super().__init__(access_key=access_key, slug=slug, **kwargs)
    def list_employees(self, only_active: bool = True) -> tuple[list[dict], list]:
        """
        Ruft eine Liste von Mock-Mitarbeitern ab.
        
        Args:
            only_active (bool): Ob nur aktive Mitarbeiter zurückgegeben werden sollen
        
        Returns:
            tuple[list[dict], list]: Tupel aus (Mock-Mitarbeiterdaten, Mock-API-Antwort)
        """
        mock_data = [{'id': 's-301', 'firstName': 'Ben', 'lastName': 'Berger', 'isActive': True, 'position': 'Developer', 'department': 'IT'}]
        data_to_return = [e for e in mock_data if e['isActive']] if only_active else mock_data
        return data_to_return, [mock_data] # Return mock raw data
    def get_employee_details(self, employee_id: str, return_history: bool = True) -> tuple[dict, dict | list]:
        """
        Ruft Mock-Details für einen Mitarbeiter ab.
        
        Args:
            employee_id (str): Die ID des Mitarbeiters (wird ignoriert)
            return_history (bool): Ob die Mock-API-Antwort zurückgegeben werden soll
        
        Returns:
            tuple[dict, dict | list]: Tupel aus (Mock-Mitarbeiterdaten, Mock-API-Antwort)
        """
        mock_details = {"isActive": True, "firstName": "Ben (Mock)", "lastName": "Berger", "details": {"Info": [{"label": "Provider", "value": "Sage HR Mock"}]}}
        return (mock_details, mock_details) if return_history else (mock_details, [])

class PersonioProvider(BaseProvider):
    """
    Provider für das Abrufen von Mitarbeiterdaten aus der Personio API.
    
    Diese Klasse implementiert die Personio-spezifische Logik für Authentifizierung,
    Datenabruf und Normalisierung von Mitarbeiterdaten.
    """
    PERSONIO_API_BASE_URL = "https://api.personio.de/v1"
    KEY_TO_LABEL_MAP = {"id":"ID","first_name":"Vorname","last_name":"Nachname","preferred_name":"Bevorzugter Name","email":"E-Mail","gender":"Geschlecht","status":"Status","position":"Position","supervisor":"Vorgesetzter","employment_type":"Beschäftigungsart","weekly_working_hours":"Wochenstunden","hire_date":"Eintrittsdatum","contract_end_date":"Vertragsende","termination_date":"Kündigungsdatum","termination_type":"Art der Kündigung","termination_reason":"Kündigungsgrund","probation_period_end":"Probezeitende","created_at":"Erstellt am","last_modified_at":"Letzte Änderung","subcompany":"Untergesellschaft","office":"Arbeitsplatz","department":"Abteilung","cost_centers":"Kostenstelle","holiday_calendar":"Feiertagskalender","absence_entitlement":"Urlaubsanspruch","work_schedule":"Arbeitszeitmodell","fix_salary":"Festgehalt","fix_salary_interval":"Gehaltsintervall","hourly_salary":"Stundenlohn","last_working_day":"Letzter Arbeitstag","team":"Team","dynamic_16291384":"Steueridentifikationsnummer","dynamic_16291391":"IBAN","dynamic_16291393":"Notfallkontakt (Name)","dynamic_16291410":"Straße","dynamic_16291415":"Persönliche E-Mail","dynamic_16291387":"Sozialversicherungsnummer","dynamic_16291392":"BIC","dynamic_16291394":"Notfallkontakt (Handy)","dynamic_16291411":"Hausnummer","dynamic_16291383":"Personalnummer","dynamic_16291385":"Lohnsteuerklasse","dynamic_16291390":"Abweichender Kontoinhaber","dynamic_16291416":"Mobilnummer","dynamic_16291381":"Geburtsdatum","dynamic_16291382":"LinkedIn","dynamic_16291395":"Notfallkontakt (Beziehung)","dynamic_16291396":"Familienstand","dynamic_16291386":"Kirchensteuer","dynamic_16291388":"Art der Krankenversicherung","dynamic_16291389":"Krankenversicherung","dynamic_16291401":"Nationalität","dynamic_16291408":"Postleitzahl","dynamic_16291397":"Haupt-/Nebenarbeitgeber","dynamic_16291399":"Studienbescheinigung gültig bis","dynamic_16291409":"Ort","dynamic_16291398":"Kinderfreibetrag","dynamic_16291400":"Abrechnungsart","dynamic_16291402":"Kündigungsfrist","dynamic_16291404":"Höchster Schulabschluss","dynamic_16291403":"Beschäftigungsart","dynamic_16291405":"Höchster Ausbildungsabschluss","dynamic_16291413":"Projekt Manager","dynamic_16291414":"Mentor"}
    KEY_TO_GROUP_MAP = {"Persönliche Informationen":["first_name","last_name","preferred_name","dynamic_16291381","gender","dynamic_16291396","dynamic_16291401","dynamic_16291382"],"Kontaktdaten":["email","dynamic_16291415","dynamic_16291416"],"Adresse":["dynamic_16291410","dynamic_16291411","dynamic_16291408","dynamic_16291409"],"Anstellung":["status","position","employment_type","dynamic_16291403","hire_date","contract_end_date","termination_date","probation_period_end","last_working_day","termination_type","termination_reason","dynamic_16291402"],"Organisation":["subcompany","office","department","team","supervisor","cost_centers","dynamic_16291413","dynamic_16291414"],"Gehalt & Finanzen":["fix_salary","fix_salary_interval","hourly_salary","dynamic_16291400","dynamic_16291397"],"Bankverbindung":["dynamic_16291391","dynamic_16291392","dynamic_16291390"],"Steuer & Sozialversicherung":["dynamic_16291383","dynamic_16291384","dynamic_16291387","dynamic_16291385","dynamic_16291398","dynamic_16291386","dynamic_16291388","dynamic_16291389"],"Systeminformationen":["id","created_at","last_modified_at"],"Sonstiges":["holiday_calendar","absence_entitlement","work_schedule","dynamic_16291399","dynamic_16291404","dynamic_16291405"],"Notfallkontakt":["dynamic_16291393","dynamic_16291394","dynamic_16291395"]}

    def __init__(self, access_key, secret_key, **kwargs):
        """
        Initialisiert den Personio-Provider.
        
        Args:
            access_key (str): Der Client-ID für Personio
            secret_key (str): Der Client-Secret für Personio
            **kwargs: Weitere Provider-spezifische Parameter
        """
        super().__init__(access_key=access_key, secret_key=secret_key, **kwargs)
        self.bearer_token, self.auth_header = None, None
        self._authenticate()
    def _authenticate(self):
        """
        Authentifiziert sich bei der Personio API, um einen Bearer-Token zu erhalten.
        
        Raises:
            ConnectionError: Bei Authentifizierungsfehlern
        
        Returns:
            None
        """
        try:
            r=requests.post(f"{self.PERSONIO_API_BASE_URL}/auth",json={"client_id":self.access_key,"client_secret":self.secret_key},timeout=10)
            r.raise_for_status()
            self.bearer_token=r.json()['data']['token']
            self.auth_header={'Authorization':f'Bearer {self.bearer_token}'}
        except Exception as e: raise ConnectionError(f"Personio-Authentifizierung fehlgeschlagen: {e}")

    def _fetch_profile_picture_as_data_uri(self, url: str) -> str | None:
        """
        Lädt ein Profilbild von einer URL und konvertiert es zu einem Data-URI.
        
        Args:
            url (str): Die URL des Profilbilds
        
        Returns:
            str | None: Der Data-URI des Bildes oder None bei Fehlern
        """
        if not url or not isinstance(url,str): return None
        try:
            r=requests.get(url,headers=self.auth_header,timeout=10)
            r.raise_for_status()
            if 'image' not in r.headers.get('Content-Type',''): return None
            return f"data:{r.headers['Content-Type']};base64,{base64.b64encode(r.content).decode('utf-8')}"
        except: return None

    def _normalize_employee_details(self, attributes: dict) -> dict:
        """
        Normalisiert rohe Personio-API-Daten in ein konsistentes internes Format.
        
        Args:
            attributes (dict): Die rohen Attribut-Daten von Personio
        
        Returns:
            dict: Die normalisierten Mitarbeiterdaten
        """
        grouped_details = {group: [] for group in self.KEY_TO_GROUP_MAP.keys()}
        grouped_details["Andere"] = []

        for key in sorted(attributes.keys()):
            attr_obj = attributes[key]
            if not isinstance(attr_obj, dict): continue
            value = attr_obj.get('value')
            if value is None or value == '' or value == []: continue

            label = self.KEY_TO_LABEL_MAP.get(key, attr_obj.get('label', key))

            if attr_obj.get('type') == 'date': value = _format_date_for_display(value)
            elif key == 'supervisor': value = f"{value.get('attributes', {}).get('first_name', {}).get('value', '')} {value.get('attributes', {}).get('last_name', {}).get('value', '')}".strip()
            elif key == 'absence_entitlement': value = ", ".join([f"{v.get('name', '')}: {v.get('entitlement', 0)}" for v in value])
            elif isinstance(value, dict) and 'attributes' in value: value = value['attributes'].get('name', str(value))
            elif key in ['dynamic_16291413', 'dynamic_16291414']:
                try:
                    value = ", ".join([item['label'] for item in json.loads(value)])
                except:
                    pass

            found_group = False
            for group_name, keys_in_group in self.KEY_TO_GROUP_MAP.items():
                if key in keys_in_group:
                    grouped_details[group_name].append({'label': label, 'value': value})
                    found_group = True
                    break

            if not found_group:
                grouped_details["Andere"].append({'label': label, 'value': value})

        final_details = {k: v for k, v in grouped_details.items() if v}
        get_value = lambda k: attributes.get(k, {}).get('value')
        
        # Department kann ein String oder ein verschachteltes Objekt sein
        department_val = get_value("department")
        if isinstance(department_val, dict):
            department_val = department_val.get('attributes', {}).get('name', department_val.get('name'))

        return {
            "id": get_value("id"),
            "isActive": str(get_value("status")).lower() == 'active',
            "firstName": get_value("first_name"),
            "lastName": get_value("last_name"),
            "position": get_value("position"),
            "department": department_val,  # Abteilung hinzugefügt
            "profilePictureUrl": self._fetch_profile_picture_as_data_uri(get_value("profile_picture")),
            "details": final_details
        }

    def list_employees(self, only_active: bool = True) -> tuple[list[dict], list]:
        """
        Ruft eine Liste aller Mitarbeiter mit ihren vollständigen, normalisierten Details ab.
        
        Args:
            only_active (bool): Ob nur aktive Mitarbeiter zurückgegeben werden sollen
        
        Returns:
            tuple[list[dict], list]: Tupel aus (normalisierte Mitarbeiterdaten, rohe API-Antwort)
        
        Raises:
            ConnectionError: Bei Authentifizierungsfehlern oder API-Fehlern
        """
        if not self.auth_header: raise ConnectionError("Nicht authentifiziert.")
        try:
            r=requests.get(f"{self.PERSONIO_API_BASE_URL}/company/employees",headers=self.auth_header,timeout=20)
            r.raise_for_status()
            raw_response = r.json()
            employees=[]
            for emp_data in raw_response.get('data',[]):
                attrs = emp_data.get('attributes',{})

                # Check active status before doing the expensive normalization
                status = attrs.get('status', {}).get('value', 'inactive')
                is_active = str(status).lower() == 'active'
                if only_active and not is_active:
                    continue

                # Normalize the full details for each employee
                normalized_details = self._normalize_employee_details(attrs)
                employees.append(normalized_details)

            return employees, [raw_response]
        except Exception as e: raise ConnectionError(f"Fehler bei Mitarbeiterliste: {e}")

    def get_employee_details(self, employee_id: str, return_history: bool = True) -> tuple[dict, dict | list]:
        """
        Ruft detaillierte Informationen für einen einzelnen Mitarbeiter ab.
        
        Args:
            employee_id (str): Die ID des Mitarbeiters
            return_history (bool): Ob die rohe API-Antwort zurückgegeben werden soll
        
        Returns:
            tuple[dict, dict | list]: Tupel aus (normalisierte Mitarbeiterdaten, rohe API-Antwort)
        
        Raises:
            ConnectionError: Bei Authentifizierungsfehlern
            ValueError: Bei Verarbeitungsfehlern
        """
        if not self.auth_header: raise ConnectionError("Nicht authentifiziert.")
        try:
            r=requests.get(f"{self.PERSONIO_API_BASE_URL}/company/employees/{employee_id}",headers=self.auth_header,timeout=10)
            r.raise_for_status()
            raw_response = r.json()
            details = self._normalize_employee_details(raw_response.get('data',{}).get('attributes',{}))
            return (details, raw_response) if return_history else (details, [])
        except Exception as e: raise ValueError(f"Fehler bei Verarbeitung der Mitarbeiterdetails: {e}")

class ProviderFactory:
    """
    Factory-Klasse zur Erstellung von Provider-Instanzen basierend auf der Konfiguration.
    
    Diese Klasse verwaltet die verfügbaren Provider-Typen und erstellt
    die entsprechenden Instanzen basierend auf dem provider_key.
    """
    @staticmethod
    def get_provider(employer_config: dict):
        """
        Erstellt eine Provider-Instanz basierend auf der Arbeitgeber-Konfiguration.
        
        Args:
            employer_config (dict): Die Konfiguration des Arbeitgebers mit provider_key
        
        Returns:
            BaseProvider: Eine Instanz des entsprechenden Providers
        
        Raises:
            ValueError: Wenn der provider_key unbekannt ist
        """
        PROVIDER_MAP = {"hrworks": HRworksProvider, "personio": PersonioProvider, "sagehr": SageHrProvider}
        provider_key = employer_config.get("provider_key")
        provider_class = PROVIDER_MAP.get(provider_key)
        if not provider_class: raise ValueError(f"Unbekannter Provider: {provider_key}")
        return provider_class(**employer_config)

# ==============================================================================
# --- SECTION 3: LOGIC FUNCTIONS (EXPORTS, STATISTICS) ---
# ==============================================================================

def _map_to_scs_schema(e: dict, employer_name: str, provider_key: str) -> dict:
    """
    Mappt einen detaillierten Mitarbeiterdatensatz auf das SCS-Tabellenkalkulationsschema.
    
    Args:
        e (dict): Die Mitarbeiterdaten
        employer_name (str): Der Name des Arbeitgebers
        provider_key (str): Der Schlüssel des Providers für provider-spezifische Logik
    
    Returns:
        dict: Die gemappten Daten im SCS-Schema
    """
    details = e.get("details", {}) if isinstance(e.get("details"), dict) else None
    graw = _getv(e, details, "geschlecht", "gender")
    gl = graw.lower()
    gmap_m = {"mann","männlich","m","male","mr","mr.","herr","1"}
    gmap_f = {"frau","weiblich","w","female","f","mrs","mrs.","2"}
    gender = "Mann" if gl in gmap_m else ("Frau" if gl in gmap_f else graw)

    # --- Provider-specific cleanups and logic ---
    # Use more specific labels for personio, with fallbacks for hrworks
    strasse = _getv(e, details, "straße", "address.street")
    hausnummer = _getv(e, details, "hausnummer", "address.streetNumber")
    # Personio uses 'Postleitzahl', HRworks uses 'PLZ' as label
    plz = _getv(e, details, "postleitzahl") or _getv(e, details, "plz", "address.zipCode")
    # Personio uses 'Ort', HRworks uses 'Stadt' as label
    ort = _getv(e, details, "ort") or _getv(e, details, "stadt", "address.city")

    # HRworks returns "n/a" for empty address fields
    if provider_key == 'hrworks':
        if strasse == 'n/a': strasse = ''
        if hausnummer == 'n/a': hausnummer = ''
        if plz == 'n/a': plz = ''
        if ort == 'n/a': ort = ''

    # Personio Email Logic: Prefer personal, fallback to work
    email = _getv(e, details, "email", "email") # Default
    if provider_key == 'personio':
        personal_email = _getv(e, details, "persönliche e-mail")
        work_email = _getv(e, details, "e-mail")
        email = personal_email or work_email

    # Status: Aktiv/Inaktiv ermitteln
    is_active = e.get("isActive")
    if is_active is None:
        status_raw = _getv(e, details, "status", "status")
        if status_raw:
            is_active = str(status_raw).lower() in ("active", "aktiv", "true", "1")
        else:
            is_active = True  # Default: aktiv
    status = "Aktiv" if is_active else "Inaktiv"

    row = {
        "Name": _getv(e, details, "nachname", "lastName"),
        "Vorname": _getv(e, details, "vorname", "firstName"),
        "Geschlecht": gender,
        "Titel": _getv(e, details, "titel"),
        "Geburtsdatum": _getv(e, details, "geburtsdatum", "birthday"),
        "Strasse": strasse,
        "Hausnummer": hausnummer,
        "PLZ": plz,
        "Ort": ort,
        "Land": _getv(e, details, "land", "address.country", default="D"),
        "Email": email,
        "Telefon": _getv(e, details, "mobilnummer"),
        "Personalnummer": _getv(e, details, "personalnummer", "personnelNumber"),
        "Position": _getv(e, details, "position", "position"),
        "Firmeneintritt": _getv(e, details, "eintrittsdatum", "joinDate"),
        "Bruttogehalt": str(_getv(e, details, "festgehalt", "salary.fixed") or ""),
        "Steuerklasse": _getv(e, details, "lohnsteuerklasse") or _getv(e, details, "steuerklasse", "taxClass"),
        "Religion": _getv(e, details, "kirchensteuer", "religion"),
        "Abteilung": _get_from_path(e, "organizationUnit.name") or _getv(e, details, "abteilung"),
        "Arbeitsplatz": _getv(e, details, "arbeitsplatz", "office"),
        "Arbeitgeber": employer_name or "",
        "Status": status,
    }
    for h in SCS_HEADERS: row.setdefault(h, "")
    return row

def generate_standard_export(e_details, n, k, d):
    """
    Generiert einen Standard-Excel-Export mit allen verfügbaren Daten.
    
    Args:
        e_details (list): Liste der Mitarbeiterdetails
        n (str): Name des Arbeitgebers
        k (str): Provider-Schlüssel
        d (str): Export-Verzeichnis
    
    Returns:
        str: Der Pfad zur generierten Excel-Datei
    """
    rows, headers = [], set(['employer', 'provider'])
    for emp in e_details:
        r = {'employer': n, 'provider': k}
        for g, items in emp.get('details', {}).items():
            for i in items:
                header = f"{g}::{i['label']}"
                headers.add(header)
                r[header] = i['value']
        rows.append(r)

    hl = sorted(list(headers))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Standard-Export"
    ws.append(hl)
    for rd in rows: ws.append([rd.get(h) for h in hl])

    t = datetime.now().strftime("%Y%m%d-%H%M%S")
    fn = f"standard_{_get_safe_employer_name(n)}_{k}_{t}.xlsx"
    fp = os.path.join(d, fn)
    wb.save(fp)
    return fp

def _json_hash(rec: dict) -> str:
    """
    Erstellt einen stabilen Hash eines Dictionarys.
    
    Args:
        rec (dict): Das zu hashende Dictionary
    
    Returns:
        str: Der SHA256-Hash des Dictionarys
    """
    payload = json.dumps(rec, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8", errors="ignore")).hexdigest()

def _flatten_record(obj, prefix="", out=None):
    """
    Flacht ein verschachteltes Dictionary auf eine einzige Ebene ab.
    
    Args:
        obj: Das zu flachende Objekt
        prefix (str): Das Präfix für die Schlüssel
        out (dict, optional): Das Ausgabe-Dictionary
    
    Returns:
        dict: Das geflachtete Dictionary
    """
    if out is None: out = {}
    if isinstance(obj, dict):
        for k, v in obj.items():
            _flatten_record(v, f"{prefix}.{k}" if prefix else str(k), out)
    elif isinstance(obj, list):
        out[prefix or "value"] = json.dumps(obj, ensure_ascii=False, separators=(",", ":"))
    else:
        out[prefix or "value"] = obj
    return out

def _person_key(detail: dict) -> str | None:
    """
    Bestimmt die stabile eindeutige ID für einen Mitarbeiter.
    
    Args:
        detail (dict): Die Mitarbeiterdetails
    
    Returns:
        str | None: Die Mitarbeiter-ID oder None, wenn nicht gefunden
    """
    pid = detail.get("personId") or detail.get("uuid") or detail.get("personnelNumber") or detail.get("id")
    return str(pid) if pid is not None else None

def generate_delta_scs_export(current_details: list[dict], employer_cfg: dict, snapshots_dir: str, exports_dir: str, always_write: bool = False):
    """
    Generiert einen Delta-Export, speichert Snapshots und gibt Änderungen zurück.
    
    Args:
        current_details (list[dict]): Die aktuellen Mitarbeiterdetails
        employer_cfg (dict): Die Arbeitgeber-Konfiguration
        snapshots_dir (str): Das Verzeichnis für Snapshots
        exports_dir (str): Das Verzeichnis für Exporte
        always_write (bool): Ob immer geschrieben werden soll, auch ohne Änderungen
    
    Returns:
        dict: Dictionary mit filepath und diff-Daten
    """
    employer_name = employer_cfg.get('name', '')
    provider_key = employer_cfg.get('provider_key', '')
    safe_emp = _get_safe_employer_name(employer_name)
    latest_path = os.path.join(snapshots_dir, f"{safe_emp}-{provider_key}-latest.json")
    dated_path = os.path.join(snapshots_dir, datetime.now().strftime(f"{safe_emp}-{provider_key}-%Y%m%d-%H%M%S.json"))

    prev = {}
    if os.path.exists(latest_path):
        try: prev = json.load(open(latest_path, "r", encoding="utf-8")) or {}
        except Exception: pass

    current = {}
    for detail in current_details:
        pid = _person_key(detail)
        if not pid: continue

        # Extract key dates using the robust _getv helper for consistency
        details_dict = detail.get("details", {})
        join_date_str = _getv(detail, details_dict, "Eintrittsdatum", "joinDate", "hire_date")
        leave_date_str = _getv(detail, details_dict, "Kündigungsdatum", "leaveDate", "termination_date", "contract_end_date")
        key_dates = {"join": join_date_str, "leave": leave_date_str}

        flat = _flatten_record(detail)
        h = _json_hash(flat)
        core = _map_to_scs_schema(detail, employer_name, provider_key)
        current[pid] = {"hash": h, "flat": flat, "core": core, "dates": key_dates}

    # --- Compare snapshots ---
    prev_pids = set(prev.keys())
    current_pids = set(current.keys())
    added_pids = current_pids - prev_pids
    removed_pids = prev_pids - current_pids
    common_pids = current_pids & prev_pids
    changed_pids = {pid for pid in common_pids if prev[pid].get('hash') != current[pid].get('hash')}

    # --- Create human-readable diff ---
    diff = {
        'added': [{'pid': pid, 'name': f"{current[pid]['core'].get('Vorname', '')} {current[pid]['core'].get('Name', '')}".strip(), 'geburtstag': current[pid]['core'].get('Geburtsdatum', '')} for pid in added_pids],
        'removed': [{'pid': pid, 'name': f"{prev[pid]['core'].get('Vorname', '')} {prev[pid]['core'].get('Name', '')}".strip(), 'geburtstag': prev[pid]['core'].get('Geburtsdatum', '')} for pid in removed_pids],
        'changed': []
    }
    comparison = _compare_snapshots(prev, current)
    diff['changed'] = comparison.get('changed', [])


    # --- Write new snapshots regardless of changes, to keep history complete ---
    try:
        with open(dated_path, "w", encoding="utf-8") as f: json.dump(current, f, ensure_ascii=False, indent=2)
        with open(latest_path, "w", encoding="utf-8") as f: json.dump(current, f, ensure_ascii=False, indent=2)
    except Exception:
        # Log this error? For now, we continue as it's not critical for the export itself.
        pass

    # --- TRIGGER SYSTEM: Evaluate and execute triggers ---
    trigger_results = []
    if diff.get('added') or diff.get('removed') or diff.get('changed'):
        try:
            trigger_engine = TriggerEngine()
            trigger_results = trigger_engine.evaluate_and_execute(
                employer_cfg=employer_cfg,
                diff=diff,
                current_data=current,
                executed_by='system'
            )
            if trigger_results:
                custom_log("TRIGGER", f"{len(trigger_results)} Trigger für '{employer_name}' ausgeführt", "cyan")
        except Exception as e:
            custom_log("TRIGGER", f"Fehler bei Trigger-Auswertung: {str(e)}", "red")

    # --- If no changes, return early ---
    if not added_pids and not changed_pids and not always_write:
        return {"filepath": None, "diff": diff, "trigger_results": trigger_results}

    # --- Generate Excel file ---
    wb = openpyxl.Workbook()
    ws_emp = wb.active
    ws_emp.title = "Mitarbeiter"
    ws_emp.append(SCS_HEADERS)
    pids_for_export = sorted(list(added_pids | changed_pids))
    for pid in pids_for_export:
        ws_emp.append([current[pid]["core"].get(h, "") for h in SCS_HEADERS])

    ws_org = wb.create_sheet("Arbeitgeber")
    ORG_HEADERS = ["Name", "Strasse", "PLZ", "Ort", "Land", "Kommentar", "Email", "Telefon", "Fax"]
    ws_org.append(ORG_HEADERS)
    addr = employer_cfg.get("address", {}) if isinstance(employer_cfg.get("address"), dict) else {}
    org_row = {
        "Name": employer_cfg.get("name", "") or "",
        "Strasse": addr.get("street") or "",
        "PLZ": addr.get("zip_code") or addr.get("zipCode") or "",
        "Ort": addr.get("city") or "",
        "Land": addr.get("country") or "D",
        "Kommentar": employer_cfg.get("comment") or "",
        "Email": employer_cfg.get("email") or "",
        "Telefon": employer_cfg.get("phone") or "",
        "Fax": employer_cfg.get("fax") or "",
    }
    ws_org.append([org_row.get(h, "") for h in ORG_HEADERS])

    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    outfile = os.path.join(exports_dir, f"delta-{safe_emp}-{provider_key}-{ts}.xlsx")
    wb.save(outfile)

    return {"filepath": outfile, "diff": diff, "trigger_results": trigger_results}

def _get_employee_history_from_snapshots(employer_cfg: dict, snapshots_dir: str) -> dict:
    """
    Verarbeitet alle Snapshots für einen Arbeitgeber, um eine umfassende Historie für jeden Mitarbeiter zu erstellen.
    
    Args:
        employer_cfg (dict): Die Arbeitgeber-Konfiguration
        snapshots_dir (str): Das Verzeichnis mit den Snapshots
    
    Returns:
        dict: Dictionary der Mitarbeiterhistorien, indexiert nach Mitarbeiter-ID
    """
    # 1. Find and sort snapshot files
    safe_emp_name = _get_safe_employer_name(employer_cfg['name'])
    provider_key = employer_cfg['provider_key']
    snapshot_prefix = f"{safe_emp_name}-{provider_key}-"

    dated_snapshots = []
    if not os.path.exists(snapshots_dir):
        return {} # No snapshots directory, no history

    for filename in os.listdir(snapshots_dir):
        if filename.startswith(snapshot_prefix) and filename.endswith('.json') and 'latest' not in filename:
            try:
                timestamp_str = filename.replace(snapshot_prefix, '').replace('.json', '')
                dt_obj = datetime.strptime(timestamp_str, '%Y%m%d-%H%M%S')
                dated_snapshots.append({'filename': filename, 'dt': dt_obj})
            except ValueError:
                continue

    dated_snapshots.sort(key=lambda x: x['dt'])

    if not dated_snapshots:
        return {}

    # 2. Process snapshots to build history
    employee_history = {}

    # Get PIDs from the final snapshot to know who is currently active in the HR system
    last_snapshot_path = os.path.join(snapshots_dir, dated_snapshots[-1]['filename'])
    try:
        with open(last_snapshot_path, 'r', encoding='utf-8') as f:
            last_snapshot_data = json.load(f)
        pids_in_last_snapshot = set(last_snapshot_data.keys())
    except (FileNotFoundError, json.JSONDecodeError):
        pids_in_last_snapshot = set()


    for snap_info in dated_snapshots:
        snap_path = os.path.join(snapshots_dir, snap_info['filename'])
        snap_dt = snap_info['dt']

        try:
            with open(snap_path, 'r', encoding='utf-8') as f:
                snap_data = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            continue # Skip corrupted or missing snapshot file

        for pid, data in snap_data.items():
            core_data = data.get('core', {})
            dates = data.get('dates')

            # If the snapshot is old and doesn't have our standardized 'dates' field,
            # skip it for this analysis to ensure data integrity.
            if not dates:
                continue

            join_date_str = dates.get("join")
            leave_date_str = dates.get("leave")

            if pid not in employee_history:
                # First time seeing this employee
                employee_history[pid] = {
                    "pid": pid,
                    "name": f"{core_data.get('Vorname', '')} {core_data.get('Name', '')}".strip(),
                    "join_date_str": join_date_str,
                    "leave_date_str": leave_date_str,
                    "last_snapshot_dt": snap_dt,
                }
            else:
                # Update existing employee, keeping the latest info
                employee_history[pid]['last_snapshot_dt'] = snap_dt
                if join_date_str: # An earlier snapshot might not have had the join date
                    employee_history[pid]['join_date_str'] = join_date_str
                if leave_date_str: # A leave date might be added later
                    employee_history[pid]['leave_date_str'] = leave_date_str

    # 3. Post-process to parse dates and determine final status
    for pid, history in employee_history.items():
        history['is_active'] = pid in pids_in_last_snapshot and not history.get('leave_date_str')
        history['join_date'] = _parse_date(history['join_date_str'])

        # Determine the final leave date
        final_leave_date = _parse_date(history['leave_date_str'])
        if not final_leave_date and not history['is_active']:
            # If no explicit leave date and they are not in the last snapshot,
            # use their last seen snapshot date as the effective leave date.
            final_leave_date = history['last_snapshot_dt']

        history['leave_date'] = final_leave_date

    return employee_history


def calculate_statistics(ce, pe):
    """
    Berechnet verschiedene Statistiken basierend auf aktuellen und vorherigen Mitarbeiterlisten.
    
    Args:
        ce (list): Aktuelle Mitarbeiterliste
        pe (list): Vorherige Mitarbeiterliste
    
    Returns:
        dict: Dictionary mit verschiedenen Statistiken
    """
    if not ce: return {}
    ae=[e for e in ce if e.get('isActive')]; t=datetime.today()
    sc={"total":len(ce),"active":len(ae),"inactive":len(ce)-len(ae)}; gd=Counter(_getv(e, e.get("details"), "geschlecht", "gender") for e in ae)
    etd=Counter(_getv(e, e.get("details"), "beschäftigungsart", "employmentType") for e in ae); dd=Counter(_getv(e, e.get("details"), "abteilung", "organizationUnit.name") for e in ae)
    t5d=dd.most_common(5); th,ch,tt,tc,tah,hac=0,0,0,0,0,0
    joins_by_month, leaves_by_month = Counter(), Counter()
    for e in ce:
        hire_date = _parse_date(_getv(e, e.get("details"), "eintrittsdatum", "joinDate"))
        if hire_date: joins_by_month[hire_date.strftime('%Y-%m')] += 1
        term_date = _parse_date(_getv(e, e.get("details"), "kündigungsdatum", "leaveDate"))
        if term_date: leaves_by_month[term_date.strftime('%Y-%m')] += 1
    labels = []
    cy, cm = t.year, t.month
    for i in range(12):
        m, y = (cm - i - 1) % 12 + 1, cy + (cm - i - 1) // 12
        labels.append(f"{y}-{m:02d}")
    labels.reverse()
    jlt = {"labels": labels, "joins": [joins_by_month.get(l, 0) for l in labels], "leaves": [leaves_by_month.get(l, 0) for l in labels]}
    for e in ae:
        jd=_parse_date(_getv(e, e.get("details"), "eintrittsdatum", "joinDate")); bd=_parse_date(_getv(e, e.get("details"), "geburtsdatum", "birthday"))
        if jd: tt+=(t-jd).days/365.25; tc+=1
        if jd and bd: tah+=(jd-bd).days/365.25; hac+=1
        h_str = _get_from_path(e, "workSchedule.weeklyWorkingHours")
        if h_str:
            try:
                h = float(str(h_str).replace(",", "."))
                th += h
                ch += 1
            except (ValueError, TypeError):
                pass
    ah=round(th/ch,2) if ch>0 else 0; at=round(tt/tc,1) if tc>0 else 0; aha=round(tah/hac,1) if hac>0 else 0
    tr,tp=0,"N/A"
    if pe:
        tp="Since Last Snapshot"
        ci={_person_key(e) for e in ce}; pi={_person_key(e) for e in pe}
        dl=len(pi-ci); hs=len([e for e in pe if e.get('isActive')])
        tr=round((dl/hs)*100,2) if hs>0 else 0
    return {"status_counts":sc,"gender_distribution":{"labels":list(gd.keys()),"data":list(gd.values())},"average_weekly_hours":ah,"employment_type_distribution":{"labels":list(etd.keys()),"data":list(etd.values())},"department_distribution":{"labels":[d[0] for d in t5d],"data":[d[1] for d in t5d]},"averages":{"tenure_years":at,"hiring_age":aha},"turnover":{"period":tp,"rate_percent":tr},"join_leave_trends":jlt}


def calculate_long_term_statistics(employee_history: dict) -> dict:
    """
    Berechnet Langzeit-Statistiken basierend auf der vollständigen Mitarbeiterhistorie aus Snapshots.
    
    Args:
        employee_history (dict): Die Mitarbeiterhistorie aus Snapshots
    
    Returns:
        dict: Dictionary mit Langzeit-Statistiken
    """
    if not employee_history:
        return {}

    entries_per_year = Counter()
    exits_per_year = Counter()
    total_tenure_days = 0
    employees_with_tenure = 0
    today = datetime.now()

    for pid, history in employee_history.items():
        join_date = history.get('join_date')
        leave_date = history.get('leave_date')

        # 1. Tally entries and exits by year
        if join_date:
            entries_per_year[join_date.year] += 1

        if leave_date:
            exits_per_year[leave_date.year] += 1

        # 2. Calculate tenure for average duration
        # Use today as the end date for active employees, as requested
        effective_leave_date = leave_date or (today if history.get('is_active') else None)

        if join_date and effective_leave_date:
            # Ensure join_date is not in the future relative to the leave date
            if effective_leave_date >= join_date:
                duration = effective_leave_date - join_date
                total_tenure_days += duration.days
                employees_with_tenure += 1

    # 3. Prepare data for the return dictionary

    # --- Entries/Exits Chart Data ---
    all_years = sorted(list(set(entries_per_year.keys()) | set(exits_per_year.keys())))
    if not all_years:
        min_year, max_year = today.year, today.year
    else:
        min_year, max_year = min(all_years), max(all_years)

    # Create a continuous list of years from min to max
    year_labels = list(range(min_year, max_year + 1))
    entry_data = [entries_per_year.get(year, 0) for year in year_labels]
    exit_data = [exits_per_year.get(year, 0) for year in year_labels]

    entries_exits_stats = {
        "labels": [str(y) for y in year_labels],
        "entries": entry_data,
        "exits": exit_data
    }

    # --- Average Tenure ---
    avg_tenure_years = 0
    if employees_with_tenure > 0:
        avg_tenure_days = total_tenure_days / employees_with_tenure
        avg_tenure_years = round(avg_tenure_days / 365.25, 1)

    avg_duration_stats = {
        "years": avg_tenure_years,
        "total_employees_included": employees_with_tenure
    }

    return {
        "entries_exits_by_year": entries_exits_stats,
        "average_employment_duration": avg_duration_stats,
    }

def _format_stats_for_export(stats: dict, mode: str) -> str:
    """
    Formatiert ein Statistik-Dictionary in einen menschenlesbaren String für TXT-Export.
    
    Args:
        stats (dict): Die zu formatierenden Statistiken
        mode (str): Der Modus ('standard' oder 'longterm')
    
    Returns:
        str: Der formatierte String für den Export
    """
    lines = []
    now = datetime.now().strftime("%d.%m.%Y %H:%M:%S")

    if mode == 'standard':
        lines.append("Standard-Statistiken")
        lines.append(f"Exportiert am: {now}")
        lines.append("="*30)

        sc = stats.get('status_counts', {})
        lines.append(f"\nMitarbeiter: {sc.get('active', 0)} Aktiv / {sc.get('total', 0)} Gesamt")

        avg = stats.get('averages', {})
        lines.append(f"Durchschnittliche Betriebszugehörigkeit: {avg.get('tenure_years', 0)} Jahre")
        lines.append(f"Durchschnittliches Eintrittsalter: {avg.get('hiring_age', 0)} Jahre")

        turnover = stats.get('turnover', {})
        lines.append(f"Fluktuation ({turnover.get('period', 'N/A')}): {turnover.get('rate_percent', 0)}%")
        lines.append("\n--- Verteilungen ---\n")

        gd = stats.get('gender_distribution', {})
        lines.append("Geschlechterverteilung:")
        for label, data in zip(gd.get('labels', []), gd.get('data', [])):
            lines.append(f"- {label}: {data}")

        etd = stats.get('employment_type_distribution', {})
        lines.append("\nBeschäftigungsart:")
        for label, data in zip(etd.get('labels', []), etd.get('data', [])):
            lines.append(f"- {label}: {data}")

    elif mode == 'longterm':
        lines.append("Langzeit-Analyse")
        lines.append(f"Exportiert am: {now}")
        lines.append("="*30)

        avg = stats.get('average_employment_duration', {})
        lines.append(f"\nDurchschnittliche Betriebszugehörigkeit (Gesamt): {avg.get('years', 0)} Jahre")
        lines.append(f"Mitarbeiter in Berechnung berücksichtigt: {avg.get('total_employees_included', 0)}")
        lines.append("\n--- Eintritte & Austritte pro Jahr ---\n")

        ee = stats.get('entries_exits_by_year', {})
        lines.append("Jahr\tEintritte\tAustritte")
        for i, year in enumerate(ee.get('labels', [])):
            entries = ee.get('entries', [])[i]
            exits = ee.get('exits', [])[i]
            lines.append(f"{year}\t{entries}\t\t{exits}")

    return "\n".join(lines)

# ==============================================================================
# --- SECTION 4: FLASK APP AND ROUTES ---
# ==============================================================================

app = Flask(__name__)

# ==============================================================================
# --- SECURITY FIX SV-004: CSRF-Schutz mit Flask-WTF ---
# ==============================================================================
# Aktiviert CSRF-Schutz für alle POST-Formulare und AJAX-Requests
try:
    from flask_wtf.csrf import CSRFProtect, CSRFError
    csrf = CSRFProtect(app)
    _csrf_enabled = True
    
    @app.errorhandler(CSRFError)
    def handle_csrf_error(e):
        """Handler für CSRF-Fehler (SV-004)."""
        flash('Sicherheitsfehler: Ungültiges oder fehlendes CSRF-Token. Bitte versuchen Sie es erneut.', 'error')
        return redirect(request.referrer or url_for('index'))
except ImportError:
    print("[WARNUNG] Flask-WTF nicht installiert. CSRF-Schutz deaktiviert.")
    _csrf_enabled = False
    csrf = None

# ==============================================================================
# --- SECURITY FIX SV-001: Secret Key aus Umgebungsvariable ---
# ==============================================================================
# Der Secret Key wird aus der Umgebungsvariable ACENCIA_SECRET_KEY geladen.
# In Produktion MUSS dieser gesetzt sein. Für Entwicklung wird ein Fallback verwendet.
_secret_key = os.environ.get('ACENCIA_SECRET_KEY')
if not _secret_key:
    import secrets as _secrets_module
    _secret_key = _secrets_module.token_hex(32)
    print("[WARNUNG] ACENCIA_SECRET_KEY nicht gesetzt! Verwende temporären Key.")
    print("[WARNUNG] Sessions werden bei Neustart invalidiert. Für Produktion: ACENCIA_SECRET_KEY setzen!")
app.secret_key = _secret_key

# ==============================================================================
# --- SECURITY FIX SV-013: Session-Timeout Konfiguration ---
# ==============================================================================
# Sessions laufen nach 8 Stunden Inaktivität ab.
# SESSION_REFRESH_EACH_REQUEST verlängert bei jeder Anfrage.
from datetime import timedelta
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)
app.config['SESSION_REFRESH_EACH_REQUEST'] = True

# ==============================================================================
# --- SECURITY FIX SV-018: Secure Cookie Flags ---
# ==============================================================================
# SESSION_COOKIE_HTTPONLY: Verhindert JavaScript-Zugriff auf Session-Cookie
# SESSION_COOKIE_SAMESITE: Schützt vor CSRF bei Cross-Site-Requests
# SESSION_COOKIE_SECURE: Sollte in Produktion mit HTTPS auf True gesetzt werden
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
# Secure-Flag nur wenn HTTPS_ENABLED Umgebungsvariable gesetzt ist
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('HTTPS_ENABLED', 'false').lower() == 'true'

# ==============================================================================
# --- SECURITY FIX SV-006: Rate-Limiting für Brute-Force-Schutz ---
# ==============================================================================
# Verwendet Flask-Limiter um Login-Versuche zu begrenzen.
# Konfiguration: 5 Versuche pro Minute auf /login, 200/Tag global
try:
    from flask_limiter import Limiter
    from flask_limiter.util import get_remote_address
    
    limiter = Limiter(
        key_func=get_remote_address,
        app=app,
        default_limits=["200 per day", "50 per hour"],
        storage_uri="memory://",  # In-Memory Storage (für Produktionscluster: Redis verwenden)
    )
    _rate_limiter_enabled = True
except ImportError:
    # Falls Flask-Limiter nicht installiert ist, deaktivieren wir Rate-Limiting
    print("[WARNUNG] Flask-Limiter nicht installiert. Rate-Limiting deaktiviert.")
    _rate_limiter_enabled = False
    limiter = None

# ==============================================================================
# --- SECURITY FIX SV-008: Security Headers Middleware ---
# ==============================================================================
# Fügt wichtige Security-Header zu allen HTTP-Responses hinzu.
# - X-Frame-Options: Verhindert Clickjacking
# - X-Content-Type-Options: Verhindert MIME-Sniffing
# - X-XSS-Protection: Legacy XSS-Schutz für ältere Browser
# - Referrer-Policy: Kontrolliert Referrer-Informationen
# - Content-Security-Policy: Beschränkt Ressourcen-Quellen
# - Strict-Transport-Security: Erzwingt HTTPS (nur wenn TLS aktiv)
@app.after_request
def add_security_headers(response):
    """Fügt Security-Header zu allen Responses hinzu (SV-008)."""
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    # CSP - permissiv für Google Fonts, Chart.js CDN und Inline-Styles/Scripts
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
        "img-src 'self' data:; "
        "connect-src 'self'"
    )
    # HSTS nur bei HTTPS-Verbindungen setzen
    if request.is_secure:
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    return response

# ==============================================================================
# --- SECURITY FIX SV-009: Arbeitgeber-Zugriffskontrolle ---
# ==============================================================================
# Decorator zur Prüfung, ob ein Benutzer Zugriff auf einen bestimmten Arbeitgeber hat.
# Master-Benutzer haben immer Zugriff. Normale Benutzer nur auf zugewiesene Arbeitgeber.
from functools import wraps

def check_employer_access(employer_id: str) -> bool:
    """
    Prüft ob der aktuelle Benutzer Zugriff auf einen Arbeitgeber hat (SV-009).
    
    Args:
        employer_id (str): Die ID des Arbeitgebers
    
    Returns:
        bool: True wenn Zugriff erlaubt, False sonst
    """
    user_info = session.get('user_info', {})
    
    # Nicht eingeloggt? Kein Zugriff
    if not user_info:
        return False
    
    # Master hat immer Zugriff
    if user_info.get('is_master'):
        return True
    
    # Prüfe allowed_employers
    allowed = user_info.get('allowed_employers', [])
    
    # Falls allowed_employers leer ist, laden wir aus der users.json
    if not allowed:
        users = load_users()
        user = next((u for u in users if u['username'] == user_info.get('username')), None)
        if user:
            allowed = user.get('allowed_employers', [])
            # Auch in Session speichern für spätere Zugriffe
            if 'user_info' in session:
                session['user_info']['allowed_employers'] = allowed
    
    return employer_id in allowed

@app.before_request
def check_employer_route_access():
    """
    Before-Request-Handler für Arbeitgeber-Zugriffskontrolle (SV-009).
    
    Prüft alle /employer/<id>/* Routen auf Berechtigung.
    Ausgenommen: /employer/add (neue Arbeitgeber hinzufügen)
    """
    # Nur für /employer/* Routen (außer /employer/add)
    if not request.path.startswith('/employer/'):
        return None
    if request.path == '/employer/add':
        return None
    
    # Extrahiere employer_id aus dem Pfad
    path_parts = request.path.split('/')
    if len(path_parts) < 3:
        return None
    
    employer_id = path_parts[2]
    
    # Prüfe Zugriff
    if not check_employer_access(employer_id):
        user_info = session.get('user_info', {})
        if not user_info:
            return redirect(url_for('login'))
        flash("Zugriff verweigert. Sie haben keine Berechtigung für diesen Arbeitgeber.", "error")
        # Audit-Log für verweigerten Zugriff
        audit_log(
            user=user_info.get('username', 'unknown'),
            action="ACCESS_DENIED",
            target=employer_id,
            details="employer_access",
            ip=request.remote_addr or 'unknown'
        )
        return redirect(url_for('index'))
    
    return None

# Path for data files
DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')
USERS_FILE = os.path.join(DATA_DIR, 'users.json')
SECRETS_FILE = os.path.join(DATA_DIR, 'secrets.json')
FORCED_LOGOUT_FILE = os.path.join(DATA_DIR, 'force_logout.txt')

def load_users():
    """
    Lädt Benutzer aus der JSON-Datei.
    
    Returns:
        list: Liste der Benutzer oder leere Liste bei Fehlern
    """
    if not os.path.exists(USERS_FILE):
        return []
    with open(USERS_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def write_users(users):
    """
    Speichert die Benutzerliste in der JSON-Datei.
    
    Args:
        users (list): Die zu speichernde Benutzerliste
    
    Returns:
        None
    """
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(USERS_FILE, 'w', encoding='utf-8') as f:
        json.dump(users, f, indent=4, ensure_ascii=False)

def load_secrets():
    """
    Lädt Geheimnisse (wie GitHub PAT) aus der JSON-Datei.
    
    SECURITY FIX SV-007: PAT wird automatisch entschlüsselt.
    
    Returns:
        dict: Dictionary der Geheimnisse mit entschlüsselten Werten
    """
    if not os.path.exists(SECRETS_FILE):
        return {}
    with open(SECRETS_FILE, 'r', encoding='utf-8') as f:
        try:
            secrets = json.load(f)
            # SV-007: PAT entschlüsseln wenn vorhanden
            if 'github_pat' in secrets and secrets['github_pat']:
                secrets['github_pat'] = decrypt_credential(secrets['github_pat'])
            return secrets
        except json.JSONDecodeError:
            return {}

def save_secrets(secrets):
    """
    Speichert Geheimnisse in der JSON-Datei.
    
    SECURITY FIX SV-007: PAT wird automatisch verschlüsselt.
    
    Args:
        secrets (dict): Die zu speichernden Geheimnisse (Klartext)
    
    Returns:
        None
    """
    os.makedirs(DATA_DIR, exist_ok=True)
    # Kopie erstellen um Original nicht zu ändern
    secrets_to_save = secrets.copy()
    # SV-007: PAT verschlüsseln wenn vorhanden und noch nicht verschlüsselt
    if 'github_pat' in secrets_to_save and secrets_to_save['github_pat']:
        if not is_encrypted(secrets_to_save['github_pat']):
            secrets_to_save['github_pat'] = encrypt_credential(secrets_to_save['github_pat'])
    with open(SECRETS_FILE, 'w', encoding='utf-8') as f:
        json.dump(secrets_to_save, f, indent=4, ensure_ascii=False)

# ==============================================================================
# --- SECURITY FIX SV-012: Passwort-Validierung ---
# ==============================================================================
def validate_password(password: str) -> tuple:
    """
    Validiert ein Passwort gegen die Sicherheits-Policy.
    
    Anforderungen:
    - Mindestens 8 Zeichen
    - Mindestens ein Großbuchstabe
    - Mindestens ein Kleinbuchstabe
    - Mindestens eine Ziffer
    
    Args:
        password (str): Das zu validierende Passwort
    
    Returns:
        tuple[bool, str]: (Gültig, Fehlermeldung wenn ungültig)
    """
    if not password:
        return False, "Passwort darf nicht leer sein."
    if len(password) < 8:
        return False, "Passwort muss mindestens 8 Zeichen lang sein."
    if not any(c.isupper() for c in password):
        return False, "Passwort muss mindestens einen Großbuchstaben enthalten."
    if not any(c.islower() for c in password):
        return False, "Passwort muss mindestens einen Kleinbuchstaben enthalten."
    if not any(c.isdigit() for c in password):
        return False, "Passwort muss mindestens eine Ziffer enthalten."
    return True, ""

# ==============================================================================
# --- SECURITY FIX SV-020: Account-Lockout Hilfsfunktionen ---
# ==============================================================================
# Konfiguration: 5 Fehlversuche führen zu 15 Minuten Sperre
LOCKOUT_MAX_ATTEMPTS = 5
LOCKOUT_DURATION_MINUTES = 15

def check_account_locked(user: dict) -> tuple:
    """
    Prüft ob ein Benutzer-Account gesperrt ist.
    
    Args:
        user (dict): Der Benutzer-Datensatz
    
    Returns:
        tuple[bool, int]: (ist_gesperrt, verbleibende_minuten)
    """
    locked_until = user.get('locked_until')
    if not locked_until:
        return False, 0
    
    current_time = datetime.utcnow().timestamp()
    if current_time < locked_until:
        remaining = int((locked_until - current_time) / 60) + 1
        return True, remaining
    return False, 0

def record_failed_login(users: list, username: str) -> bool:
    """
    Zeichnet einen fehlgeschlagenen Login-Versuch auf und sperrt ggf. den Account.
    
    Args:
        users (list): Die Benutzerliste
        username (str): Der Benutzername
    
    Returns:
        bool: True wenn Account jetzt gesperrt wurde
    """
    user = next((u for u in users if u['username'] == username), None)
    if not user:
        return False
    
    # Zähler erhöhen
    user['failed_attempts'] = user.get('failed_attempts', 0) + 1
    
    # Bei Erreichen des Limits: sperren
    if user['failed_attempts'] >= LOCKOUT_MAX_ATTEMPTS:
        user['locked_until'] = datetime.utcnow().timestamp() + (LOCKOUT_DURATION_MINUTES * 60)
        write_users(users)
        return True
    
    write_users(users)
    return False

def reset_failed_attempts(users: list, username: str):
    """
    Setzt den Fehlversuch-Zähler nach erfolgreichem Login zurück.
    
    Args:
        users (list): Die Benutzerliste
        username (str): Der Benutzername
    """
    user = next((u for u in users if u['username'] == username), None)
    if user:
        user['failed_attempts'] = 0
        user['locked_until'] = None
        write_users(users)

def get_forced_logout_time():
    """
    Liest den Zeitstempel für erzwungenes Abmelden aus einer Datei.
    
    Returns:
        float | None: Der Zeitstempel oder None bei Fehlern
    """
    if not os.path.exists(FORCED_LOGOUT_FILE):
        return None
    with open(FORCED_LOGOUT_FILE, 'r') as f:
        try:
            return float(f.read().strip())
        except (ValueError, TypeError):
            return None

def set_forced_logout_time():
    """
    Schreibt den aktuellen Zeitstempel in die Datei für erzwungenes Abmelden.
    
    Returns:
        None
    """
    with open(FORCED_LOGOUT_FILE, 'w') as f:
        f.write(str(datetime.utcnow().timestamp()))

# ==============================================================================
# --- SECURITY FIX SV-006: Rate-Limiting auf Login-Route ---
# ==============================================================================
# Begrenzt Login-Versuche auf 5 pro Minute pro IP-Adresse
def _login_rate_limit_decorator(f):
    """Wendet Rate-Limiting an, falls Flask-Limiter verfügbar ist."""
    if _rate_limiter_enabled and limiter:
        return limiter.limit("5 per minute")(f)
    return f

@app.route('/login', methods=['GET', 'POST'])
@_login_rate_limit_decorator
def login():
    """
    Behandelt die Anmeldung von Benutzern.
    
    GET: Zeigt das Anmeldeformular an
    POST: Verarbeitet die Anmeldedaten und authentifiziert den Benutzer
    
    Rate-Limit: 5 Versuche pro Minute (SV-006)
    
    Returns:
        Response: Weiterleitung zur Hauptseite oder Anmeldeformular
    """
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        users = load_users()

        user = next((u for u in users if u['username'] == username), None)
        
        # ==============================================================================
        # --- SECURITY FIX SV-020: Account-Lockout Prüfung ---
        # ==============================================================================
        if user:
            is_locked, remaining_minutes = check_account_locked(user)
            if is_locked:
                flash(f'Ihr Account ist gesperrt. Versuchen Sie es in {remaining_minutes} Minuten erneut.', 'error')
                client_ip = request.remote_addr or 'unknown'
                custom_log("SECURITY", f"Zugriff auf gesperrten Account '{username[:3]}***' von IP {client_ip}", "red")
                return render_template('login.html')

        if user and check_password_hash(user['password_hash'], password):
            # Erfolgreicher Login: Fehlversuche zurücksetzen (SV-020)
            reset_failed_attempts(users, username)
            session['user_id'] = user['username'] # Using username as a simple ID
            session['user_info'] = {
                'username': user['username'],
                'kuerzel': user['kuerzel'],
                'is_master': user.get('is_master', False),
                'color': user.get('color'),
                'theme': user.get('theme', 'light') # Load theme, default to light
            }
            session['login_time'] = datetime.utcnow().timestamp()
            session.permanent = True  # Für PERMANENT_SESSION_LIFETIME (SV-013)
            custom_log(user['kuerzel'], "wurde angemeldet", user.get('color'))
            return redirect(url_for('index'))
        else:
            # ==============================================================================
            # --- SECURITY FIX SV-020: Fehlversuch aufzeichnen ---
            # ==============================================================================
            account_locked = False
            if user:
                account_locked = record_failed_login(users, username)
            
            if account_locked:
                flash(f'Ihr Account wurde nach {LOCKOUT_MAX_ATTEMPTS} Fehlversuchen für {LOCKOUT_DURATION_MINUTES} Minuten gesperrt.', 'error')
            else:
                flash('Ungültiger Benutzername oder Passwort.', 'error')
            
            # ==============================================================================
            # --- SECURITY FIX SV-021: Failed Login Logging ---
            # ==============================================================================
            # Protokolliert fehlgeschlagene Login-Versuche mit IP-Adresse und Benutzername
            client_ip = request.remote_addr or 'unknown'
            # Benutzername anonymisieren (nur erste 3 Zeichen + ***)
            safe_username = (username[:3] + '***') if username and len(username) > 3 else '***'
            log_msg = f"Fehlgeschlagener Login-Versuch für '{safe_username}' von IP {client_ip}"
            if account_locked:
                log_msg += " - ACCOUNT GESPERRT"
            custom_log("SECURITY", log_msg, "red")

    return render_template('login.html')

@app.route('/logout')
def logout():
    """
    Behandelt die Abmeldung von Benutzern.
    
    Returns:
        Response: Weiterleitung zur Anmeldeseite
    """
    session.clear()
    flash('Sie wurden erfolgreich abgemeldet.', 'success')
    return redirect(url_for('login'))

@app.route('/settings', methods=['GET', 'POST'])
def settings():
    """
    Behandelt die Master-Einstellungen der Anwendung.
    
    GET: Zeigt die Einstellungsseite an
    POST: Verarbeitet Änderungen an Benutzern oder Einstellungen
    
    Returns:
        Response: Einstellungsseite oder Weiterleitung
    """
    user_info = session.get('user_info', {})
    if not user_info.get('is_master'):
        flash("Zugriff verweigert. Nur für Master-Benutzer.", "error")
        return redirect(url_for('index'))

    if request.method == 'POST':
        action = request.form.get('action')
        users = load_users()

        if action == 'add_user':
            username = request.form.get('username')
            password = request.form.get('password')
            kuerzel = request.form.get('kuerzel')
            color = request.form.get('color')
            is_master = request.form.get('is_master') == 'true'

            if any(u['username'] == username for u in users):
                flash(f"Benutzername '{username}' existiert bereits.", "error")
            else:
                # SECURITY FIX SV-012: Passwort-Policy prüfen bei Benutzer-Erstellung
                is_valid, error_msg = validate_password(password)
                if not is_valid:
                    flash(error_msg, "error")
                else:
                    new_user = {
                        "username": username,
                        "password_hash": generate_password_hash(password),
                        "kuerzel": kuerzel,
                        "color": color,
                        "is_master": is_master
                    }
                    users.append(new_user)
                    write_users(users)
                    flash(f"Benutzer '{username}' erfolgreich angelegt.", "success")
                    # SECURITY FIX SV-016: Audit-Log für Benutzer-Erstellung
                    audit_log(
                        user=user_info.get('username', 'unknown'),
                        action="CREATE_USER",
                        target=username,
                        details=f"is_master={is_master}",
                        ip=request.remote_addr or 'unknown'
                    )

        elif action == 'delete_user':
            username_to_delete = request.form.get('username')
            if username_to_delete == user_info.get('username'):
                flash("Sie können sich nicht selbst löschen.", "error")
            else:
                users = [u for u in users if u['username'] != username_to_delete]
                write_users(users)
                flash(f"Benutzer '{username_to_delete}' gelöscht.", "success")
                # SECURITY FIX SV-016: Audit-Log für Benutzer-Löschung
                audit_log(
                    user=user_info.get('username', 'unknown'),
                    action="DELETE_USER",
                    target=username_to_delete,
                    ip=request.remote_addr or 'unknown'
                )

        elif action == 'logout_all':
            set_forced_logout_time()
            flash("Alle anderen Benutzer werden bei ihrer nächsten Aktion abgemeldet.", "success")
            # SECURITY FIX SV-016: Audit-Log für Force-Logout
            audit_log(
                user=user_info.get('username', 'unknown'),
                action="FORCE_LOGOUT_ALL",
                target="all_users",
                ip=request.remote_addr or 'unknown'
            )

        elif action == 'save_pat':
            pat = request.form.get('github_pat', '').strip()
            secrets = load_secrets()
            secrets['github_pat'] = pat
            save_secrets(secrets)
            flash("GitHub Personal Access Token gespeichert.", "success")
            # SECURITY FIX SV-016: Audit-Log für PAT-Änderung
            audit_log(
                user=user_info.get('username', 'unknown'),
                action="UPDATE_GITHUB_PAT",
                target="secrets.json",
                details="PAT updated",
                ip=request.remote_addr or 'unknown'
            )

        return redirect(url_for('settings'))

    # For GET request
    custom_log(user_info.get('kuerzel'), "Master-Einstellungen aufgerufen", user_info.get('color'))
    all_users = load_users()
    secrets = load_secrets()
    github_pat = secrets.get('github_pat', '')
    return render_template('settings.html', users=all_users, github_pat=github_pat)


# ==============================================================================
# --- TRIGGER SYSTEM: Routes ---
# ==============================================================================
# Diese Routen verwalten die Trigger-Konfiguration und das Ausführungsprotokoll.

# Initialize stores (singletons)
trigger_store = TriggerStore()
trigger_log_store = TriggerLogStore()


@app.route('/settings/triggers')
def triggers_overview():
    """
    Zeigt die Übersicht aller Trigger an (Master-only).
    
    Returns:
        Response: Trigger-Übersichtsseite
    """
    user_info = session.get('user_info', {})
    if not user_info.get('is_master'):
        flash("Zugriff verweigert. Nur für Master-Benutzer.", "error")
        return redirect(url_for('index'))
    
    custom_log(user_info.get('kuerzel'), "Trigger-Verwaltung aufgerufen", user_info.get('color'))
    
    triggers = trigger_store.get_all_triggers()
    employers = employer_store.get_all()
    smtp_config = trigger_store.get_smtp_config()
    smtp_configured = bool(smtp_config.get('host'))
    
    return render_template('triggers.html', 
                          triggers=triggers, 
                          employers=employers,
                          smtp_configured=smtp_configured,
                          scs_headers=SCS_HEADERS,
                          trigger_events=TriggerStore.TRIGGER_EVENTS,
                          condition_operators=TriggerStore.CONDITION_OPERATORS,
                          action_types=TriggerStore.ACTION_TYPES)


@app.route('/settings/triggers/new', methods=['GET', 'POST'])
def trigger_new():
    """
    Erstellt einen neuen Trigger (Master-only).
    
    GET: Zeigt das Formular zur Trigger-Erstellung an
    POST: Speichert den neuen Trigger
    
    Returns:
        Response: Trigger-Formular oder Weiterleitung
    """
    user_info = session.get('user_info', {})
    if not user_info.get('is_master'):
        flash("Zugriff verweigert. Nur für Master-Benutzer.", "error")
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        try:
            trigger_data = _parse_trigger_form(request.form)
            trigger = trigger_store.add_trigger(trigger_data, created_by=user_info.get('username', 'system'))
            
            audit_log(
                user=user_info.get('username', 'unknown'),
                action="CREATE_TRIGGER",
                target=trigger['id'],
                details=f"name={trigger['name']}",
                ip=request.remote_addr or 'unknown'
            )
            
            flash(f"Trigger '{trigger['name']}' erfolgreich erstellt.", "success")
            return redirect(url_for('triggers_overview'))
        except Exception as e:
            flash(f"Fehler beim Erstellen des Triggers: {str(e)}", "error")
    
    employers = employer_store.get_all()
    return render_template('trigger_form.html',
                          trigger=None,
                          employers=employers,
                          scs_headers=SCS_HEADERS,
                          trigger_events=TriggerStore.TRIGGER_EVENTS,
                          condition_operators=TriggerStore.CONDITION_OPERATORS,
                          action_types=TriggerStore.ACTION_TYPES)


@app.route('/settings/triggers/<trigger_id>/edit', methods=['GET', 'POST'])
def trigger_edit(trigger_id):
    """
    Bearbeitet einen bestehenden Trigger (Master-only).
    
    Args:
        trigger_id (str): Die Trigger-ID
    
    Returns:
        Response: Trigger-Formular oder Weiterleitung
    """
    user_info = session.get('user_info', {})
    if not user_info.get('is_master'):
        flash("Zugriff verweigert. Nur für Master-Benutzer.", "error")
        return redirect(url_for('index'))
    
    trigger = trigger_store.get_trigger_by_id(trigger_id)
    if not trigger:
        flash("Trigger nicht gefunden.", "error")
        return redirect(url_for('triggers_overview'))
    
    if request.method == 'POST':
        try:
            trigger_data = _parse_trigger_form(request.form)
            trigger_store.update_trigger(trigger_id, trigger_data)
            
            audit_log(
                user=user_info.get('username', 'unknown'),
                action="UPDATE_TRIGGER",
                target=trigger_id,
                details=f"name={trigger_data.get('name')}",
                ip=request.remote_addr or 'unknown'
            )
            
            flash(f"Trigger '{trigger_data['name']}' erfolgreich aktualisiert.", "success")
            return redirect(url_for('triggers_overview'))
        except Exception as e:
            flash(f"Fehler beim Aktualisieren des Triggers: {str(e)}", "error")
    
    employers = employer_store.get_all()
    return render_template('trigger_form.html',
                          trigger=trigger,
                          employers=employers,
                          scs_headers=SCS_HEADERS,
                          trigger_events=TriggerStore.TRIGGER_EVENTS,
                          condition_operators=TriggerStore.CONDITION_OPERATORS,
                          action_types=TriggerStore.ACTION_TYPES)


@app.route('/settings/triggers/<trigger_id>/delete', methods=['POST'])
def trigger_delete(trigger_id):
    """
    Löscht einen Trigger (Master-only).
    
    Args:
        trigger_id (str): Die Trigger-ID
    
    Returns:
        Response: Weiterleitung zur Trigger-Übersicht
    """
    user_info = session.get('user_info', {})
    if not user_info.get('is_master'):
        flash("Zugriff verweigert. Nur für Master-Benutzer.", "error")
        return redirect(url_for('index'))
    
    trigger = trigger_store.get_trigger_by_id(trigger_id)
    if trigger:
        trigger_store.delete_trigger(trigger_id)
        audit_log(
            user=user_info.get('username', 'unknown'),
            action="DELETE_TRIGGER",
            target=trigger_id,
            details=f"name={trigger.get('name')}",
            ip=request.remote_addr or 'unknown'
        )
        flash(f"Trigger '{trigger.get('name')}' erfolgreich gelöscht.", "success")
    else:
        flash("Trigger nicht gefunden.", "error")
    
    return redirect(url_for('triggers_overview'))


@app.route('/settings/triggers/<trigger_id>/toggle', methods=['POST'])
def trigger_toggle(trigger_id):
    """
    Aktiviert oder deaktiviert einen Trigger (Master-only).
    
    Args:
        trigger_id (str): Die Trigger-ID
    
    Returns:
        Response: JSON-Antwort mit neuem Status
    """
    user_info = session.get('user_info', {})
    if not user_info.get('is_master'):
        return jsonify({'status': 'error', 'message': 'Zugriff verweigert'}), 403
    
    new_status = trigger_store.toggle_trigger(trigger_id)
    if new_status is not None:
        trigger = trigger_store.get_trigger_by_id(trigger_id)
        audit_log(
            user=user_info.get('username', 'unknown'),
            action="TOGGLE_TRIGGER",
            target=trigger_id,
            details=f"enabled={new_status}",
            ip=request.remote_addr or 'unknown'
        )
        return jsonify({'status': 'success', 'enabled': new_status, 'name': trigger.get('name')})
    
    return jsonify({'status': 'error', 'message': 'Trigger nicht gefunden'}), 404


@app.route('/settings/smtp', methods=['GET', 'POST'])
def smtp_settings():
    """
    Verwaltet die SMTP-Konfiguration (Master-only).
    
    GET: Zeigt die SMTP-Einstellungen an
    POST: Speichert die SMTP-Konfiguration
    
    Returns:
        Response: SMTP-Einstellungsseite oder Weiterleitung
    """
    user_info = session.get('user_info', {})
    if not user_info.get('is_master'):
        flash("Zugriff verweigert. Nur für Master-Benutzer.", "error")
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        smtp_config = {
            'host': request.form.get('smtp_host', '').strip(),
            'port': int(request.form.get('smtp_port', 587)),
            'username': request.form.get('smtp_username', '').strip(),
            'password': request.form.get('smtp_password', '').strip(),
            'from_email': request.form.get('smtp_from_email', '').strip(),
            'use_tls': request.form.get('smtp_use_tls') == 'on'
        }
        
        # If password is empty, keep the existing one
        existing_config = trigger_store.get_smtp_config()
        if not smtp_config['password'] and existing_config.get('password'):
            smtp_config['password'] = existing_config['password']
        
        trigger_store.update_smtp_config(smtp_config)
        
        audit_log(
            user=user_info.get('username', 'unknown'),
            action="UPDATE_SMTP_CONFIG",
            target="triggers.json",
            details=f"host={smtp_config['host']}",
            ip=request.remote_addr or 'unknown'
        )
        
        flash("SMTP-Konfiguration erfolgreich gespeichert.", "success")
        return redirect(url_for('smtp_settings'))
    
    custom_log(user_info.get('kuerzel'), "SMTP-Einstellungen aufgerufen", user_info.get('color'))
    smtp_config = trigger_store.get_smtp_config()
    # Don't show password in form
    smtp_config_display = dict(smtp_config)
    smtp_config_display['password'] = '********' if smtp_config.get('password') else ''
    
    return render_template('smtp_settings.html', smtp_config=smtp_config_display)


@app.route('/settings/smtp/test', methods=['POST'])
def smtp_test():
    """
    Sendet eine Test-E-Mail über die konfigurierte SMTP-Verbindung (Master-only).
    
    Returns:
        Response: JSON-Antwort mit Test-Ergebnis
    """
    user_info = session.get('user_info', {})
    if not user_info.get('is_master'):
        return jsonify({'status': 'error', 'message': 'Zugriff verweigert'}), 403
    
    test_email = request.json.get('test_email', '').strip()
    if not test_email:
        return jsonify({'status': 'error', 'message': 'Keine Test-E-Mail-Adresse angegeben'}), 400
    
    smtp_config = trigger_store.get_smtp_config()
    if not smtp_config.get('host'):
        return jsonify({'status': 'error', 'message': 'SMTP nicht konfiguriert'}), 400
    
    try:
        import smtplib
        import socket
        from email.message import EmailMessage
        
        msg = EmailMessage()
        
        from_email = smtp_config.get('from_email', smtp_config.get('username', ''))
        msg['From'] = from_email
        msg['To'] = test_email
        msg['Subject'] = 'ACENCIA Hub - SMTP Test'
        
        body = """Dies ist eine Test-E-Mail von ACENCIA Hub.

Wenn Sie diese E-Mail erhalten, ist die SMTP-Konfiguration korrekt.

Zeitstempel: """ + datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        msg.set_content(body)
        
        # Explizit ASCII-sicheren lokalen Hostnamen setzen (vermeidet Umlaute im Computernamen)
        local_hostname = 'localhost'
        try:
            hostname = socket.gethostname()
            # Nur ASCII-sichere Hostnamen verwenden
            hostname.encode('ascii')
            local_hostname = hostname
        except (UnicodeEncodeError, socket.error):
            local_hostname = 'localhost'
        
        if smtp_config.get('use_tls', True):
            server = smtplib.SMTP(smtp_config['host'], smtp_config.get('port', 587), local_hostname=local_hostname)
            server.starttls()
        else:
            server = smtplib.SMTP(smtp_config['host'], smtp_config.get('port', 25), local_hostname=local_hostname)
        
        if smtp_config.get('username') and smtp_config.get('password'):
            server.login(smtp_config['username'], smtp_config['password'])
        
        server.send_message(msg)
        server.quit()
        
        custom_log(user_info.get('kuerzel'), f"SMTP-Test an {test_email} erfolgreich", user_info.get('color'))
        return jsonify({'status': 'success', 'message': f'Test-E-Mail erfolgreich an {test_email} gesendet'})
    
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        custom_log(user_info.get('kuerzel'), f"SMTP-Test fehlgeschlagen: {str(e)}", "red")
        print(f"[DEBUG SMTP] Full traceback:\n{error_details}")
        return jsonify({'status': 'error', 'message': f'SMTP-Fehler: {str(e)}'}), 500


@app.route('/api/triggers/fields')
def api_trigger_fields():
    """
    API-Endpunkt zum Abrufen der verfügbaren SCS-Felder für Trigger-Bedingungen.
    
    Returns:
        Response: JSON mit verfügbaren Feldern, Events und Operatoren
    """
    if not session.get('user_id'):
        return jsonify({'status': 'error', 'message': 'Nicht angemeldet'}), 401
    
    return jsonify({
        'fields': SCS_HEADERS,
        'events': TriggerStore.TRIGGER_EVENTS,
        'operators': TriggerStore.CONDITION_OPERATORS,
        'action_types': TriggerStore.ACTION_TYPES
    })


@app.route('/api/trigger-log')
def api_trigger_log():
    """
    API-Endpunkt zum Abrufen des Trigger-Ausführungsprotokolls.
    
    Query-Parameter:
    - employer_id: Filter nach Arbeitgeber
    - trigger_id: Filter nach Trigger
    - status: Filter nach Status (success, error, skipped)
    - from: Filter ab Datum (ISO-Format)
    - to: Filter bis Datum (ISO-Format)
    - limit: Maximale Anzahl (default: 100)
    - offset: Offset für Pagination
    
    Returns:
        Response: JSON mit Ausführungsprotokoll
    """
    if not session.get('user_id'):
        return jsonify({'status': 'error', 'message': 'Nicht angemeldet'}), 401
    
    result = trigger_log_store.get_executions(
        employer_id=request.args.get('employer_id'),
        trigger_id=request.args.get('trigger_id'),
        status=request.args.get('status'),
        from_date=request.args.get('from'),
        to_date=request.args.get('to'),
        limit=int(request.args.get('limit', 100)),
        offset=int(request.args.get('offset', 0))
    )
    
    return jsonify(result)


@app.route('/api/trigger-log/<execution_id>/retry', methods=['POST'])
def api_trigger_retry(execution_id):
    """
    Wiederholt eine fehlgeschlagene Trigger-Ausführung.
    
    Args:
        execution_id (str): Die ID der ursprünglichen Ausführung
    
    Returns:
        Response: JSON mit Ergebnis
    """
    user_info = session.get('user_info', {})
    if not user_info.get('is_master'):
        return jsonify({'status': 'error', 'message': 'Zugriff verweigert'}), 403
    
    trigger_engine = TriggerEngine()
    result = trigger_engine.retry_execution(
        execution_id=execution_id,
        executed_by=user_info.get('username', 'system')
    )
    
    if result:
        audit_log(
            user=user_info.get('username', 'unknown'),
            action="RETRY_TRIGGER",
            target=execution_id,
            details=f"new_execution_id={result.get('id')}",
            ip=request.remote_addr or 'unknown'
        )
        return jsonify({
            'status': 'success',
            'message': 'Trigger wurde erneut ausgeführt',
            'new_execution': result
        })
    
    return jsonify({'status': 'error', 'message': 'Ausführung nicht gefunden oder nicht wiederholbar'}), 404


@app.route('/settings/trigger-log')
def trigger_log_view():
    """
    Zeigt das Trigger-Ausführungsprotokoll an (Master-only).
    
    Returns:
        Response: Trigger-Log-Seite
    """
    user_info = session.get('user_info', {})
    if not user_info.get('is_master'):
        flash("Zugriff verweigert. Nur für Master-Benutzer.", "error")
        return redirect(url_for('index'))
    
    custom_log(user_info.get('kuerzel'), "Trigger-Protokoll aufgerufen", user_info.get('color'))
    
    triggers = trigger_store.get_all_triggers()
    employers = employer_store.get_all()
    
    return render_template('trigger_log.html', triggers=triggers, employers=employers)


def _safe_int(value, default=0):
    """
    Konvertiert einen Wert sicher zu int.
    
    Args:
        value: Der zu konvertierende Wert
        default: Fallback-Wert bei Fehler
    
    Returns:
        int: Der konvertierte Wert oder default
    """
    try:
        return int(value)
    except (ValueError, TypeError):
        return default


def _parse_trigger_form(form):
    """
    Parst das Trigger-Formular und erstellt ein Trigger-Daten-Dictionary.
    
    Args:
        form: Das Flask-Request-Formular
    
    Returns:
        dict: Die geparsten Trigger-Daten
    """
    # Parse conditions from form
    conditions = []
    try:
        condition_count = int(form.get('condition_count', 1))
    except (ValueError, TypeError):
        condition_count = 1
    for i in range(condition_count):
        field = form.get(f'condition_field_{i}')
        operator = form.get(f'condition_operator_{i}')
        if field and operator:
            conditions.append({
                'field': field,
                'operator': operator,
                'from_value': form.get(f'condition_from_{i}') or None,
                'to_value': form.get(f'condition_to_{i}') or None
            })
    
    # Parse excluded employers
    excluded_employers = form.getlist('excluded_employers')
    
    # Parse action config
    action_type = form.get('action_type', 'email')
    action_config = {}
    
    if action_type == 'email':
        recipients = form.get('email_recipients', '')
        action_config = {
            'recipients': [r.strip() for r in recipients.split(',') if r.strip()],
            'subject': form.get('email_subject', ''),
            'body': form.get('email_body', ''),
            'send_individual': form.get('email_send_individual') == 'on'
        }
    elif action_type == 'api':
        # Parse headers JSON sicher
        headers_str = form.get('api_headers', '{}') or '{}'
        try:
            headers = json.loads(headers_str)
        except json.JSONDecodeError:
            headers = {}
        
        action_config = {
            'url': form.get('api_url', ''),
            'method': form.get('api_method', 'POST'),
            'headers': headers,
            'auth': {
                'type': form.get('api_auth_type', 'none'),
                'token': form.get('api_auth_token', ''),
                'username': form.get('api_auth_username', ''),
                'password': form.get('api_auth_password', ''),
                'api_key': form.get('api_auth_api_key', ''),
                'api_key_header': form.get('api_auth_api_key_header', 'X-API-Key')
            },
            'body': form.get('api_body', ''),
            'timeout_seconds': _safe_int(form.get('api_timeout', 30), 30),
            'retry_on_failure': form.get('api_retry') == 'on',
            'send_individual': form.get('api_send_individual') == 'on'
        }
    
    trigger_data = {
        'name': form.get('trigger_name', '').strip(),
        'enabled': form.get('trigger_enabled') == 'on',
        'trigger_type': 'employee',  # Currently only employee triggers supported
        'event': form.get('trigger_event', 'employee_changed'),
        'conditions': conditions if form.get('trigger_event') == 'employee_changed' else [],
        'condition_logic': form.get('condition_logic', 'AND'),
        'excluded_employers': excluded_employers,
        'action': {
            'type': action_type,
            'config': action_config
        }
    }
    
    return trigger_data


@app.route('/user/settings', methods=['GET', 'POST'])
def user_settings():
    """
    Behandelt die Benutzereinstellungen für den aktuellen Benutzer.
    
    GET: Zeigt die Benutzereinstellungsseite an
    POST: Verarbeitet Passwortänderungen
    
    Returns:
        Response: Benutzereinstellungsseite oder Weiterleitung
    """
    user_id = session.get('user_id')
    if not user_id:
        return redirect(url_for('login'))

    users = load_users()
    user = next((u for u in users if u['username'] == user_id), None)
    if not user:
        flash("Benutzer nicht gefunden.", "error")
        return redirect(url_for('index'))

    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'change_password':
            current_password = request.form.get('current_password')
            new_password = request.form.get('new_password')
            confirm_password = request.form.get('confirm_password')

            if not check_password_hash(user['password_hash'], current_password):
                flash("Das aktuelle Passwort ist nicht korrekt.", "error")
            elif new_password != confirm_password:
                flash("Die neuen Passwörter stimmen nicht überein.", "error")
            else:
                # SECURITY FIX SV-012: Passwort-Policy prüfen
                is_valid, error_msg = validate_password(new_password)
                if not is_valid:
                    flash(error_msg, "error")
                else:
                    user['password_hash'] = generate_password_hash(new_password)
                    write_users(users)
                    flash("Ihr Passwort wurde erfolgreich geändert.", "success")

            return redirect(url_for('user_settings'))

    return render_template('user_settings.html')


@app.route('/api/user/theme', methods=['POST'])
def update_theme():
    """
    API-Endpunkt zum Aktualisieren des Benutzer-Themes.
    
    Returns:
        Response: JSON-Antwort mit Status der Theme-Aktualisierung
    """
    user_id = session.get('user_id')
    if not user_id:
        return jsonify({"status": "error", "message": "Nicht angemeldet"}), 401

    data = request.get_json()
    new_theme = data.get('theme')
    if new_theme not in ['light', 'dark']:
        return jsonify({"status": "error", "message": "Ungültiges Theme"}), 400

    users = load_users()
    user = next((u for u in users if u['username'] == user_id), None)
    if user:
        user['theme'] = new_theme
        write_users(users)
        # Update session as well
        user_info = session.get('user_info', {})
        user_info['theme'] = new_theme
        session['user_info'] = user_info
        return jsonify({"status": "success", "message": "Theme aktualisiert"}), 200
    else:
        return jsonify({"status": "error", "message": "Benutzer nicht gefunden"}), 404


@app.route('/api/system/restart', methods=['POST'])
def system_restart():
    """
    API-Endpunkt zum Neustarten des Systems.
    
    Returns:
        Response: JSON-Antwort und Systemneustart
    """
    user_info = session.get('user_info', {})
    if not user_info.get('is_master'):
        return jsonify({"status": "error", "message": "Zugriff verweigert"}), 403

    custom_log(user_info.get('kuerzel'), "Server-Neustart ausgelöst", user_info.get('color'))

    # We need a way to shut down the Waitress server gracefully.
    # A simple and effective way for this setup is to exit the process.
    # The start.bat script will be modified to loop and restart the server.
    def shutdown_server():
        # Using os._exit() is a hard exit, which is suitable for this restart mechanism.
        # It bypasses finally blocks, so it's not "graceful" in a traditional sense,
        # but it's effective for forcing a restart via an external script.
        os._exit(0)

    # It's better to perform the shutdown after the response has been sent.
    # However, since the server is going down, we can't easily do that.
    # We will just send a success message and then shut down immediately.
    # A small delay could be added, but os._exit is immediate.
    # For a more robust solution, a separate thread could be used.
    # For this application's context, an immediate exit is acceptable.

    # This response might not always be delivered if the shutdown is too fast.
    # The frontend will handle this by assuming success on a 200 OK response.
    shutdown_server()
    return jsonify({"status": "success", "message": "Server wird neugestartet..."})


@app.route('/api/logs')
def api_get_logs():
    """
    API-Endpunkt zum Abrufen der System-Logs.
    
    Returns:
        Response: JSON-Antwort mit den letzten 200 Log-Einträgen
    """
    user_info = session.get('user_info', {})
    if not user_info.get('is_master'):
        return jsonify({"error": "Zugriff verweigert"}), 403

    log_file_path = os.path.abspath(os.path.join(APP_ROOT, '..', 'server.log'))

    try:
        with open(log_file_path, 'r', encoding='utf-8', errors='ignore') as f:
            # Read all lines and take the last 200
            lines = f.readlines()
            last_lines = lines[-200:]
            # Reverse the list to show newest logs first
            last_lines.reverse()
            log_content = "".join(last_lines)
            return jsonify({"logs": log_content})
    except FileNotFoundError:
        return jsonify({"logs": "Log-Datei nicht gefunden. Sie wird beim ersten Start erstellt."})
    except Exception as e:
        return jsonify({"error": f"Fehler beim Lesen der Log-Datei: {e}"}), 500


@app.before_request
def before_request_handler():
    """
    Handler, der vor jeder Anfrage ausgeführt wird.
    
    Behandelt Logging, Authentifizierung und erzwungenes Abmelden.
    
    Returns:
        Response | None: Weiterleitung zur Anmeldung oder None
    """
    # --- 1. Logging Logic ---
    # Log unauthenticated access to the login page. All other logging is handled in the specific routes.
    if 'user_id' not in session and request.endpoint == 'login':
        custom_log(None, "LOGIN", ip=request.remote_addr)

    # --- 2. Authentication and Forced Logout Logic ---
    user_info = session.get('user_info')
    if user_info:
        is_master = user_info.get('is_master', False)
        if not is_master:
            forced_logout_time = get_forced_logout_time()
            login_time = session.get('login_time')
            if forced_logout_time and login_time and login_time < forced_logout_time:
                session.clear()
                flash("Sie wurden von einem Administrator abgemeldet.", "info")
                return redirect(url_for('login'))

    if 'user_id' not in session and request.endpoint not in ['login', 'static']:
        return redirect(url_for('login'))

@app.context_processor
def inject_user():
    """
    Injiziert Benutzerinformationen in alle Templates.
    
    Returns:
        dict: Dictionary mit Benutzerinformationen
    """
    return dict(user_info=session.get('user_info'))

@app.context_processor
def inject_now():
    """
    Injiziert das aktuelle Jahr in alle Templates.
    
    Returns:
        dict: Dictionary mit dem aktuellen Jahr
    """
    return {'now': datetime.utcnow()}

# Configure paths relative to the application's root
APP_ROOT = app.root_path
app.config.update(
    EMPLOYERS_FILE=os.path.join(APP_ROOT, 'employers.json'),
    EXPORTS_DIR=os.path.join(APP_ROOT, 'exports'),
    SNAPSHOTS_DIR=os.path.join(APP_ROOT, '_snapshots'),
    HISTORY_DIR=os.path.join(APP_ROOT, '_history'),
)

# Ensure directories exist
for d in ['EXPORTS_DIR', 'SNAPSHOTS_DIR', 'HISTORY_DIR']:
    os.makedirs(app.config[d], exist_ok=True)

employer_store = EmployerStore(filepath=app.config['EMPLOYERS_FILE'])

@app.route('/')
def index():
    """
    Zeigt die Hauptseite mit der Arbeitgeber-Auswahl an.
    
    Returns:
        Response: Die Hauptseite mit allen Arbeitgebern
    """
    user_info = session.get('user_info', {})
    custom_log(user_info.get('kuerzel'), "Arbeitgeber-Auswahl aufgerufen", user_info.get('color'))
    return render_template('index.html', employers=employer_store.get_all())

# ==============================================================================
# --- SECURITY FIX SV-015: Input-Validierung für Arbeitgeber ---
# ==============================================================================
# Erlaubte Provider-Keys (Whitelist)
ALLOWED_PROVIDER_KEYS = {'personio', 'hrworks', 'sagehr'}

def validate_employer_input(form_data: dict) -> tuple:
    """
    Validiert die Eingabedaten für einen neuen Arbeitgeber (SV-015).
    
    Args:
        form_data (dict): Die Formulardaten
    
    Returns:
        tuple[bool, str]: (Gültig, Fehlermeldung wenn ungültig)
    """
    name = form_data.get('name', '').strip()
    provider_key = form_data.get('provider_key', '').strip().lower()
    access_key = form_data.get('access_key', '').strip()
    
    # Name-Validierung
    if not name:
        return False, "Der Name des Arbeitgebers ist erforderlich."
    if len(name) > 200:
        return False, "Der Name darf maximal 200 Zeichen lang sein."
    if not all(c.isalnum() or c in ' -_.,&()' for c in name):
        return False, "Der Name enthält ungültige Zeichen."
    
    # Provider-Key Whitelist
    if provider_key not in ALLOWED_PROVIDER_KEYS:
        return False, f"Ungültiger Provider. Erlaubt: {', '.join(ALLOWED_PROVIDER_KEYS)}"
    
    # Access-Key Validierung
    if not access_key:
        return False, "Der Access Key ist erforderlich."
    if len(access_key) > 500:
        return False, "Der Access Key ist zu lang."
    
    return True, ""

@app.route('/employer/add', methods=['GET', 'POST'])
def add_employer():
    """
    Behandelt das Hinzufügen neuer Arbeitgeber.
    
    GET: Zeigt das Formular zum Hinzufügen eines Arbeitgebers an
    POST: Verarbeitet die Arbeitgeberdaten und fügt sie hinzu
    
    SECURITY FIX SV-015: Input-Validierung mit Whitelist für provider_key
    
    Returns:
        Response: Formular oder Weiterleitung zur Hauptseite
    """
    if request.method == 'POST':
        # SV-015: Input-Validierung
        is_valid, error_msg = validate_employer_input(request.form)
        if not is_valid:
            flash(error_msg, "error")
            return render_template('add_employer.html')
        
        employer_data = {
            "id": str(uuid.uuid4()),
            "name": request.form.get('name', '').strip(),
            "provider_key": request.form.get('provider_key', '').strip().lower(),
            "access_key": request.form.get('access_key', '').strip(),
            "secret_key": request.form.get('secret_key', '').strip() if request.form.get('secret_key') else None,
            "address": {
                "street": request.form.get('street', '').strip(),
                "zip_code": request.form.get('zip_code', '').strip(),
                "city": request.form.get('city', '').strip()
            },
            "is_demo": request.form.get('is_demo') == 'true'
        }
        employer_store.add(employer_data)
        flash(f"Arbeitgeber '{employer_data['name']}' erfolgreich hinzugefügt.", "success")

        user_info = session.get('user_info', {})
        custom_log(user_info.get('kuerzel'), f"Arbeitgeber '{employer_data['name']}' angelegt", user_info.get('color'))
        
        # SV-016: Audit-Log für Arbeitgeber-Erstellung
        audit_log(
            user=user_info.get('username', 'unknown'),
            action="CREATE_EMPLOYER",
            target=employer_data['id'],
            details=f"name={employer_data['name']}, provider={employer_data['provider_key']}",
            ip=request.remote_addr or 'unknown'
        )

        return redirect(url_for('index'))
    return render_template('add_employer.html')

@app.route('/employer/<employer_id>')
def employer_dashboard(employer_id):
    """
    Zeigt das Dashboard für einen spezifischen Arbeitgeber an.
    
    Args:
        employer_id (str): Die ID des Arbeitgebers
    
    Returns:
        Response: Das Arbeitgeber-Dashboard oder Weiterleitung zur Hauptseite
    """
    e = employer_store.get_by_id(employer_id)
    if not e: return redirect(url_for('index'))

    user_info = session.get('user_info', {})
    custom_log(user_info.get('kuerzel'), f"'{e.get('name')}' Mitarbeiter aufgerufen", user_info.get('color'))

    sf = request.args.get('status', 'active')
    try:
        p = ProviderFactory.get_provider(e)
        ae, raw_history = p.list_employees(only_active=False)

        # Save the entire raw response (which could be a list of pages) to one file
        save_history_entry(app.config['HISTORY_DIR'], e, raw_history)

        if sf == 'active': emps = [i for i in ae if i.get('isActive')]
        elif sf == 'former': emps = [i for i in ae if not i.get('isActive')]
        else: emps = ae
    except Exception as err:
        flash(f"Fehler bei Mitarbeiterdaten: {err}", "error")
        emps = []
    return render_template('employer_dashboard.html', employer=e, employees=emps, status=sf, active_tab='employees')

@app.route('/employer/<employer_id>/employee/<employee_id>')
def employee_detail(employer_id, employee_id):
    """
    Zeigt die Detailansicht für einen spezifischen Mitarbeiter an.
    
    Args:
        employer_id (str): Die ID des Arbeitgebers
        employee_id (str): Die ID des Mitarbeiters
    
    Returns:
        Response: Die Mitarbeiter-Detailansicht oder Weiterleitung
    """
    e = employer_store.get_by_id(employer_id)
    if not e: return redirect(url_for('index'))
    try:
        p = ProviderFactory.get_provider(e)
        emp, raw_history = p.get_employee_details(employee_id)

        # Save the entire raw response (which could be a list of pages) to one file
        save_history_entry(app.config['HISTORY_DIR'], e, raw_history)

        user_info = session.get('user_info', {})
        emp_name = f"{emp.get('firstName', '')} {emp.get('lastName', '')}".strip()
        custom_log(user_info.get('kuerzel'), f"'{e.get('name')}' Mitarbeiterdetails für '{emp_name}' aufgerufen", user_info.get('color'))
    except Exception as err:
        flash(f"Fehler bei Mitarbeiterdetails: {err}", "error")
        return redirect(url_for('employer_dashboard', employer_id=employer_id))
    return render_template('employee_detail.html', employer_id=employer_id, employer=e, employee=emp)

@app.route('/employer/<employer_id>/exports')
def employer_exports(employer_id):
    """
    Zeigt die Export-Seite für einen Arbeitgeber an.
    
    Args:
        employer_id (str): Die ID des Arbeitgebers
    
    Returns:
        Response: Die Export-Seite oder Weiterleitung zur Hauptseite
    """
    e = employer_store.get_by_id(employer_id)
    if not e: return redirect(url_for('index'))

    user_info = session.get('user_info', {})
    custom_log(user_info.get('kuerzel'), f"'{e.get('name')}' Exporte aufgerufen", user_info.get('color'))

    return render_template('exports.html', employer=e, active_tab='exports')

@app.route('/employer/<employer_id>/settings', methods=['GET', 'POST'])
def employer_settings(employer_id):
    """
    Behandelt die Einstellungen für einen spezifischen Arbeitgeber.
    
    Args:
        employer_id (str): Die ID des Arbeitgebers
    
    GET: Zeigt die Arbeitgeber-Einstellungsseite an
    POST: Verarbeitet Änderungen an den Arbeitgeberdaten
    
    Returns:
        Response: Einstellungsseite oder Weiterleitung
    """
    e = employer_store.get_by_id(employer_id)
    if not e:
        return redirect(url_for('index'))

    user_info = session.get('user_info', {})
    if request.method == 'GET':
        custom_log(user_info.get('kuerzel'), f"'{e.get('name')}' Einstellungen aufgerufen", user_info.get('color'))

    if request.method == 'POST':
        # Create a dictionary with the updated data
        updated_data = {
            "address": {
                "street": request.form.get('street'),
                "zip_code": request.form.get('zip_code'),
                "city": request.form.get('city'),
                "country": request.form.get('country')
            },
            "email": request.form.get('email'),
            "phone": request.form.get('phone'),
            "fax": request.form.get('fax'),
            "comment": request.form.get('comment')
        }

        # The update method in EmployerStore now handles merging
        employer_store.update(employer_id, updated_data)
        flash("Arbeitgeberdaten erfolgreich aktualisiert.", "success")
        return redirect(url_for('employer_settings', employer_id=employer_id))

    # For GET request, ensure address key exists for template rendering
    if 'address' not in e:
        e['address'] = {}

    return render_template('employer_settings.html', employer=e, active_tab='settings')


@app.route('/employer/<employer_id>/triggers')
def employer_triggers(employer_id):
    """
    Zeigt die Trigger-Übersicht für einen spezifischen Arbeitgeber an.
    
    Zeigt alle Trigger mit ihrem Status für diesen Arbeitgeber (aktiv/ausgeschlossen)
    und ermöglicht das Ein-/Ausschließen von Triggern.
    
    Args:
        employer_id (str): Die ID des Arbeitgebers
    
    Returns:
        Response: Trigger-Übersichtsseite für den Arbeitgeber
    """
    e = employer_store.get_by_id(employer_id)
    if not e:
        return redirect(url_for('index'))
    
    user_info = session.get('user_info', {})
    custom_log(user_info.get('kuerzel'), f"'{e.get('name')}' Trigger aufgerufen", user_info.get('color'))
    
    # Hole alle Trigger und markiere welche für diesen AG ausgeschlossen sind
    all_triggers = trigger_store.get_all_triggers()
    triggers_with_status = []
    for trigger in all_triggers:
        is_excluded = employer_id in trigger.get('excluded_employers', [])
        triggers_with_status.append({
            **trigger,
            'is_excluded_for_employer': is_excluded,
            'is_active_for_employer': trigger.get('enabled', True) and not is_excluded
        })
    
    # Hole Ausführungen für diesen Arbeitgeber
    recent_executions = trigger_log_store.get_executions(
        employer_id=employer_id,
        limit=10
    )
    
    return render_template('employer_triggers.html', 
                          employer=e, 
                          triggers=triggers_with_status,
                          recent_executions=recent_executions.get('executions', []),
                          active_tab='triggers')


@app.route('/employer/<employer_id>/triggers/<trigger_id>/toggle-exclude', methods=['POST'])
def employer_trigger_toggle_exclude(employer_id, trigger_id):
    """
    Schaltet den Ausschluss eines Triggers für einen Arbeitgeber um.
    
    Args:
        employer_id (str): Die Arbeitgeber-ID
        trigger_id (str): Die Trigger-ID
    
    Returns:
        Response: JSON mit neuem Status
    """
    user_info = session.get('user_info', {})
    if not user_info.get('is_master'):
        return jsonify({'status': 'error', 'message': 'Zugriff verweigert'}), 403
    
    trigger = trigger_store.get_trigger_by_id(trigger_id)
    if not trigger:
        return jsonify({'status': 'error', 'message': 'Trigger nicht gefunden'}), 404
    
    employer = employer_store.get_by_id(employer_id)
    if not employer:
        return jsonify({'status': 'error', 'message': 'Arbeitgeber nicht gefunden'}), 404
    
    is_excluded = employer_id in trigger.get('excluded_employers', [])
    
    if is_excluded:
        trigger_store.include_employer(trigger_id, employer_id)
        new_status = False
        action_detail = "included"
    else:
        trigger_store.exclude_employer(trigger_id, employer_id)
        new_status = True
        action_detail = "excluded"
    
    audit_log(
        user=user_info.get('username', 'unknown'),
        action="TOGGLE_TRIGGER_EMPLOYER",
        target=f"{trigger_id}/{employer_id}",
        details=action_detail,
        ip=request.remote_addr or 'unknown'
    )
    
    return jsonify({
        'status': 'success',
        'is_excluded': new_status,
        'trigger_name': trigger.get('name'),
        'employer_name': employer.get('name')
    })


@app.route('/employer/<employer_id>/statistics')
def employer_statistics(employer_id):
    """
    Zeigt die Statistik-Seite für einen Arbeitgeber an.
    
    Args:
        employer_id (str): Die ID des Arbeitgebers
    
    Returns:
        Response: Die Statistik-Seite oder Weiterleitung zur Hauptseite
    """
    e = employer_store.get_by_id(employer_id)
    if not e: return redirect(url_for('index'))

    user_info = session.get('user_info', {})
    custom_log(user_info.get('kuerzel'), f"'{e.get('name')}' Statistiken aufgerufen", user_info.get('color'))

    return render_template('statistics.html', employer=e, active_tab='statistics')

@app.route('/employer/<employer_id>/snapshots')
def snapshot_comparison(employer_id):
    """
    Zeigt die Snapshot-Vergleichsseite für einen Arbeitgeber an.
    
    Args:
        employer_id (str): Die ID des Arbeitgebers
    
    Returns:
        Response: Die Snapshot-Vergleichsseite oder Weiterleitung zur Hauptseite
    """
    employer = employer_store.get_by_id(employer_id)
    if not employer:
        return redirect(url_for('index'))

    user_info = session.get('user_info', {})
    custom_log(user_info.get('kuerzel'), f"'{employer.get('name')}' Snapshots aufgerufen", user_info.get('color'))

    snapshots_dir = app.config['SNAPSHOTS_DIR']
    safe_emp_name = _get_safe_employer_name(employer['name'])
    provider_key = employer['provider_key']
    snapshot_prefix = f"{safe_emp_name}-{provider_key}-"

    dated_snapshots = []
    for filename in os.listdir(snapshots_dir):
        if filename.startswith(snapshot_prefix) and filename.endswith('.json'):
            # Exclude the 'latest' snapshot from the list
            if 'latest' in filename:
                continue

            filepath = os.path.join(snapshots_dir, filename)
            timestamp_str = filename.replace(snapshot_prefix, '').replace('.json', '')

            try:
                dt_obj = datetime.strptime(timestamp_str, '%Y%m%d-%H%M%S')
                dated_snapshots.append({
                    'filename': filename,
                    'dt': dt_obj,
                    'size': os.path.getsize(filepath)
                })
            except ValueError:
                continue # Skip files with unparseable date formats

    # Sort snapshots by datetime object, newest first
    dated_snapshots.sort(key=lambda x: x['dt'], reverse=True)

    # Prepare the list for the template
    available_snapshots = []
    for i, snap in enumerate(dated_snapshots):
        display_timestamp = snap['dt'].strftime('%d.%m.%Y %H:%M:%S')
        # Mark the newest one
        if i == 0:
            display_timestamp += " (Neuster)"

        available_snapshots.append({
            'filename': snap['filename'],
            'timestamp': display_timestamp,
            'size': snap['size']
        })

    return render_template('snapshot_comparison.html', employer=employer, snapshots=available_snapshots, active_tab='snapshots', comparison_results=None)

def _compare_snapshots(data1, data2):
    """
    Vergleicht zwei Snapshots und gibt einen detaillierten Diff zurück.
    
    Args:
        data1 (dict): Das erste Snapshot-Dictionary
        data2 (dict): Das zweite Snapshot-Dictionary
    
    Returns:
        dict: Dictionary mit hinzugefügten, entfernten und geänderten Mitarbeitern
    """
    pids1 = set(data1.keys())
    pids2 = set(data2.keys())

    added_pids = pids2 - pids1
    removed_pids = pids1 - pids2
    common_pids = pids1 & pids2

    changed_employees = []
    for pid in common_pids:
        rec1 = data1[pid]
        rec2 = data2[pid]
        if rec1.get('hash') != rec2.get('hash'):
            flat1 = rec1.get('flat', {})
            flat2 = rec2.get('flat', {})
            changes = {}
            all_keys = set(flat1.keys()) | set(flat2.keys())
            for key in all_keys:
                val1 = flat1.get(key)
                val2 = flat2.get(key)
                if val1 != val2:
                    changes[key] = {'from': val1, 'to': val2}

            vorname = rec2.get('core', {}).get('Vorname', '')
            nachname = rec2.get('core', {}).get('Name', '')
            geburtstag = rec2.get('core', {}).get('Geburtsdatum', '')

            name_parts = [f"{vorname} {nachname}".strip(), geburtstag]
            name_str = ", ".join(filter(None, name_parts))

            changed_employees.append({
                'pid': pid,
                'name': name_str,
                'changes': changes
            })

    return {
        'added': [{'pid': pid, 'name': f"{data2[pid].get('core', {}).get('Vorname', '')} {data2[pid].get('core', {}).get('Nachname', '')}".strip()} for pid in added_pids],
        'removed': [{'pid': pid, 'name': f"{data1[pid].get('core', {}).get('Vorname', '')} {data1[pid].get('core', {}).get('Nachname', '')}".strip()} for pid in removed_pids],
        'changed': changed_employees
    }

@app.route('/employer/<employer_id>/snapshots/compare', methods=['POST'])
def compare_snapshots(employer_id):
    """
    Behandelt den Vergleich von zwei Snapshots.
    
    Args:
        employer_id (str): Die ID des Arbeitgebers
    
    Returns:
        Response: Die Snapshot-Vergleichsseite mit Ergebnissen oder Weiterleitung
    """
    employer = employer_store.get_by_id(employer_id)
    if not employer:
        return redirect(url_for('index'))

    user_info = session.get('user_info', {})
    custom_log(user_info.get('kuerzel'), f"'{employer.get('name')}' Snapshots verglichen", user_info.get('color'))

    form = request.form
    file_a = form.get('snapshot1')
    file_b = form.get('snapshot2')
    direction = form.get('direction', 'forward')

    if not file_a or not file_b:
        flash("Bitte wählen Sie zwei Snapshots zum Vergleichen aus.", "error")
        return redirect(url_for('snapshot_comparison', employer_id=employer_id))

    if file_a == file_b:
        flash("Bitte wählen Sie zwei unterschiedliche Snapshots zum Vergleichen aus.", "warning")
        return redirect(url_for('snapshot_comparison', employer_id=employer_id))

    # --- Determine comparison order ---
    snapshot_prefix = f"{_get_safe_employer_name(employer['name'])}-{employer['provider_key']}-"
    try:
        ts_a = datetime.strptime(file_a.replace(snapshot_prefix, '').replace('.json', ''), '%Y%m%d-%H%M%S')
        ts_b = datetime.strptime(file_b.replace(snapshot_prefix, '').replace('.json', ''), '%Y%m%d-%H%M%S')
    except ValueError:
        flash("Die ausgewählten Snapshot-Dateinamen haben ein ungültiges Datumsformat.", "error")
        return redirect(url_for('snapshot_comparison', employer_id=employer_id))

    # file1 is always "from", file2 is always "to"
    if ts_a < ts_b:
        older_file, newer_file = file_a, file_b
    else:
        older_file, newer_file = file_b, file_a

    if direction == 'forward':
        file1, file2 = older_file, newer_file
    else: # backward
        file1, file2 = newer_file, older_file
    # --- End comparison order ---

    snapshots_dir = app.config['SNAPSHOTS_DIR']
    path1 = os.path.join(snapshots_dir, file1)
    path2 = os.path.join(snapshots_dir, file2)

    try:
        with open(path1, 'r', encoding='utf-8') as f:
            data1 = json.load(f)
        with open(path2, 'r', encoding='utf-8') as f:
            data2 = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        flash(f"Fehler beim Laden der Snapshot-Dateien: {e}", "error")
        return redirect(url_for('snapshot_comparison', employer_id=employer_id))

    comparison_results = _compare_snapshots(data1, data2)

    # Post-process changes for better readability
    for emp in comparison_results.get('changed', []):
        processed_changes = {}
        for field, values in emp.get('changes', {}).items():
            from_val_raw = values.get('from')
            to_val_raw = values.get('to')

            from_dict, to_dict = {}, {}

            try:
                from_val = json.loads(from_val_raw) if isinstance(from_val_raw, str) else from_val_raw
                if isinstance(from_val, list):
                    from_dict = {str(item.get('label', '')): item.get('value') for item in from_val if isinstance(item, dict)}
            except (json.JSONDecodeError, TypeError):
                pass # Keep from_dict empty

            try:
                to_val = json.loads(to_val_raw) if isinstance(to_val_raw, str) else to_val_raw
                if isinstance(to_val, list):
                    to_dict = {str(item.get('label', '')): item.get('value') for item in to_val if isinstance(item, dict)}
            except (json.JSONDecodeError, TypeError):
                pass # Keep to_dict empty

            if from_dict or to_dict:
                all_labels = sorted(list(set(from_dict.keys()) | set(to_dict.keys())))
                for label in all_labels:
                    v1 = from_dict.get(label)
                    v2 = to_dict.get(label)
                    if v1 != v2:
                        processed_changes[f"{field}::{label}"] = {'from': v1, 'to': v2}
            else:
                # If no dicts could be parsed, keep the original change
                processed_changes[field] = values

        emp['changes'] = processed_changes

    # --- Format title and metadata ---
    ts1 = datetime.strptime(file1.replace(snapshot_prefix, '').replace('.json', ''), '%Y%m%d-%H%M%S')
    ts2 = datetime.strptime(file2.replace(snapshot_prefix, '').replace('.json', ''), '%Y%m%d-%H%M%S')
    # Desired format: "29.8.2025 - 07:09."
    f = lambda d: d.strftime('%d.%#m.%Y - %H:%M.') if os.name != 'nt' else d.strftime('%d.%m.%Y - %H:%M.') # %#m is not available on Windows

    title = f"Vergleichsergebnis für: {f(ts2)} -VS- {f(ts1)}"
    metadata = {
        'from_ts': f(ts1),
        'to_ts': f(ts2)
    }
    # --- End format title ---

    # Re-fetch snapshot list to render the page again
    # This logic is duplicated from the snapshot_comparison route, consider refactoring in the future
    dated_snapshots = []
    for filename in os.listdir(snapshots_dir):
        if filename.startswith(snapshot_prefix) and filename.endswith('.json'):
            if 'latest' in filename: continue
            filepath = os.path.join(snapshots_dir, filename)
            timestamp_str = filename.replace(snapshot_prefix, '').replace('.json', '')
            try:
                dt_obj = datetime.strptime(timestamp_str, '%Y%m%d-%H%M%S')
                dated_snapshots.append({'filename': filename, 'dt': dt_obj, 'size': os.path.getsize(filepath)})
            except ValueError:
                continue
    dated_snapshots.sort(key=lambda x: x['dt'], reverse=True)
    available_snapshots = []
    for i, snap in enumerate(dated_snapshots):
        display_timestamp = snap['dt'].strftime('%d.%m.%Y %H:%M:%S')
        if i == 0: display_timestamp += " (Neuster)"
        available_snapshots.append({'filename': snap['filename'], 'timestamp': display_timestamp, 'size': snap['size']})


    return render_template('snapshot_comparison.html',
                             employer=employer,
                             snapshots=available_snapshots,
                             comparison_results=comparison_results,
                             comparison_title=title,
                             comparison_metadata=metadata,
                             active_tab='snapshots')

@app.route('/employer/<employer_id>/snapshots/delete_latest', methods=['POST'])
def delete_latest_snapshot(employer_id):
    """
    Löscht den neuesten Snapshot für einen Arbeitgeber.
    
    Args:
        employer_id (str): Die ID des Arbeitgebers
    
    Returns:
        Response: Weiterleitung zur Snapshot-Vergleichsseite
    """
    employer = employer_store.get_by_id(employer_id)
    if not employer:
        flash("Arbeitgeber nicht gefunden.", "error")
        return redirect(url_for('index'))

    snapshots_dir = app.config['SNAPSHOTS_DIR']
    safe_emp_name = _get_safe_employer_name(employer['name'])
    provider_key = employer['provider_key']
    latest_snapshot_filename = f"{safe_emp_name}-{provider_key}-latest.json"
    latest_snapshot_path = os.path.join(snapshots_dir, latest_snapshot_filename)

    if os.path.exists(latest_snapshot_path):
        try:
            os.remove(latest_snapshot_path)
            flash("Der neueste Snapshot wurde erfolgreich gelöscht.", "success")
        except OSError as e:
            flash(f"Fehler beim Löschen des Snapshots: {e}", "error")
    else:
        flash("Kein 'latest' Snapshot zum Löschen gefunden.", "info")

    return redirect(url_for('snapshot_comparison', employer_id=employer_id))


@app.route('/employer/<employer_id>/delete', methods=['POST'])
def delete_employer(employer_id):
    """
    Löscht einen Arbeitgeber.
    
    Args:
        employer_id (str): Die ID des zu löschenden Arbeitgebers
    
    Returns:
        Response: Weiterleitung zur Hauptseite
    """
    employer = employer_store.get_by_id(employer_id)
    if employer:
        employer_name = employer.get("name", "Unbekannt")
        employer_store.delete(employer_id)
        flash(f"Arbeitgeber '{employer_name}' erfolgreich gelöscht.", "success")

        user_info = session.get('user_info', {})
        custom_log(user_info.get('kuerzel'), f"Arbeitgeber '{employer_name}' gelöscht", user_info.get('color'))
    else:
        flash("Arbeitgeber nicht gefunden.", "error")

    return redirect(url_for('index'))

@app.route('/api/employer/<employer_id>/statistics')
def api_employer_statistics(employer_id):
    """
    API-Endpunkt zum Abrufen der Standard-Statistiken für einen Arbeitgeber.
    
    Args:
        employer_id (str): Die ID des Arbeitgebers
    
    Returns:
        Response: JSON-Antwort mit Statistiken oder Fehlermeldung
    """
    e = employer_store.get_by_id(employer_id)
    if not e: return jsonify({"error": "Employer not found"}), 404
    try:
        p = ProviderFactory.get_provider(e)
        current_employee_details, list_raw_history = p.list_employees(only_active=False)
        save_history_entry(app.config['HISTORY_DIR'], e, list_raw_history)

        # Logic to get previous snapshot data would go here
        previous_employee_details = []

        stats = calculate_statistics(current_employee_details, previous_employee_details)
        return jsonify(stats)
    except Exception as err:
        print(f"ERROR in statistics API for {employer_id}: {err}")
        return jsonify({"error": "Internal error"}), 500

@app.route('/api/employer/<employer_id>/long_term_statistics')
def api_long_term_statistics(employer_id):
    """
    API-Endpunkt zum Abrufen der Langzeit-Statistiken für einen Arbeitgeber.
    
    Args:
        employer_id (str): Die ID des Arbeitgebers
    
    Returns:
        Response: JSON-Antwort mit Langzeit-Statistiken oder Fehlermeldung
    """
    e = employer_store.get_by_id(employer_id)
    if not e:
        return jsonify({"error": "Employer not found"}), 404

    try:
        # Step 1: Get the complete employee history from all snapshots
        employee_history = _get_employee_history_from_snapshots(e, app.config['SNAPSHOTS_DIR'])

        if not employee_history:
            return jsonify({"error": "Keine kompatiblen Verlaufsdaten gefunden. Die Langzeit-Statistik wird aufgebaut, sobald neue \"Delta-SCS-Exporte\" generiert werden."}), 404

        # Step 2: Calculate the long-term statistics based on that history
        stats = calculate_long_term_statistics(employee_history)

        return jsonify(stats)
    except Exception as err:
        # It's good practice to log the full error for debugging
        error_details = traceback.format_exc()
        print(f"ERROR in long-term statistics API for {employer_id}: {err}\n{error_details}")
        return jsonify({"error": f"An internal error occurred: {err}"}), 500

@app.route('/employer/<employer_id>/export/standard')
def download_standard_export(employer_id):
    """
    Generiert und lädt einen Standard-Export für einen Arbeitgeber herunter.
    
    Args:
        employer_id (str): Die ID des Arbeitgebers
    
    Returns:
        Response: Excel-Datei-Download oder Weiterleitung
    """
    e = employer_store.get_by_id(employer_id)
    if not e: return redirect(url_for('index'))
    try:
        p = ProviderFactory.get_provider(e)
        ad, list_raw_history = p.list_employees(only_active=False)
        save_history_entry(app.config['HISTORY_DIR'], e, list_raw_history)
        fp = generate_standard_export(ad, e['name'], e['provider_key'], app.config['EXPORTS_DIR'])
        flash(f"Standard-Export erfolgreich generiert.", "success")

        # Log the download action
        user_info = session.get('user_info', {})
        filename = os.path.basename(fp)
        custom_log(user_info.get('kuerzel'), f"Download '{filename}' angefordert", user_info.get('color'))

        return send_from_directory(app.config['EXPORTS_DIR'], filename, as_attachment=True)
    except Exception as err:
        flash(f"Fehler beim Export: {err}", "error")
        return redirect(url_for('employer_exports', employer_id=employer_id))

@app.route('/api/employer/<employer_id>/export/delta_scs')
def api_delta_scs_export(employer_id):
    """
    API-Endpunkt zum Generieren eines Delta-SCS-Exports.
    
    Args:
        employer_id (str): Die ID des Arbeitgebers
    
    Returns:
        Response: JSON-Antwort mit Export-Status und Download-URL
    """
    e = employer_store.get_by_id(employer_id)
    if not e:
        return jsonify({"status": "error", "message": "Arbeitgeber nicht gefunden."}), 404

    user_info = session.get('user_info', {})
    custom_log(user_info.get('kuerzel'), f"'{e.get('name')}' Delta Export generiert", user_info.get('color'))

    try:
        p = ProviderFactory.get_provider(e)
        ad, list_raw_history = p.list_employees(only_active=False)
        save_history_entry(app.config['HISTORY_DIR'], e, list_raw_history)
        force = request.args.get('force') == '1'

        result = generate_delta_scs_export(ad, e, app.config['SNAPSHOTS_DIR'], app.config['EXPORTS_DIR'], always_write=force)
        filepath = result.get("filepath")
        diff_data = result.get("diff")

        if not filepath:
            return jsonify({
                "status": "no_changes",
                "message": "Keine neuen oder geänderten Daten gefunden."
            })

        # Post-process the diff data for better readability before sending to frontend
        for emp in diff_data.get('changed', []):
            processed_changes = {}
            for field, values in emp.get('changes', {}).items():
                # This logic is borrowed from the snapshot comparison view
                # It unpacks nested JSON strings for a clearer UI
                from_val_raw, to_val_raw = values.get('from'), values.get('to')
                try:
                    from_val = json.loads(from_val_raw) if isinstance(from_val_raw, str) else from_val_raw
                    to_val = json.loads(to_val_raw) if isinstance(to_val_raw, str) else to_val_raw
                    if isinstance(from_val, list) or isinstance(to_val, list):
                        from_dict = {str(item.get('label', '')): item.get('value') for item in (from_val or []) if isinstance(item, dict)}
                        to_dict = {str(item.get('label', '')): item.get('value') for item in (to_val or []) if isinstance(item, dict)}
                        all_labels = sorted(list(set(from_dict.keys()) | set(to_dict.keys())))
                        for label in all_labels:
                            v1, v2 = from_dict.get(label), to_dict.get(label)
                            if v1 != v2:
                                processed_changes[f"{field.split('.')[-1]}::{label}"] = {'from': v1, 'to': v2}
                    else:
                        processed_changes[field.split('.')[-1]] = values
                except (json.JSONDecodeError, TypeError):
                    processed_changes[field.split('.')[-1]] = values
            emp['changes'] = processed_changes


        return jsonify({
            "status": "success",
            "message": "Delta-SCS-Export erfolgreich generiert.",
            "download_url": url_for('download_past_export', filename=os.path.basename(filepath)),
            "diff": diff_data
        })

    except Exception as err:
        error_details = traceback.format_exc()
        print(f"Error during delta export for {employer_id}: {err}\n{error_details}")
        return jsonify({"status": "error", "message": f"Ein unerwarteter Fehler ist aufgetreten: {err}"}), 500

@app.route('/api/employer/<employer_id>/past_exports')
def api_get_past_exports(employer_id):
    """
    API-Endpunkt zum Abrufen vergangener Exporte für einen Arbeitgeber.
    
    Args:
        employer_id (str): Die ID des Arbeitgebers
    
    Returns:
        Response: JSON-Antwort mit Liste der vergangenen Exporte
    """
    employer = employer_store.get_by_id(employer_id)
    if not employer:
        return jsonify({"error": "Arbeitgeber nicht gefunden."}), 404

    exports_dir = app.config['EXPORTS_DIR']
    safe_emp_name = _get_safe_employer_name(employer['name'])
    provider_key = employer['provider_key']
    prefix = f"delta-{safe_emp_name}-{provider_key}-"

    past_exports = []
    try:
        for filename in os.listdir(exports_dir):
            if filename.startswith(prefix) and filename.endswith('.xlsx'):
                ts_part = filename.replace(prefix, '').replace('.xlsx', '')
                try:
                    dt_obj = datetime.strptime(ts_part, '%Y%m%d-%H%M%S')
                    # Format as "dd.MM.yyyy__HH:mm"
                    display_name = dt_obj.strftime('%d.%m.%Y__%H:%M')
                    past_exports.append({
                        "filename": filename,
                        "display_name": display_name,
                        "dt": dt_obj # include for sorting
                    })
                except ValueError:
                    continue # Skip files with malformed timestamps

        # Sort by datetime object, newest first
        past_exports.sort(key=lambda x: x['dt'], reverse=True)

        # Remove the temporary dt object before sending to client
        for pe in past_exports:
            del pe['dt']

        return jsonify(past_exports)
    except Exception as e:
        return jsonify({"error": f"Fehler beim Abrufen der Exporte: {e}"}), 500

@app.route('/download/past_export/<path:filename>')
def download_past_export(filename):
    """
    Stellt eine zuvor generierte Export-Datei zum Download bereit.
    
    Args:
        filename (str): Der Name der Export-Datei
    
    Returns:
        Response: Die Export-Datei zum Download
    """
    # Log the download action
    user_info = session.get('user_info', {})
    custom_log(user_info.get('kuerzel'), f"Download '{filename}' angefordert", user_info.get('color'))

    return send_from_directory(
        app.config['EXPORTS_DIR'],
        filename,
        as_attachment=True
    )

@app.route('/employer/<employer_id>/export/statistics/standard')
def export_standard_statistics(employer_id):
    """
    Exportiert die Standard-Statistiken als TXT-Datei.
    
    Args:
        employer_id (str): Die ID des Arbeitgebers
    
    Returns:
        Response: TXT-Datei-Download oder Weiterleitung
    """
    e = employer_store.get_by_id(employer_id)
    if not e: return redirect(url_for('index'))
    try:
        p = ProviderFactory.get_provider(e)
        current_employee_details, _ = p.list_employees(only_active=False)

        previous_employee_details = []
        stats = calculate_statistics(current_employee_details, previous_employee_details)

        export_content = _format_stats_for_export(stats, 'standard')

        filename = f"standard_stats_{_get_safe_employer_name(e['name'])}_{datetime.now().strftime('%Y%m%d')}.txt"
        return Response(
            export_content,
            mimetype="text/plain",
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )
    except Exception as err:
        flash(f"Fehler beim Export der Standard-Statistiken: {err}", "error")
        return redirect(url_for('employer_statistics', employer_id=employer_id))

@app.route('/employer/<employer_id>/export/statistics/longterm')
def export_longterm_statistics(employer_id):
    """
    Exportiert die Langzeit-Statistiken als TXT-Datei.
    
    Args:
        employer_id (str): Die ID des Arbeitgebers
    
    Returns:
        Response: TXT-Datei-Download oder Weiterleitung
    """
    e = employer_store.get_by_id(employer_id)
    if not e: return redirect(url_for('index'))
    try:
        employee_history = _get_employee_history_from_snapshots(e, app.config['SNAPSHOTS_DIR'])
        if not employee_history:
            flash("Keine Verlaufsdaten für den Langzeit-Export gefunden.", "warning")
            return redirect(url_for('employer_statistics', employer_id=employer_id))

        stats = calculate_long_term_statistics(employee_history)

        export_content = _format_stats_for_export(stats, 'longterm')

        filename = f"langzeit_stats_{_get_safe_employer_name(e['name'])}_{datetime.now().strftime('%Y%m%d')}.txt"
        return Response(
            export_content,
            mimetype="text/plain",
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )
    except Exception as err:
        flash(f"Fehler beim Export der Langzeit-Statistiken: {err}", "error")
        return redirect(url_for('employer_statistics', employer_id=employer_id))

@app.route('/styleguide')
def styleguide():
    """
    Rendert die Styleguide-Seite.
    
    Returns:
        Response: Die Styleguide-Seite
    """
    user_info = session.get('user_info', {})
    custom_log(user_info.get('kuerzel'), "Styleguide aufgerufen", user_info.get('color'))
    return render_template('styleguide.html')

# ==============================================================================
# --- SECURITY FIX SV-025: Health-Check Endpoints ---
# ==============================================================================
# Endpoints für Monitoring und Load-Balancer ohne Authentifizierung

@app.route('/health')
def health():
    """
    Basis Health-Check (SV-025).
    
    Gibt einen einfachen Status zurück um zu prüfen, ob der Server läuft.
    Kein Login erforderlich.
    
    Returns:
        Response: JSON mit Status "healthy"
    """
    return jsonify({"status": "healthy", "version": "1.0.0"}), 200

@app.route('/ready')
def ready():
    """
    Readiness-Check mit Dependency-Prüfung (SV-025).
    
    Prüft ob alle notwendigen Dateien und Verzeichnisse verfügbar sind.
    Kein Login erforderlich.
    
    Returns:
        Response: JSON mit detailliertem Status
    """
    checks = {
        "data_dir": os.path.isdir(DATA_DIR),
        "employers_file": os.path.exists(os.path.join(os.path.dirname(__file__), 'employers.json')),
    }
    all_ok = all(checks.values())
    return jsonify({
        "status": "ready" if all_ok else "not_ready",
        "checks": checks
    }), 200 if all_ok else 503

if __name__ == '__main__':
    port = 5001
    print(" * Serving Flask app 'acencia_hub.app'")
    print(f" * Running on http://127.0.0.1:{port}")
    print("   Press CTRL+C to quit")
    # ==============================================================================
    # --- SECURITY FIX SV-005: Debug-Modus aus Umgebungsvariable ---
    # ==============================================================================
    # Debug-Modus wird nur aktiviert wenn FLASK_DEBUG explizit auf 'true' gesetzt ist.
    # Default: False (sicher für Produktion)
    _debug_mode = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    if _debug_mode:
        print("[WARNUNG] Debug-Modus ist aktiviert! Nicht für Produktion verwenden!")
    app.run(debug=_debug_mode, port=port)
